"""Modal effective mass fractions module.

Reads EFMFACS from an OP2 and displays per-mode fractions with
cumulative sums for each direction (Tx-Rz).  Supports comparison
between two OP2 files with mode matching by number and by MEFFMASS
cosine similarity.
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
import scipy.sparse


DIRECTIONS = ['Tx', 'Ty', 'Tz', 'Rx', 'Ry', 'Rz']

# Column definitions: (id, header_text, width)
_SINGLE_COLS = [('mode', 'Mode', 60), ('freq', 'Freq (Hz)', 100)]
for _d in DIRECTIONS:
    _SINGLE_COLS.append((f'{_d}_frac', 'Frac', 75))
    _SINGLE_COLS.append((f'{_d}_sum', 'Sum', 75))

_NUM_COLS = [
    ('mode', 'Mode', 60),
    ('freq_a', 'Freq A', 90), ('freq_b', 'Freq B', 90),
    ('delta_hz', '\u0394 Hz', 80), ('delta_pct', '\u0394 %', 70),
]
for _d in DIRECTIONS:
    _NUM_COLS.append((f'd{_d}', f'\u0394{_d}', 70))

_MEFF_COLS = [
    ('mode_a', 'Mode A', 65), ('match_b', 'Match B', 65),
    ('sim', 'Similarity', 80),
    ('freq_a', 'Freq A', 90), ('freq_b', 'Freq B', 90),
    ('delta_hz', '\u0394 Hz', 80), ('delta_pct', '\u0394 %', 70),
]
for _d in DIRECTIONS:
    _MEFF_COLS.append((f'd{_d}', f'\u0394{_d}', 70))

COL_IDS = [c[0] for c in _SINGLE_COLS]


# ---------------------------------------------------------------- comparison logic

def _match_modes_by_meff(frac_a, frac_b):
    """Match modes by cosine similarity of 6-D MEFFMASS fraction vectors.

    Uses absolute value of cosine similarity to handle eigenvector sign
    flips between solver runs.  Greedy best-match (not Hungarian) so
    multiple A-modes can map to the same B-mode, which reveals mode
    disappearance or splitting.

    Returns
    -------
    best_idx : ndarray of int, shape (n_a,)
        Index into *frac_b* of the best match for each row of *frac_a*.
    best_sim : ndarray of float, shape (n_a,)
        Cosine similarity score in [0, 1] for each match.
    """
    norms_a = np.linalg.norm(frac_a, axis=1, keepdims=True)
    norms_b = np.linalg.norm(frac_b, axis=1, keepdims=True)

    # Guard zero-norm rows (modes with no effective mass)
    safe_a = np.where(norms_a == 0, 1.0, norms_a)
    safe_b = np.where(norms_b == 0, 1.0, norms_b)

    unit_a = frac_a / safe_a
    unit_b = frac_b / safe_b

    sim_matrix = np.abs(unit_a @ unit_b.T)

    best_idx = np.argmax(sim_matrix, axis=1)
    best_sim = np.max(sim_matrix, axis=1)

    # Zero-norm rows in A have no meaningful similarity
    best_sim[norms_a.ravel() == 0] = 0.0

    return best_idx, best_sim


def compare_meff_data(data_a, data_b):
    """Compare MEFFMASS data from two files.

    Parameters
    ----------
    data_a, data_b : dict
        Each has keys *modes*, *freqs*, *frac*, *cumsum*.

    Returns
    -------
    dict with keys ``'by_number'`` and ``'by_meff'``.
    """
    modes_a, freqs_a, frac_a = data_a['modes'], data_a['freqs'], data_a['frac']
    modes_b, freqs_b, frac_b = data_b['modes'], data_b['freqs'], data_b['frac']

    # --- By mode number (common modes only) ---
    idx_b = {int(m): i for i, m in enumerate(modes_b)}
    bn_mode, bn_fa, bn_fb = [], [], []
    bn_dhz, bn_dpct, bn_dfrac = [], [], []

    for i, m in enumerate(modes_a):
        m_int = int(m)
        if m_int in idx_b:
            j = idx_b[m_int]
            fa, fb = float(freqs_a[i]), float(freqs_b[j])
            dhz = fb - fa
            dpct = (dhz / fa * 100) if fa != 0 else 0.0
            bn_mode.append(m_int)
            bn_fa.append(fa)
            bn_fb.append(fb)
            bn_dhz.append(dhz)
            bn_dpct.append(dpct)
            bn_dfrac.append(frac_b[j] - frac_a[i])

    by_number = {
        'mode': bn_mode, 'freq_a': bn_fa, 'freq_b': bn_fb,
        'delta_hz': bn_dhz, 'delta_pct': bn_dpct,
        'delta_frac': np.array(bn_dfrac) if bn_dfrac else np.empty((0, 6)),
    }

    # --- By MEFF similarity ---
    best_idx, best_sim = _match_modes_by_meff(frac_a, frac_b)
    bm_dhz, bm_dpct, bm_dfrac = [], [], []
    for i, bi in enumerate(best_idx):
        fa = float(freqs_a[i])
        fb = float(freqs_b[bi])
        bm_dhz.append(fb - fa)
        bm_dpct.append((fb - fa) / fa * 100 if fa != 0 else 0.0)
        bm_dfrac.append(frac_b[bi] - frac_a[i])

    by_meff = {
        'mode_a': [int(m) for m in modes_a],
        'match_b': [int(modes_b[bi]) for bi in best_idx],
        'similarity': best_sim.tolist(),
        'freq_a': [float(f) for f in freqs_a],
        'freq_b': [float(freqs_b[bi]) for bi in best_idx],
        'delta_hz': bm_dhz, 'delta_pct': bm_dpct,
        'delta_frac': np.array(bm_dfrac) if bm_dfrac else np.empty((0, 6)),
    }

    return {'by_number': by_number, 'by_meff': by_meff}


# --------------------------------------------------------- Excel helpers

def make_meff_styles():
    """Return a dict of openpyxl style objects for MEFFMASS Excel sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        'dark_fill': PatternFill("solid", fgColor="1F4E79"),
        'mid_fill': PatternFill("solid", fgColor="2E75B6"),
        'white_bold': Font(bold=True, color="FFFFFF", size=11),
        'sub_font': Font(bold=True, color="FFFFFF", size=10),
        'center': Alignment(horizontal="center", vertical="center"),
        'right': Alignment(horizontal="right", vertical="center"),
        'cell_border': Border(bottom=Side(style='thin', color="B4C6E7")),
        'weak_font': Font(color="FF0000"),
        'num4': '0.0000',
        'num2': '0.00',
    }


def write_meff_single_sheet(ws, data, styles):
    """Write a single-file MEFFMASS fraction sheet (row-1 merged direction
    headers, row-2 sub-headers, data from row 3)."""
    from openpyxl.utils import get_column_letter

    s = styles
    modes, freqs = data['modes'], data['freqs']
    frac, cumsum = data['frac'], data['cumsum']

    # Row 1: direction group headers
    ws.cell(row=1, column=1, value="").fill = s['dark_fill']
    ws.cell(row=1, column=2, value="").fill = s['dark_fill']
    for idx, d in enumerate(DIRECTIONS):
        c1, c2 = 3 + idx * 2, 4 + idx * 2
        cell = ws.cell(row=1, column=c1, value=d)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        ws.cell(row=1, column=c2).fill = s['dark_fill']

    # Row 2: sub-headers
    sub = ['Mode', 'Freq (Hz)']
    for _ in DIRECTIONS:
        sub.extend(['Frac', 'Sum'])
    for ci, h in enumerate(sub, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data rows
    for i in range(len(modes)):
        row = i + 3
        ws.cell(row=row, column=1, value=int(modes[i])).alignment = s['right']
        c = ws.cell(row=row, column=2, value=float(freqs[i]))
        c.number_format = s['num4']
        c.alignment = s['right']
        for j in range(6):
            fc = ws.cell(row=row, column=3 + j * 2, value=float(frac[i, j]))
            fc.number_format = s['num4']
            fc.alignment = s['right']
            sc = ws.cell(row=row, column=4 + j * 2, value=float(cumsum[i, j]))
            sc.number_format = s['num4']
            sc.alignment = s['right']
        for ci in range(1, len(sub) + 1):
            ws.cell(row=row, column=ci).border = s['cell_border']

    # Column widths
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 12
    for ci in range(3, len(sub) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.freeze_panes = 'A3'


def write_comparison_number_sheet(ws, comparison, styles):
    """Write comparison-by-mode-number sheet."""
    from openpyxl.utils import get_column_letter

    s = styles
    bn = comparison['by_number']
    base = 5  # Mode, Freq A, Freq B, delta Hz, delta %

    # Row 1: direction labels over delta columns
    for ci in range(1, base + 1):
        ws.cell(row=1, column=ci, value="").fill = s['dark_fill']
    for idx, d in enumerate(DIRECTIONS):
        cell = ws.cell(row=1, column=base + 1 + idx, value=d)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']

    # Row 2: sub-headers
    sub = ['Mode', 'Freq A', 'Freq B', '\u0394 Hz', '\u0394 %']
    for d in DIRECTIONS:
        sub.append(f'\u0394{d}')
    for ci, h in enumerate(sub, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data
    n = len(bn['mode'])
    for i in range(n):
        row = i + 3
        ws.cell(row=row, column=1, value=bn['mode'][i]).alignment = s['right']
        c = ws.cell(row=row, column=2, value=bn['freq_a'][i])
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=3, value=bn['freq_b'][i])
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=4, value=bn['delta_hz'][i])
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=5, value=bn['delta_pct'][i])
        c.number_format = s['num2']
        c.alignment = s['right']
        for j in range(6):
            c = ws.cell(row=row, column=base + 1 + j,
                        value=float(bn['delta_frac'][i, j]))
            c.number_format = s['num4']
            c.alignment = s['right']
        for ci in range(1, len(sub) + 1):
            ws.cell(row=row, column=ci).border = s['cell_border']

    # Widths
    ws.column_dimensions['A'].width = 7
    for ci in range(2, len(sub) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = 'A3'


def write_comparison_meff_sheet(ws, comparison, styles):
    """Write comparison-by-MEFF-similarity sheet."""
    from openpyxl.utils import get_column_letter

    s = styles
    bm = comparison['by_meff']
    base = 7  # Mode A, Match B, Similarity, Freq A, Freq B, delta Hz, delta %

    # Row 1: direction labels over delta columns
    for ci in range(1, base + 1):
        ws.cell(row=1, column=ci, value="").fill = s['dark_fill']
    for idx, d in enumerate(DIRECTIONS):
        cell = ws.cell(row=1, column=base + 1 + idx, value=d)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']

    # Row 2: sub-headers
    sub = ['Mode A', 'Match B', 'Similarity', 'Freq A', 'Freq B',
           '\u0394 Hz', '\u0394 %']
    for d in DIRECTIONS:
        sub.append(f'\u0394{d}')
    for ci, h in enumerate(sub, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data
    n = len(bm['mode_a'])
    for i in range(n):
        row = i + 3
        sim = bm['similarity'][i]
        weak = sim < 0.5

        ws.cell(row=row, column=1, value=bm['mode_a'][i]).alignment = s['right']
        ws.cell(row=row, column=2, value=bm['match_b'][i]).alignment = s['right']
        c = ws.cell(row=row, column=3, value=sim)
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=4, value=bm['freq_a'][i])
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=5, value=bm['freq_b'][i])
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=6, value=bm['delta_hz'][i])
        c.number_format = s['num4']
        c.alignment = s['right']
        c = ws.cell(row=row, column=7, value=bm['delta_pct'][i])
        c.number_format = s['num2']
        c.alignment = s['right']
        for j in range(6):
            c = ws.cell(row=row, column=base + 1 + j,
                        value=float(bm['delta_frac'][i, j]))
            c.number_format = s['num4']
            c.alignment = s['right']
        for ci in range(1, len(sub) + 1):
            cell = ws.cell(row=row, column=ci)
            cell.border = s['cell_border']
            if weak:
                cell.font = s['weak_font']

    # Widths
    for ci in range(1, len(sub) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = 'A3'


# ---------------------------------------------------------------- GUI module

class MeffModule:
    name = "Effective Mass Fractions"

    def __init__(self, parent):
        self.frame = ttk.Frame(parent)
        self.data = None
        self.data_b = None
        self.comparison = None
        self._view_mode = 'single'
        self._build_ui()

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        toolbar = ttk.Frame(self.frame)
        toolbar.pack(fill=tk.X, padx=5, pady=(5, 0))

        self.export_btn = ttk.Button(toolbar, text="Export to Excel\u2026",
                                     command=self._export_excel)
        self.export_btn.pack(side=tk.RIGHT)

        # Comparison radio buttons (hidden until comparison loaded)
        self._radio_var = tk.StringVar(value='number')
        self._radio_frame = ttk.Frame(toolbar)
        ttk.Radiobutton(self._radio_frame, text="By Mode Number",
                        variable=self._radio_var, value='number',
                        command=self._on_radio_change).pack(side=tk.LEFT,
                                                            padx=(0, 8))
        ttk.Radiobutton(self._radio_frame, text="By MEFF Match",
                        variable=self._radio_var, value='meff',
                        command=self._on_radio_change).pack(side=tk.LEFT)
        # Not packed yet -- shown when comparison is loaded

        # Table
        container = ttk.Frame(self.frame)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.tree = ttk.Treeview(container, columns=[c[0] for c in _SINGLE_COLS],
                                 show='headings', selectmode='browse')
        vsb = ttk.Scrollbar(container, orient=tk.VERTICAL,
                            command=self.tree.yview)
        hsb = ttk.Scrollbar(container, orient=tk.HORIZONTAL,
                            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        for col_id, heading, width in _SINGLE_COLS:
            self.tree.heading(col_id, text=heading)
            self.tree.column(col_id, width=width, minwidth=50, anchor=tk.E)

        # Tag for weak matches (red text)
        self.tree.tag_configure('weak', foreground='red')

    def _configure_tree(self, col_defs):
        """Reconfigure Treeview columns dynamically."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.tree['columns'] = [c[0] for c in col_defs]
        for col_id, heading, width in col_defs:
            self.tree.heading(col_id, text=heading)
            self.tree.column(col_id, width=width, minwidth=50, anchor=tk.E)

    # -------------------------------------------------------------- load
    def load(self, op2):
        """Populate table from OP2 data."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.data = None

        if not op2.eigenvalues:
            return

        eigval_table = next(iter(op2.eigenvalues.values()))
        modes = np.array(eigval_table.mode)
        freqs = np.array(eigval_table.cycles)

        meff_frac = op2.modal_effective_mass_fraction
        if meff_frac is None:
            messagebox.showwarning(
                "No MEFFMASS Data",
                "No MEFFMASS matrices found in this OP2.\n\n"
                "Add to your Nastran case control:\n"
                "  MEFFMASS(PLOT) = ALL")
            return

        raw = meff_frac.data
        if scipy.sparse.issparse(raw):
            raw = raw.toarray()
        raw = np.asarray(raw)

        frac = raw.T  # (nmodes, 6)
        cumsum = np.cumsum(frac, axis=0)
        nmodes = min(frac.shape[0], len(modes))

        self.data = {
            'modes': modes[:nmodes],
            'freqs': freqs[:nmodes],
            'frac': frac[:nmodes],
            'cumsum': cumsum[:nmodes],
        }

        # Clear any existing comparison
        self.data_b = None
        self.comparison = None
        self._view_mode = 'single'
        self._radio_frame.pack_forget()
        self._show_single_view()

    def load_comparison(self, op2_b):
        """Load a second OP2 for comparison."""
        if self.data is None:
            return

        if not op2_b.eigenvalues:
            return

        eigval_table = next(iter(op2_b.eigenvalues.values()))
        modes = np.array(eigval_table.mode)
        freqs = np.array(eigval_table.cycles)

        meff_frac = op2_b.modal_effective_mass_fraction
        if meff_frac is None:
            messagebox.showwarning(
                "No MEFFMASS Data",
                "No MEFFMASS matrices found in comparison OP2.\n\n"
                "Add to your Nastran case control:\n"
                "  MEFFMASS(PLOT) = ALL")
            return

        raw = meff_frac.data
        if scipy.sparse.issparse(raw):
            raw = raw.toarray()
        raw = np.asarray(raw)

        frac = raw.T
        cumsum = np.cumsum(frac, axis=0)
        nmodes = min(frac.shape[0], len(modes))

        self.data_b = {
            'modes': modes[:nmodes],
            'freqs': freqs[:nmodes],
            'frac': frac[:nmodes],
            'cumsum': cumsum[:nmodes],
        }

        self.comparison = compare_meff_data(self.data, self.data_b)
        self._view_mode = 'number'
        self._radio_var.set('number')
        self._radio_frame.pack(side=tk.LEFT, padx=(8, 0))
        self._show_by_number_view()

    def clear_comparison(self):
        """Reset to single-file view."""
        self.data_b = None
        self.comparison = None
        self._view_mode = 'single'
        self._radio_frame.pack_forget()
        if self.data is not None:
            self._show_single_view()

    # ---------------------------------------------------------- view helpers
    def _on_radio_change(self):
        view = self._radio_var.get()
        if view == 'number':
            self._show_by_number_view()
        else:
            self._show_by_meff_view()

    def _show_single_view(self):
        self._view_mode = 'single'
        self._configure_tree(_SINGLE_COLS)
        if self.data is None:
            return
        modes, freqs = self.data['modes'], self.data['freqs']
        frac, cumsum = self.data['frac'], self.data['cumsum']
        for i in range(len(modes)):
            vals = [int(modes[i]), f"{freqs[i]:.4f}"]
            for j in range(6):
                vals.extend([f"{frac[i, j]:.4f}", f"{cumsum[i, j]:.4f}"])
            self.tree.insert('', tk.END, values=vals)

    def _show_by_number_view(self):
        self._view_mode = 'number'
        self._configure_tree(_NUM_COLS)
        if self.comparison is None:
            return
        bn = self.comparison['by_number']
        for i in range(len(bn['mode'])):
            vals = [
                bn['mode'][i],
                f"{bn['freq_a'][i]:.4f}", f"{bn['freq_b'][i]:.4f}",
                f"{bn['delta_hz'][i]:.4f}", f"{bn['delta_pct'][i]:.2f}",
            ]
            for j in range(6):
                vals.append(f"{bn['delta_frac'][i, j]:.4f}")
            self.tree.insert('', tk.END, values=vals)

    def _show_by_meff_view(self):
        self._view_mode = 'meff'
        self._configure_tree(_MEFF_COLS)
        if self.comparison is None:
            return
        bm = self.comparison['by_meff']
        for i in range(len(bm['mode_a'])):
            sim = bm['similarity'][i]
            vals = [
                bm['mode_a'][i], bm['match_b'][i],
                f"{sim:.4f}",
                f"{bm['freq_a'][i]:.4f}", f"{bm['freq_b'][i]:.4f}",
                f"{bm['delta_hz'][i]:.4f}", f"{bm['delta_pct'][i]:.2f}",
            ]
            for j in range(6):
                vals.append(f"{bm['delta_frac'][i, j]:.4f}")
            tags = ('weak',) if sim < 0.5 else ()
            self.tree.insert('', tk.END, values=vals, tags=tags)

    # ------------------------------------------------------------ export
    def _export_excel(self):
        if self.data is None:
            messagebox.showinfo("Nothing to export",
                                "Open an OP2 file first.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return

        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for Excel export.\n\n"
                "pip install openpyxl")
            return

        wb = Workbook()
        styles = make_meff_styles()
        ws = wb.active

        if self.comparison is not None and self.data_b is not None:
            ws.title = "File A - MEFFMASS"
            write_meff_single_sheet(ws, self.data, styles)

            ws_b = wb.create_sheet("File B - MEFFMASS")
            write_meff_single_sheet(ws_b, self.data_b, styles)

            ws_num = wb.create_sheet("Compare - Mode Number")
            write_comparison_number_sheet(ws_num, self.comparison, styles)

            ws_meff = wb.create_sheet("Compare - MEFF Match")
            write_comparison_meff_sheet(ws_meff, self.comparison, styles)
        else:
            ws.title = "Effective Mass Fractions"
            write_meff_single_sheet(ws, self.data, styles)

        try:
            wb.save(path)
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
