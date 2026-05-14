"""ASD Overlay — Acceleration Spectral Density comparison across 1–2 OP2 files.

Reads PSD acceleration results from Nastran random response OP2 files (SOL 111).
Plots ASD (g²/Hz) vs frequency on log-log axes for selected nodes.  Up to two
OP2 files can be overlaid.  RMS is shown in the legend.

Note: this module deliberately avoids importing matplotlib.pyplot so that it
does not interfere with the Agg backend used by the Miles Equation popup.
"""

import csv
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

def _sc_int(key):
    """Normalize a pyNastran result dict key to a plain integer subcase ID."""
    return int(key[0]) if isinstance(key, tuple) else int(key)


def _subcase_options(result_dict):
    """Return [(sc_id, display_label), ...] sorted by sc_id.

    Pulls SUBTITLE then LABEL from the pyNastran table (CASE CONTROL cards).
    Falls back to the bare integer string when neither is set.
    """
    seen = {}
    for key in sorted(result_dict.keys(), key=_sc_int):
        sc = _sc_int(key)
        if sc in seen:
            continue
        tbl = result_dict[key]
        sub = (getattr(tbl, "subtitle", "") or "").strip()
        lab = (getattr(tbl, "label", "") or "").strip()
        hint = sub or lab
        display = f"{sc} — {hint}" if hint else str(sc)
        seen[sc] = display
    return list(seen.items())


def _lookup_subcase(result_dict, subcase_int):
    """Fetch a result table by integer subcase ID regardless of key format."""
    if subcase_int in result_dict:
        return result_dict[subcase_int]
    for key, val in result_dict.items():
        if _sc_int(key) == subcase_int:
            return val
    return None


_NODE_COLORS = (
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728",
    "#9467bd", "#8c564b", "#e377c2", "#bcbd22",
    "#17becf", "#7f7f7f",
)
_SLOT_TAGS = ("A", "B")
_SLOT_LINES = ("-", "--")
_SC_LINES = ("-", "--", "-.", ":")  # linestyle cycle when multiple subcases selected

RESPONSE_TYPES = {
    "Acceleration": {
        "psd_attr": "accelerations",
        "rms_attr": "accelerations",
        "frf_attr": "accelerations",
        "id_attr": "node_gridtype",
        "entity_label": "Node",
        "dof_labels": ("T1 (X)", "T2 (Y)", "T3 (Z)"),
        "unit_choices": ["in/s²", "m/s²"],
        "unit_factors": {"in/s²": 386.089, "m/s²": 9.80665},
        "psd_units": "g²/Hz",
        "rms_units": "g",
        "rms_fmt": ".3g",
        "frf_units": "g/g",
        "input_label": "Input ASD",
    },
    "Displacement": {
        "psd_attr": "displacements",
        "rms_attr": "displacements",
        "frf_attr": "displacements",
        "id_attr": "node_gridtype",
        "entity_label": "Node",
        "dof_labels": ("T1 (X)", "T2 (Y)", "T3 (Z)"),
        "unit_choices": ["in", "mm", "m"],
        "unit_factors": {"in": 1.0, "mm": 0.0393701, "m": 39.3701},
        "psd_units": "in²/Hz",
        "rms_units": "in (RMS)",
        "rms_fmt": ".2e",
        "frf_units": "in/g",
        "input_label": "Input PSD",
    },
    "SPC Force": {
        "psd_attr": "spc_forces",
        "rms_attr": "spc_forces",
        "frf_attr": "spc_forces",
        "id_attr": "node_gridtype",
        "entity_label": "Node",
        "dof_labels": ("T1 (X)", "T2 (Y)", "T3 (Z)"),
        "unit_choices": ["lbf", "N"],
        "unit_factors": {"lbf": 1.0, "N": 0.224809},
        "psd_units": "lbf²/Hz",
        "rms_units": "lbf (RMS)",
        "rms_fmt": ".3g",
        "frf_units": "lbf/g",
        "input_label": "Input PSD",
    },
    "CBUSH Force": {
        "psd_attr": "cbush_force",
        "rms_attr": "cbush_force",
        "frf_attr": "cbush_force",
        "id_attr": "element",
        "entity_label": "Element",
        "dof_labels": ("F1", "F2", "F3", "M1", "M2", "M3"),
        "unit_choices": ["lbf", "N"],
        "unit_factors": {"lbf": 1.0, "N": 0.224809},
        "psd_units": "lbf²/Hz",
        "rms_units": "lbf (RMS)",
        "rms_fmt": ".3g",
        "frf_units": "lbf/g",
        "input_label": "Input PSD",
    },
}
_DARK_BG = "#2b2b2b"
_THEMES = {
    "dark": {
        "fig_bg":    "#2b2b2b",
        "plot_bg":   "#1e1e1e",
        "grid":      "#3a3a3a",
        "text":      "#c0c0c0",
        "spine":     "#505050",
        "legend_bg": "#383838",
    },
    "light": {
        "fig_bg":    "#f5f5f5",
        "plot_bg":   "white",
        "grid":      "#cccccc",
        "text":      "#222222",
        "spine":     "#888888",
        "legend_bg": "white",
    },
}


class AsdOverlayModule:
    name = "ASD Overlay"

    DOF_LABELS = ("T1 (X)", "T2 (Y)", "T3 (Z)")  # default; DOF dropdown is dynamic
    UNIT_OPTIONS = ("in/s²", "m/s²")
    UNIT_FACTORS = {"in/s²": 386.089, "m/s²": 9.80665}

    _REF_COLORS = (
        "#000000", "#7f4f24", "#5a189a",
        "#bb3e03", "#005f73", "#404040",
    )

    _GUIDE_TEXT = """\
Random Response Overlay — Quick Guide

PURPOSE
  Compare random-response PSDs from 1–2 Nastran OP2 files.
  Plots response vs frequency on log-log axes.
  The RMS value for each curve is shown in the legend.

RESPONSE TYPE  (Type: dropdown in toolbar)
  Acceleration | Displacement | SPC Force | CBUSH Force
  One global setting — both OP2 slots use the same type.
  CBUSH Force uses element IDs (not node IDs);
    DOF labels are F1/F2/F3 (forces) and M1/M2/M3 (moments).

WORKFLOW
  1. Pick Response Type in the toolbar.
  2. Open OP2 A (required).  Select subcase and units.
  3. Optional: Open OP2 B for overlay comparison.
  4. Add nodes/elements — the section appears when an OP2 is loaded.
       Paste IDs one per line.  Optional label after the ID:
         1001          1001 Tip mass          1001, Tip mass
       Or use Import to load a CSV/text file (grid_id, label columns).
  5. Check/uncheck IDs to show or hide curves.
  6. DOF dropdown: which response component to plot.

PLOT MODE
  ASD            — PSD vs frequency
  Cumulative RMS — cumulative log-log integral (FEMCI method)

VIEW MODES
  Manual | All grids cycle DOF | One grid DOF×grid |
  One grid all DOFs cycle grid | Cycle subcases
  Use Prev/Next to step through frames in multi-frame modes.

REFERENCE ASDs
  Load spec/qual envelopes.  Dotted thick lines — informational only.

PICK PEAKS
  Toggle "Pick Peaks", then click near a curve peak to annotate.
  Annotations include frequency (and value if Label = Freq + value).
  All picked peaks export to the Excel Summary sheet.

CLEAR OP2
  Each slot has a Clear button to drop a loaded file without restarting.

SAVE / OPEN SESSION
  Save Session writes a .asdsession.json capturing all loaded files,
  nodes, references, view state, and picked peaks.
  Open Session restores everything; missing files are listed in a dialog.

EXPORT EXCEL
  Summary sheet: title, environment, curves, RMS values, references,
    picked peaks.  Curve labels include the user-typed node/element name.
  ASD sheet: raw PSD data columns per curve.
  Cumulative RMS sheet: integrated curves.

UNITS  (per slot — set to match OP2 output)
  Acceleration: g, in/s², m/s²
  Displacement: in, mm, m
  SPC Force / CBUSH Force: lbf, N

RMS IN LEGEND
  First checks whether the OP2 contains a Nastran-integrated RMS table
  (matches F06 output).  Falls back to numerical integration of the
  displayed PSD curve using FEMCI log-log integration.

LINE STYLES
  Solid   — OP2 A      Dashed  — OP2 B      Dotted  — Reference ASDs
"""

    def __init__(self, parent):
        self.frame = ctk.CTkFrame(parent)

        # Slot state: {0: slot_A, 1: slot_B}
        self._op2_slots = {0: self._empty_slot(), 1: self._empty_slot()}

        # Node rows — keyed by response type
        self._nodes_by_rt = {rt: [] for rt in RESPONSE_TYPES}

        # Reference ASD rows
        self._refs = []

        self._dof_var = ctk.StringVar(value="T1 (X)")

        # Global response type (single selector for both slots)
        self._rt_global_var = ctk.StringVar(value="Acceleration")

        # Section UI cache — populated in _build_ui
        self._section_widgets = {}     # rt -> {'frame', 'rows_frame'}
        self._sections_container = None
        self._sections_placeholder = None

        # Per-slot UI widgets populated in _build_ui
        self._open_btn = [None, None]
        self._file_label = [None, None]
        self._unit_var = [ctk.StringVar(value="in/s²"), ctk.StringVar(value="in/s²")]
        self._sc_btn = [None, None]
        self._mode_var = [ctk.StringVar(value="PSD (RANDOM)"),
                          ctk.StringVar(value="PSD (RANDOM)")]
        self._frf_row = [None, None]
        self._input_asd_btn = [None, None]
        self._input_asd_label = [None, None]
        self._input_asd_db_var = [tk.StringVar(value="0"), tk.StringVar(value="0")]
        self._unit_menu = [None, None]
        self._clear_btn = [None, None]
        self._dof_menu = None

        # Per-slot analysis name (auto-fills from OP2 stem, user-editable)
        self._name_var = [ctk.StringVar(value=""), ctk.StringVar(value="")]
        self._name_user_edited = [False, False]
        self._suppress_name_trace = [False, False]
        for i in range(2):
            self._name_var[i].trace_add(
                "write", lambda *_a, idx=i: self._on_name_var_write(idx))

        self._same_input_asd_var = tk.BooleanVar(value=False)

        self._plot_theme = "light"
        self._theme_btn = None
        self._view_mode_var = ctk.StringVar(value="Manual")
        self._cycle_index = 0
        self._prev_btn = None
        self._next_btn = None
        self._cycle_label = None
        self._title_var = ctk.StringVar(value="")
        self._env_var = ctk.StringVar(value="")
        self._env_user_edited = False
        self._suppress_env_trace = False
        self._title_var.trace_add("write", self._on_title_var_write)
        self._env_var.trace_add("write", self._on_env_var_write)

        # Plot mode and y-axis scale
        self._plot_mode_var = ctk.StringVar(value="ASD")
        self._yscale_var = ctk.StringVar(value="Log")

        # Peak picking
        self._pick_peaks_mode = False
        self._pick_btn = None
        self._peak_label_style = ctk.StringVar(value="Freq only")
        self._picked_peaks = []
        self._last_drawn_curves = []
        self._mpl_cid_click = None

        # X/Y reference lines
        self._aux_lines = []

        self._build_ui()

    @staticmethod
    def _empty_slot():
        return {
            "op2": None, "path": None, "subcase": None,
            "subcases": [],
            "subcase_options": [],
            "mode": "PSD",
            "input_asd_path": None,
            "input_asd_freqs": None,
            "input_asd_g2hz": None,
            "input_asd_g2hz_raw": None,
            "input_asd_db": 0.0,
        }

    # ── UI construction ──────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(self.frame, fg_color="transparent")
        toolbar.pack(fill=tk.X, padx=5, pady=(5, 0))

        # Shared grid so both slot rows have aligned columns
        slot_grid = ctk.CTkFrame(toolbar, fg_color="transparent")
        slot_grid.pack(fill=tk.X, pady=1)
        slot_grid.grid_columnconfigure(2, weight=1)  # file label absorbs slack

        for i, tag in enumerate(_SLOT_TAGS):
            btn = ctk.CTkButton(
                slot_grid, text=f"Open OP2 {tag}…", width=120,
                command=lambda idx=i: self._open_op2(idx),
            )
            btn.grid(row=i, column=0, sticky="w", padx=(0, 4), pady=2)
            self._open_btn[i] = btn

            clr_btn = ctk.CTkButton(
                slot_grid, text="Clear", width=60,
                fg_color="#8b1a1a",
                command=lambda idx=i: self._clear_slot(idx),
            )
            clr_btn.grid(row=i, column=1, sticky="w", padx=(0, 6), pady=2)
            self._clear_btn[i] = clr_btn

            lbl = ctk.CTkLabel(slot_grid, text="(no file)", text_color="gray",
                               anchor=tk.W, width=400)
            lbl.grid(row=i, column=2, sticky="ew", padx=(0, 8))
            self._file_label[i] = lbl

            ctk.CTkLabel(slot_grid, text="Name:").grid(
                row=i, column=3, padx=(0, 2))
            ctk.CTkEntry(slot_grid, textvariable=self._name_var[i], width=130,
                         placeholder_text="Analysis name",
                         ).grid(row=i, column=4, padx=(0, 10))

            ctk.CTkLabel(slot_grid, text="Units:").grid(
                row=i, column=5, padx=(0, 2))
            unit_cfg = RESPONSE_TYPES["Acceleration"]
            umenu = ctk.CTkOptionMenu(
                slot_grid, variable=self._unit_var[i],
                values=unit_cfg["unit_choices"],
                command=lambda _v, idx=i: self._refresh_plot(),
                width=80,
            )
            umenu.grid(row=i, column=6, padx=(0, 10))
            self._unit_menu[i] = umenu

            ctk.CTkLabel(slot_grid, text="Subcase:").grid(
                row=i, column=7, padx=(0, 2))
            scbtn = ctk.CTkButton(
                slot_grid, text="(none)",
                command=lambda idx=i: self._open_subcase_picker(idx),
                width=180, anchor="w",
            )
            scbtn.grid(row=i, column=8, padx=(0, 10))
            self._sc_btn[i] = scbtn

            ctk.CTkLabel(slot_grid, text="Mode:").grid(
                row=i, column=9, padx=(0, 2))
            ctk.CTkOptionMenu(
                slot_grid, variable=self._mode_var[i],
                values=["PSD (RANDOM)", "FRF + Input ASD"],
                command=lambda _v, idx=i: self._on_mode_change(idx),
                width=140,
            ).grid(row=i, column=10, sticky="w")

            # FRF subrow — hidden until mode is switched
            frf_row = ctk.CTkFrame(toolbar, fg_color="transparent")
            self._frf_row[i] = frf_row

            if i == 1:
                ctk.CTkCheckBox(frf_row, text="Same as A",
                                variable=self._same_input_asd_var,
                                command=self._on_same_input_asd_toggle,
                                ).pack(side=tk.LEFT, padx=(0, 8))

            asd_btn = ctk.CTkButton(
                frf_row, text="Load Input ASD…", width=140,
                command=lambda idx=i: self._load_input_asd(idx),
            )
            asd_btn.pack(side=tk.LEFT, padx=(0, 6))
            self._input_asd_btn[i] = asd_btn
            asd_lbl = ctk.CTkLabel(frf_row, text="(no file)", text_color="gray",
                                   anchor=tk.W, width=180)
            asd_lbl.pack(side=tk.LEFT, padx=(0, 8))
            self._input_asd_label[i] = asd_lbl
            ctk.CTkLabel(frf_row, text="dB:").pack(side=tk.LEFT, padx=(0, 2))
            _db_entry = ctk.CTkEntry(frf_row, textvariable=self._input_asd_db_var[i], width=46)
            _db_entry.pack(side=tk.LEFT, padx=(0, 8))
            _db_entry.bind("<Return>",   lambda _e, idx=i: self._on_input_asd_db_change(idx))
            _db_entry.bind("<FocusOut>", lambda _e, idx=i: self._on_input_asd_db_change(idx))

        # ── Type / DOF row ────────────────────────────────────────────────────
        dof_row = ctk.CTkFrame(toolbar, fg_color="transparent")
        dof_row.pack(fill=tk.X, pady=(4, 0))

        ctk.CTkLabel(dof_row, text="Type:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkOptionMenu(
            dof_row, variable=self._rt_global_var,
            values=list(RESPONSE_TYPES.keys()),
            command=self._on_rt_global_change,
            width=130,
        ).pack(side=tk.LEFT, padx=(0, 16))

        ctk.CTkLabel(dof_row, text="DOF:").pack(side=tk.LEFT, padx=(0, 2))
        self._dof_menu = ctk.CTkOptionMenu(
            dof_row, variable=self._dof_var, values=list(self.DOF_LABELS),
            command=lambda _: self._refresh_plot(), width=120,
        )
        self._dof_menu.pack(side=tk.LEFT, padx=(0, 12))

        ctk.CTkButton(
            dof_row, text="Help", width=60, font=ctk.CTkFont(weight="bold"),
            command=self._show_guide,
        ).pack(side=tk.LEFT, padx=(0, 6))

        self._status_label = ctk.CTkLabel(
            dof_row, text="Open an OP2 to begin", text_color="gray")
        self._status_label.pack(side=tk.LEFT, padx=(10, 0))

        # ── Cycle / View row ─────────────────────────────────────────────────
        view_row = ctk.CTkFrame(toolbar, fg_color="transparent")
        view_row.pack(fill=tk.X, pady=(2, 0))

        ctk.CTkLabel(view_row, text="View:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkOptionMenu(
            view_row, variable=self._view_mode_var,
            values=["Manual", "All grids, cycle DOF",
                    "One grid, cycle DOF×grid", "One grid all DOFs, cycle grid",
                    "Cycle subcases"],
            command=self._on_view_mode_change,
            width=200,
        ).pack(side=tk.LEFT, padx=(0, 10))

        self._prev_btn = ctk.CTkButton(
            view_row, text="◀ Prev", width=70, state=tk.DISABLED,
            command=lambda: self._step_cycle(-1),
        )
        self._prev_btn.pack(side=tk.LEFT, padx=(0, 4))

        self._cycle_label = ctk.CTkLabel(view_row, text="", text_color="gray",
                                         width=260, anchor=tk.W)
        self._cycle_label.pack(side=tk.LEFT, padx=(0, 4))

        self._next_btn = ctk.CTkButton(
            view_row, text="Next ▶", width=70, state=tk.DISABLED,
            command=lambda: self._step_cycle(1),
        )
        self._next_btn.pack(side=tk.LEFT)

        # ── Title / Environment row ───────────────────────────────────────────
        title_row = ctk.CTkFrame(toolbar, fg_color="transparent")
        title_row.pack(fill=tk.X, pady=(2, 4))

        ctk.CTkLabel(title_row, text="Title:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkEntry(title_row, textvariable=self._title_var, width=280,
                     placeholder_text="Plot title").pack(side=tk.LEFT, padx=(0, 16))
        ctk.CTkLabel(title_row, text="Environment:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkEntry(title_row, textvariable=self._env_var, width=220,
                     placeholder_text="Auto-fills on file load").pack(side=tk.LEFT)

        # ── Annotations row ──────────────────────────────────────────────────
        annot_row = ctk.CTkFrame(toolbar, fg_color="transparent")
        annot_row.pack(fill=tk.X, pady=(0, 4))

        ctk.CTkLabel(annot_row, text="Plot:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkOptionMenu(
            annot_row, variable=self._plot_mode_var,
            values=["ASD", "Cumulative RMS"],
            command=self._on_plot_mode_change,
            width=160,
        ).pack(side=tk.LEFT, padx=(0, 14))

        ctk.CTkLabel(annot_row, text="Y-axis:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkOptionMenu(
            annot_row, variable=self._yscale_var,
            values=["Log", "Linear"],
            command=lambda _: self._refresh_plot(),
            width=80,
        ).pack(side=tk.LEFT, padx=(0, 14))

        self._pick_btn = ctk.CTkButton(
            annot_row, text="Pick Peaks", width=100,
            command=self._toggle_pick_mode,
        )
        self._pick_btn.pack(side=tk.LEFT, padx=(0, 4))

        ctk.CTkButton(annot_row, text="Clear Peaks", width=90,
                      command=self._clear_peaks,
                      ).pack(side=tk.LEFT, padx=(0, 14))

        ctk.CTkLabel(annot_row, text="Label:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkOptionMenu(annot_row, variable=self._peak_label_style,
                          values=["Freq only", "Freq + value"],
                          command=lambda _: self._refresh_plot(),
                          width=120,
                          ).pack(side=tk.LEFT)

        ctk.CTkButton(annot_row, text="X/Y Lines…", width=100,
                      command=self._aux_lines_dialog,
                      ).pack(side=tk.LEFT, padx=(14, 0))

        ctk.CTkButton(annot_row, text="Pin Vline", width=80,
                      command=self._pin_last_peak_as_vline,
                      ).pack(side=tk.LEFT, padx=(4, 0))

        ctk.CTkButton(annot_row, text="Export Excel…", width=120,
                      command=self._export_excel,
                      ).pack(side=tk.LEFT, padx=(14, 0))

        ctk.CTkButton(annot_row, text="Save Session…", width=110,
                      command=self._save_session,
                      ).pack(side=tk.LEFT, padx=(8, 0))
        ctk.CTkButton(annot_row, text="Open Session…", width=110,
                      command=self._open_session,
                      ).pack(side=tk.LEFT, padx=(4, 0))

        # ── Body: node panel + plot ───────────────────────────────────────────
        body = ctk.CTkFrame(self.frame, fg_color="transparent")
        body.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # Left: node management panel
        node_panel = ctk.CTkFrame(body, width=320)
        node_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        node_panel.pack_propagate(False)

        ctk.CTkLabel(
            node_panel, text="Nodes / Elements",
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(pady=(8, 4))

        self._sections_container = ctk.CTkScrollableFrame(
            node_panel, fg_color="transparent", label_text="")
        self._sections_container.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 5))

        # Placeholder shown when no OP2 is loaded
        self._sections_placeholder = ctk.CTkLabel(
            self._sections_container,
            text="Load an OP2 to add\nnodes/elements",
            text_color="gray", justify="center")
        self._sections_placeholder.pack(pady=20)

        # Pre-build one frame per RT — packed/unpacked by _rebuild_sections()
        for _rt, _cfg in RESPONSE_TYPES.items():
            _entity = _cfg['entity_label']
            _frame = ctk.CTkFrame(self._sections_container, fg_color="transparent")
            ctk.CTkLabel(
                _frame, text=f"{_rt} {_entity}s",
                font=ctk.CTkFont(size=12, weight="bold"),
            ).pack(anchor=tk.W, padx=4, pady=(8, 2))
            _br1 = ctk.CTkFrame(_frame, fg_color="transparent")
            _br1.pack(fill=tk.X, padx=4)
            ctk.CTkButton(_br1, text="Add…", width=60,
                          command=lambda r=_rt: self._add_nodes_dialog(r),
                          ).pack(side=tk.LEFT)
            ctk.CTkButton(_br1, text="Import", width=60,
                          command=lambda r=_rt: self._import_nodes(r),
                          ).pack(side=tk.LEFT, padx=3)
            ctk.CTkButton(_br1, text="Clear", width=55,
                          command=lambda r=_rt: self._clear_nodes(r),
                          ).pack(side=tk.LEFT)
            _br2 = ctk.CTkFrame(_frame, fg_color="transparent")
            _br2.pack(fill=tk.X, padx=4, pady=(3, 5))
            ctk.CTkButton(_br2, text="All", width=60,
                          command=lambda r=_rt: self._select_all(r, True),
                          ).pack(side=tk.LEFT)
            ctk.CTkButton(_br2, text="None", width=60,
                          command=lambda r=_rt: self._select_all(r, False),
                          ).pack(side=tk.LEFT, padx=3)
            _rows = ctk.CTkFrame(_frame, fg_color="transparent")
            _rows.pack(fill=tk.BOTH, expand=True, padx=4)
            self._section_widgets[_rt] = {'frame': _frame, 'rows_frame': _rows}

        # Reference ASDs section
        ctk.CTkLabel(
            node_panel, text="Reference ASDs",
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(pady=(6, 4))

        ref_btn_row = ctk.CTkFrame(node_panel, fg_color="transparent")
        ref_btn_row.pack(fill=tk.X, padx=6)
        ctk.CTkButton(ref_btn_row, text="Load…", width=70,
                      command=self._load_reference_asd).pack(side=tk.LEFT)
        ctk.CTkButton(ref_btn_row, text="Manual…", width=76,
                      command=self._manual_asd_dialog).pack(side=tk.LEFT, padx=3)
        ctk.CTkButton(ref_btn_row, text="Clear", width=55,
                      command=self._clear_references).pack(side=tk.LEFT, padx=3)

        self._ref_scroll = ctk.CTkScrollableFrame(
            node_panel, fg_color="transparent", label_text="", height=120)
        self._ref_scroll.pack(fill=tk.X, padx=4, pady=(2, 6))

        # Right: matplotlib plot
        plot_container = ctk.CTkFrame(body, corner_radius=0)
        plot_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Thin header for theme toggle
        plot_header = ctk.CTkFrame(plot_container, fg_color="transparent",
                                   height=28)
        plot_header.pack(side=tk.TOP, fill=tk.X)
        plot_header.pack_propagate(False)
        self._theme_btn = ctk.CTkButton(
            plot_header, text="☾ Dark", width=80,
            command=self._toggle_theme,
        )
        self._theme_btn.pack(side=tk.RIGHT, padx=4, pady=2)

        ctk.CTkButton(
            plot_header, text="Copy Figure", width=100,
            command=self._copy_figure,
        ).pack(side=tk.RIGHT, padx=(0, 4), pady=2)

        self._fig = Figure(figsize=(8, 5), dpi=100, facecolor=_DARK_BG)
        self._ax = self._fig.add_subplot(111)

        self._canvas = FigureCanvasTkAgg(self._fig, master=plot_container)
        self._canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        toolbar_tk = tk.Frame(plot_container)
        toolbar_tk.pack(side=tk.BOTTOM, fill=tk.X)
        self._mpl_toolbar = NavigationToolbar2Tk(self._canvas, toolbar_tk)
        self._mpl_toolbar.update()

        self._mpl_cid_click = self._canvas.mpl_connect(
            "button_press_event", self._on_canvas_click)

        self._draw_empty_axes()

    def _toggle_theme(self):
        self._plot_theme = "light" if self._plot_theme == "dark" else "dark"
        self._theme_btn.configure(
            text="☀ Light" if self._plot_theme == "dark" else "☾ Dark")
        self._refresh_plot()

    def _copy_figure(self):
        import io, os, tempfile, subprocess
        buf = io.BytesIO()
        try:
            self._fig.savefig(buf, format='png', dpi=200, bbox_inches='tight',
                              facecolor=self._fig.get_facecolor())
        except Exception as exc:
            messagebox.showerror("Copy Error", f"Could not render figure:\n{exc}")
            return
        buf.seek(0)
        tmp = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
                f.write(buf.getvalue())
                tmp = f.name
            if os.name == 'nt':
                ps = (
                    'Add-Type -Assembly System.Windows.Forms;'
                    '[Windows.Forms.Clipboard]::SetImage('
                    f'[System.Drawing.Image]::FromFile("{tmp}"))'
                )
                subprocess.run(['powershell', '-Command', ps], check=True)
            elif os.uname().sysname == 'Darwin':
                subprocess.run(
                    ['osascript', '-e',
                     f'set the clipboard to '
                     f'(read (POSIX file "{tmp}") as «class PNGf»)'],
                    check=True)
            else:
                subprocess.run(
                    ['xclip', '-selection', 'clipboard',
                     '-t', 'image/png', '-i', tmp],
                    check=True)
        except Exception as exc:
            messagebox.showerror("Copy Error", str(exc))
        finally:
            if tmp:
                try:
                    os.unlink(tmp)
                except Exception:
                    pass

    def _draw_empty_axes(self):
        t = _THEMES[self._plot_theme]
        ax = self._ax
        ax.clear()
        ax.set_facecolor(t["plot_bg"])
        ax.set_xlabel("Frequency (Hz)", color=t["text"])
        ax.set_ylabel("PSD", color=t["text"])
        ax.tick_params(colors=t["text"], which="both")
        for spine in ax.spines.values():
            spine.set_edgecolor(t["spine"])
        self._fig.set_facecolor(t["fig_bg"])
        ax.text(0.5, 0.5, "Load an OP2 file and add nodes to begin",
                transform=ax.transAxes,
                ha="center", va="center", color="gray", fontsize=11)
        self._canvas.draw_idle()

    # ── Guide ────────────────────────────────────────────────────────────────

    def _show_guide(self):
        try:
            from structures_tools import show_guide
        except ImportError:
            return
        show_guide(self.frame.winfo_toplevel(), "ASD Overlay Guide",
                   self._GUIDE_TEXT)

    # ── Background threading ─────────────────────────────────────────────────

    def _run_in_background(self, label, work_fn, done_fn):
        self._status_label.configure(text=label, text_color="gray")
        for btn in self._open_btn:
            btn.configure(state=tk.DISABLED)

        container = {}

        def _worker():
            try:
                container['result'] = work_fn()
            except Exception as exc:
                container['error'] = exc

        def _poll():
            if thread.is_alive():
                self.frame.after(50, _poll)
            else:
                for btn in self._open_btn:
                    btn.configure(state=tk.NORMAL)
                done_fn(container.get('result'), container.get('error'))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()
        self.frame.after(50, _poll)

    # ── OP2 loading ──────────────────────────────────────────────────────────

    def _open_op2(self, slot_idx):
        tag = _SLOT_TAGS[slot_idx]
        path = filedialog.askopenfilename(
            title=f"Open OP2 {tag}",
            filetypes=[("OP2 files", "*.op2"), ("All files", "*.*")])
        if not path:
            return

        def _work():
            from pyNastran.op2.op2 import OP2
            op2 = OP2(mode='nx', debug=False)
            op2.read_op2(path)
            return op2

        def _done(op2, error):
            if error is not None:
                messagebox.showerror("Error", f"Could not read OP2:\n{error}")
                self._status_label.configure(text="Load failed", text_color="red")
                return

            rt = self._rt_global_var.get()
            cfg = RESPONSE_TYPES[rt]
            psd_dict = getattr(op2.op2_results.psd, cfg['psd_attr'], None) or {}
            frf_dict = getattr(op2, cfg['frf_attr'], None) or {}
            if psd_dict:
                mode = "PSD"
                result_dict = psd_dict
            elif frf_dict:
                mode = "FRF"
                result_dict = frf_dict
            else:
                messagebox.showwarning(
                    f"No {rt} Data",
                    f"OP2 {tag} contains no {rt.lower()} results.\n\n"
                    "Check that the deck includes:\n"
                    f"  {'ACCELERATION' if rt == 'Acceleration' else rt.upper()}(PLOT) = ALL\n\n"
                    "For PSD output also add:\n"
                    "  RANDOM = <sid>")
                self._file_label[slot_idx].configure(
                    text="(no data)", text_color="orange")
                return

            self._op2_slots[slot_idx]['mode'] = mode
            mode_label = "PSD (RANDOM)" if mode == "PSD" else "FRF + Input ASD"
            self._mode_var[slot_idx].set(mode_label)
            if mode == "FRF":
                self._frf_row[slot_idx].pack(fill=tk.X, pady=1)
            else:
                self._frf_row[slot_idx].pack_forget()

            self._op2_slots[slot_idx]['op2'] = op2
            self._op2_slots[slot_idx]['path'] = path

            sc_pairs = _subcase_options(result_dict)
            self._op2_slots[slot_idx]['subcase_options'] = sc_pairs
            self._op2_slots[slot_idx]['subcase'] = sc_pairs[0][0]
            self._op2_slots[slot_idx]['subcases'] = [sc_pairs[0][0]]
            self._sc_btn[slot_idx].configure(text=self._sc_btn_label(slot_idx))

            stem = os.path.splitext(os.path.basename(path))[0]
            self._maybe_autofill_name(slot_idx, stem)
            if mode == "PSD":
                self._maybe_autofill_env(stem)

            self._file_label[slot_idx].configure(
                text=os.path.basename(path), text_color=("gray10", "gray90"))
            self._status_label.configure(
                text=f"OP2 {tag}: {os.path.basename(path)} "
                     f"({len(sc_pairs)} subcase{'s' if len(sc_pairs) != 1 else ''})",
                text_color=("gray10", "gray90"))
            self._update_dof_dropdown()
            self._rebuild_sections()
            self._refresh_plot()

        self._run_in_background(f"Loading OP2 {tag}…", _work, _done)

    def _sc_btn_label(self, slot_idx):
        """Button text for the subcase picker button."""
        scs = self._op2_slots[slot_idx].get('subcases', [])
        opts = self._op2_slots[slot_idx].get('subcase_options', [])
        if not scs:
            return "(none)"
        if len(scs) == 1:
            lbl = next((l for sc, l in opts if sc == scs[0]), str(scs[0]))
            return lbl[:35] + ("…" if len(lbl) > 35 else "")
        return f"{len(scs)} subcases"

    def _open_subcase_picker(self, slot_idx):
        """Toplevel checkbox list for selecting one or more subcases."""
        opts = self._op2_slots[slot_idx].get('subcase_options', [])
        if not opts:
            return
        current = set(self._op2_slots[slot_idx].get('subcases', []))
        tag = _SLOT_TAGS[slot_idx]

        dlg = ctk.CTkToplevel(self.frame)
        dlg.title(f"Subcases — Slot {tag}")
        dlg.resizable(False, False)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text=f"Select subcases for slot {tag}:",
                     anchor="w").pack(fill=tk.X, padx=12, pady=(10, 4))

        # Checkbox rows
        check_vars = []
        scroll = ctk.CTkScrollableFrame(dlg, height=min(30 * len(opts), 280))
        scroll.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
        for sc_id, sc_lbl in opts:
            var = tk.BooleanVar(value=(sc_id in current))
            check_vars.append((sc_id, var))
            ctk.CTkCheckBox(scroll, text=sc_lbl, variable=var).pack(
                anchor="w", pady=2)

        # Select all / Clear buttons
        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=12, pady=(2, 4))
        ctk.CTkButton(btn_row, text="All", width=60,
                      command=lambda: [v.set(True) for _, v in check_vars]).pack(
            side=tk.LEFT, padx=(0, 4))
        ctk.CTkButton(btn_row, text="Clear", width=60,
                      command=lambda: [v.set(False) for _, v in check_vars]).pack(
            side=tk.LEFT)

        def _ok():
            selected = [sc for sc, v in check_vars if v.get()]
            self._op2_slots[slot_idx]['subcases'] = selected
            # Keep legacy 'subcase' in sync with first selection
            self._op2_slots[slot_idx]['subcase'] = selected[0] if selected else None
            self._sc_btn[slot_idx].configure(text=self._sc_btn_label(slot_idx))
            dlg.destroy()
            self._refresh_plot()

        ok_row = ctk.CTkFrame(dlg, fg_color="transparent")
        ok_row.pack(fill=tk.X, padx=12, pady=(4, 10))
        ctk.CTkButton(ok_row, text="OK", width=80, command=_ok).pack(
            side=tk.RIGHT, padx=(4, 0))
        ctk.CTkButton(ok_row, text="Cancel", width=80,
                      command=dlg.destroy).pack(side=tk.RIGHT)

    def _on_mode_change(self, slot_idx):
        mode_label = self._mode_var[slot_idx].get()
        mode = "FRF" if mode_label == "FRF + Input ASD" else "PSD"
        self._op2_slots[slot_idx]['mode'] = mode
        if mode == "FRF":
            self._frf_row[slot_idx].pack(fill=tk.X, pady=1)
        else:
            self._frf_row[slot_idx].pack_forget()
        self._op2_slots[slot_idx]['op2'] = None
        self._op2_slots[slot_idx]['subcase'] = None
        self._op2_slots[slot_idx]['subcases'] = []
        self._op2_slots[slot_idx]['subcase_options'] = []
        self._sc_btn[slot_idx].configure(text="(none)")
        self._file_label[slot_idx].configure(text="(no file)", text_color="gray")
        self._refresh_plot()

    # ── Input ASD loading ────────────────────────────────────────────────────

    @staticmethod
    def _parse_asd_text(text_str):
        """Parse 2-column ASD text (freq, g²/Hz). Returns (freqs, g2hz) arrays or (None, None)."""
        freqs, asds = [], []
        for line in text_str.splitlines():
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('$'):
                continue
            parts = line.replace(',', ' ').split()
            if len(parts) < 2:
                continue
            try:
                freqs.append(float(parts[0]))
                asds.append(float(parts[1]))
            except ValueError:
                continue
        if len(freqs) < 2:
            return None, None
        freqs_arr = np.array(freqs)
        asds_arr = np.array(asds)
        order = np.argsort(freqs_arr)
        return freqs_arr[order], asds_arr[order]

    @staticmethod
    def _parse_asd_text_file(path):
        """Parse a 2-column ASD text file (freq, g²/Hz). Returns (freqs, asds) or (None, None)."""
        try:
            with open(path, encoding='utf-8') as f:
                text = f.read()
        except Exception as exc:
            messagebox.showerror("Load Error", str(exc))
            return None, None
        freqs, asds = AsdOverlayModule._parse_asd_text(text)
        if freqs is None:
            messagebox.showerror("Load Error", "Need at least 2 frequency points.")
            return None, None
        return freqs, asds

    def _load_input_asd(self, slot_idx):
        path = filedialog.askopenfilename(
            title="Load Input ASD",
            filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv"),
                       ("All files", "*.*")])
        if not path:
            return

        freqs_arr, asds_arr = self._parse_asd_text_file(path)
        if freqs_arr is None:
            return

        slot = self._op2_slots[slot_idx]
        slot['input_asd_path'] = path
        slot['input_asd_freqs'] = freqs_arr
        slot['input_asd_g2hz_raw'] = asds_arr
        slot['input_asd_db'] = 0.0
        self._input_asd_db_var[slot_idx].set("0")
        slot['input_asd_g2hz'] = asds_arr

        self._input_asd_label[slot_idx].configure(
            text=f"{os.path.basename(path)} "
                 f"({len(freqs_arr)} pts, "
                 f"{freqs_arr[0]:.1f}–{freqs_arr[-1]:.1f} Hz)",
            text_color=("gray10", "gray90"))
        self._maybe_autofill_env(os.path.splitext(os.path.basename(path))[0])
        self._refresh_plot()

    def _on_input_asd_db_change(self, slot_idx):
        slot = self._op2_slots[slot_idx]
        if slot['input_asd_g2hz_raw'] is None:
            return
        try:
            db = float(self._input_asd_db_var[slot_idx].get())
        except ValueError:
            db = 0.0
        slot['input_asd_db'] = db
        slot['input_asd_g2hz'] = slot['input_asd_g2hz_raw'] * 10.0 ** (db / 10.0)
        self._refresh_plot()

    def _on_same_input_asd_toggle(self):
        use_same = self._same_input_asd_var.get()
        self._input_asd_btn[1].configure(
            state=tk.DISABLED if use_same else tk.NORMAL)
        if use_same:
            a = self._op2_slots[0]
            if a['input_asd_freqs'] is None:
                self._input_asd_label[1].configure(
                    text="(slot A has no Input ASD)", text_color="orange")
            else:
                self._input_asd_label[1].configure(
                    text=f"← same as A: {os.path.basename(a['input_asd_path'])}",
                    text_color=("gray10", "gray90"))
        else:
            b = self._op2_slots[1]
            if b['input_asd_freqs'] is not None:
                self._input_asd_label[1].configure(
                    text=f"{os.path.basename(b['input_asd_path'])} "
                         f"({len(b['input_asd_freqs'])} pts)",
                    text_color=("gray10", "gray90"))
            else:
                self._input_asd_label[1].configure(
                    text="(no file)", text_color="gray")
        self._refresh_plot()

    # ── Reference ASD management ─────────────────────────────────────────────

    def _load_reference_asd(self):
        path = filedialog.askopenfilename(
            title="Load Reference ASD",
            filetypes=[("Text/CSV files", "*.txt *.csv"),
                       ("All files", "*.*")])
        if not path:
            return
        freqs, asds = self._parse_asd_text_file(path)
        if freqs is None:
            return
        name = os.path.splitext(os.path.basename(path))[0]
        self._load_reference_asd_from_data(path, freqs, asds, name, True)
        self._refresh_plot()

    def _remove_reference(self, ref):
        ref['row_frame'].destroy()
        self._refs.remove(ref)
        self._refresh_plot()

    def _clear_references(self):
        for r in self._refs:
            r['row_frame'].destroy()
        self._refs.clear()
        self._refresh_plot()

    def _commit_ref_name(self, ref):
        new_name = ref['name_var'].get().strip() or ref['name']
        if new_name == ref['name']:
            return
        ref['name'] = new_name
        ref['name_var'].set(new_name)
        self._refresh_plot()

    def _on_ref_db_change(self, ref):
        try:
            db = float(ref['db_var'].get())
        except ValueError:
            db = 0.0
        ref['db'] = db
        ref['g2hz'] = ref['g2hz_raw'] * 10.0 ** (db / 10.0)
        grms = float(np.sqrt(max(self._grms_loglog(ref['freqs'], ref['g2hz']), 0.0)))
        if ref.get('grms_label') is not None:
            ref['grms_label'].configure(text=f"{grms:.3g} g")
        self._refresh_plot()

    def _manual_asd_dialog(self):
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("Manual ASD Entry")
        dlg.geometry("420x370")
        dlg.transient(self.frame.winfo_toplevel())
        dlg.grab_set()

        top_row = ctk.CTkFrame(dlg, fg_color="transparent")
        top_row.pack(fill=tk.X, padx=12, pady=(12, 4))
        ctk.CTkLabel(top_row, text="Name:").pack(side=tk.LEFT, padx=(0, 2))
        name_var = ctk.StringVar(value="")
        ctk.CTkEntry(top_row, textvariable=name_var, width=170,
                     placeholder_text="ASD name").pack(side=tk.LEFT, padx=(0, 12))
        ctk.CTkLabel(top_row, text="dB scale:").pack(side=tk.LEFT, padx=(0, 2))
        db_var = ctk.StringVar(value="0")
        db_entry = ctk.CTkEntry(top_row, textvariable=db_var, width=60)
        db_entry.pack(side=tk.LEFT)

        ctk.CTkLabel(dlg, text="Enter freq, PSD pairs (one per line, comma or space separated):",
                     anchor=tk.W).pack(padx=12, pady=(4, 2), fill=tk.X)

        tb = ctk.CTkTextbox(dlg, wrap="none")
        tb.pack(fill=tk.BOTH, expand=True, padx=12)

        status_var = ctk.StringVar(value="Parsed: 0 points")
        status_lbl = ctk.CTkLabel(dlg, textvariable=status_var, text_color="gray",
                                  anchor=tk.W)
        status_lbl.pack(padx=12, pady=(2, 0), fill=tk.X)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=12, pady=8)
        ok_btn = ctk.CTkButton(btn_row, text="OK", state=tk.DISABLED, command=lambda: _ok())
        ok_btn.pack(side=tk.LEFT)
        ctk.CTkButton(btn_row, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=5)

        def _update_readout(*_):
            text = tb.get("1.0", "end")
            freqs, g2hz = self._parse_asd_text(text)
            try:
                db = float(db_var.get())
            except ValueError:
                db = 0.0
            if freqs is None:
                status_var.set("Parsed: 0 points — enter freq, PSD pairs")
                status_lbl.configure(text_color="gray")
                ok_btn.configure(state=tk.DISABLED)
                return
            g2hz_scaled = g2hz * 10.0 ** (db / 10.0)
            grms = float(np.sqrt(max(self._grms_loglog(freqs, g2hz_scaled), 0.0)))
            status_var.set(f"Parsed: {len(freqs)} points    g_RMS: {grms:.3g} g")
            status_lbl.configure(text_color=("gray10", "gray90"))
            ok_btn.configure(state=tk.NORMAL)

        tb.bind("<KeyRelease>", _update_readout)
        db_entry.bind("<KeyRelease>", _update_readout)

        def _ok():
            text = tb.get("1.0", "end")
            freqs, g2hz = self._parse_asd_text(text)
            if freqs is None:
                messagebox.showerror("Error", "Need at least 2 valid freq, PSD pairs.",
                                     parent=dlg)
                return
            try:
                db = float(db_var.get())
            except ValueError:
                db = 0.0
            name = name_var.get().strip() or "Manual ASD"
            self._load_reference_asd_from_data(
                path=None, freqs_raw=freqs, g2hz_raw=g2hz, name=name,
                checked=True, db=db, is_manual=True, manual_text=text.strip())
            self._refresh_plot()
            dlg.destroy()

    def _curve_style_dialog(self, curve_dict, is_ref=False):
        from tkinter import colorchooser
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("Edit Curve Style")
        dlg.geometry("400x260")
        dlg.resizable(False, False)
        dlg.transient(self.frame.winfo_toplevel())
        dlg.grab_set()

        st = curve_dict.get('style', {})

        # Label override
        row = ctk.CTkFrame(dlg, fg_color="transparent")
        row.pack(fill=tk.X, padx=14, pady=(14, 4))
        ctk.CTkLabel(row, text="Label override:", width=110, anchor=tk.W).pack(side=tk.LEFT)
        lbl_var = ctk.StringVar(value=st.get('label_override') or "")
        ctk.CTkEntry(row, textvariable=lbl_var, width=220,
                     placeholder_text="blank = auto").pack(side=tk.LEFT)

        # Color
        color_row = ctk.CTkFrame(dlg, fg_color="transparent")
        color_row.pack(fill=tk.X, padx=14, pady=4)
        ctk.CTkLabel(color_row, text="Color:", width=110, anchor=tk.W).pack(side=tk.LEFT)
        color_var = tk.StringVar(value=st.get('color') or "")
        color_entry = ctk.CTkEntry(color_row, textvariable=color_var, width=90)
        color_entry.pack(side=tk.LEFT, padx=(0, 6))
        swatch = ctk.CTkFrame(color_row, width=22, height=22,
                              fg_color=st.get('color') or "gray50")
        swatch.pack(side=tk.LEFT, padx=(0, 6))
        swatch.pack_propagate(False)

        def _pick_color():
            init = color_var.get() or None
            result = colorchooser.askcolor(color=init, title="Pick Color", parent=dlg)
            if result and result[1]:
                color_var.set(result[1])
                swatch.configure(fg_color=result[1])

        color_entry.bind("<FocusOut>", lambda _e: swatch.configure(
            fg_color=color_var.get() if color_var.get() else "gray50"))
        ctk.CTkButton(color_row, text="Pick…", width=60,
                      command=_pick_color).pack(side=tk.LEFT)

        # Linestyle
        ls_row = ctk.CTkFrame(dlg, fg_color="transparent")
        ls_row.pack(fill=tk.X, padx=14, pady=4)
        ctk.CTkLabel(ls_row, text="Linestyle:", width=110, anchor=tk.W).pack(side=tk.LEFT)
        _ls_map = {"Auto": None, "Solid": "-", "Dashed": "--", "Dotted": ":", "Dash-dot": "-."}
        _ls_rev = {v: k for k, v in _ls_map.items()}
        cur_ls_name = _ls_rev.get(st.get('linestyle'), "Auto")
        ls_var = ctk.StringVar(value=cur_ls_name)
        ctk.CTkOptionMenu(ls_row, variable=ls_var, values=list(_ls_map.keys()),
                          width=130).pack(side=tk.LEFT)

        # Linewidth
        lw_row = ctk.CTkFrame(dlg, fg_color="transparent")
        lw_row.pack(fill=tk.X, padx=14, pady=4)
        ctk.CTkLabel(lw_row, text="Linewidth:", width=110, anchor=tk.W).pack(side=tk.LEFT)
        lw_var = ctk.StringVar(value=str(st.get('linewidth') or ""))
        ctk.CTkEntry(lw_row, textvariable=lw_var, width=60,
                     placeholder_text="auto").pack(side=tk.LEFT)

        # Buttons
        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=14, pady=(8, 14))

        def _reset():
            curve_dict['style'] = {}
            self._refresh_plot()
            dlg.destroy()

        def _ok():
            new_st = {}
            lbl = lbl_var.get().strip()
            if lbl:
                new_st['label_override'] = lbl
            col = color_var.get().strip()
            if col:
                new_st['color'] = col
            ls_name = ls_var.get()
            ls_val = _ls_map.get(ls_name)
            if ls_val is not None:
                new_st['linestyle'] = ls_val
            try:
                lw = float(lw_var.get())
                new_st['linewidth'] = lw
            except ValueError:
                pass
            curve_dict['style'] = new_st
            self._refresh_plot()
            dlg.destroy()

        ctk.CTkButton(btn_row, text="Reset to defaults", width=130,
                      command=_reset).pack(side=tk.LEFT)
        ctk.CTkButton(btn_row, text="OK", width=70, command=_ok).pack(side=tk.LEFT, padx=(8, 4))
        ctk.CTkButton(btn_row, text="Cancel", width=70,
                      command=dlg.destroy).pack(side=tk.LEFT)

    def _aux_lines_dialog(self):
        import copy
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("X/Y Reference Lines")
        dlg.geometry("660x380")
        dlg.transient(self.frame.winfo_toplevel())
        dlg.grab_set()

        working = copy.deepcopy(self._aux_lines)
        row_widgets = []

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill=tk.BOTH, expand=True, padx=8, pady=(8, 0))

        hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        hdr.pack(fill=tk.X, pady=(0, 2))
        for txt, w in [("Vis", 28), ("Axis", 50), ("Value", 80), ("Label", 170),
                       ("Color", 80), ("Style", 80), ("LW", 50), ("", 28)]:
            ctk.CTkLabel(hdr, text=txt, width=w, anchor=tk.W,
                         font=ctk.CTkFont(weight="bold")).pack(side=tk.LEFT, padx=2)

        def _make_row(ln):
            from tkinter import colorchooser as cc
            rf = ctk.CTkFrame(scroll, fg_color="transparent")
            rf.pack(fill=tk.X, pady=1)

            vis_var = tk.BooleanVar(value=ln.get('visible', True))
            axis_var = ctk.StringVar(value=ln.get('axis', 'x').upper())
            val_var = ctk.StringVar(value=str(ln.get('value', 0.0)))
            lbl_var = ctk.StringVar(value=ln.get('label', ''))
            col_var = ctk.StringVar(value=ln.get('color', '#888888'))
            ls_var = ctk.StringVar(value=ln.get('linestyle', '--'))
            lw_var = ctk.StringVar(value=str(ln.get('linewidth', 1.0)))

            ctk.CTkCheckBox(rf, text="", variable=vis_var, width=28).pack(side=tk.LEFT, padx=2)
            ctk.CTkOptionMenu(rf, variable=axis_var, values=["X", "Y"],
                              width=50).pack(side=tk.LEFT, padx=2)
            ctk.CTkEntry(rf, textvariable=val_var, width=80).pack(side=tk.LEFT, padx=2)
            ctk.CTkEntry(rf, textvariable=lbl_var, width=170).pack(side=tk.LEFT, padx=2)

            swatch = ctk.CTkButton(rf, text="", width=30, height=22,
                                   fg_color=ln.get('color', '#888888'),
                                   hover_color=ln.get('color', '#888888'),
                                   command=lambda cv=col_var, sw=None: None)

            def _pick(cv=col_var, sw_ref=[None]):
                result = cc.askcolor(color=cv.get(), title="Pick Color", parent=dlg)
                if result and result[1]:
                    cv.set(result[1])
                    swatch.configure(fg_color=result[1], hover_color=result[1])

            swatch.configure(command=_pick)
            swatch.pack(side=tk.LEFT, padx=2)

            ctk.CTkOptionMenu(rf, variable=ls_var,
                              values=["-", "--", ":", "-."],
                              width=80).pack(side=tk.LEFT, padx=2)
            ctk.CTkEntry(rf, textvariable=lw_var, width=50).pack(side=tk.LEFT, padx=2)

            def _del(r=rf):
                r.destroy()
                for item in row_widgets:
                    if item['frame'] is r:
                        row_widgets.remove(item)
                        break

            ctk.CTkButton(rf, text="✕", width=28,
                          fg_color="transparent", hover_color=("gray75", "gray30"),
                          text_color=("gray40", "gray60"),
                          command=_del).pack(side=tk.LEFT, padx=2)

            row_widgets.append({
                'frame': rf, 'vis': vis_var, 'axis': axis_var,
                'value': val_var, 'label': lbl_var, 'color': col_var,
                'linestyle': ls_var, 'linewidth': lw_var,
            })

        for ln in working:
            _make_row(ln)

        def _add():
            _make_row({"axis": "x", "value": 100.0, "label": "",
                       "color": "#888888", "linestyle": "--",
                       "linewidth": 1.0, "visible": True})

        add_btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        add_btn_row.pack(fill=tk.X, padx=8, pady=(4, 0))
        ctk.CTkButton(add_btn_row, text="+ Add Line", width=100,
                      command=_add).pack(side=tk.LEFT)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=8, pady=(4, 10))

        def _ok():
            new_lines = []
            for w in row_widgets:
                try:
                    val = float(w['value'].get())
                except ValueError:
                    continue
                try:
                    lw = float(w['linewidth'].get())
                except ValueError:
                    lw = 1.0
                new_lines.append({
                    "axis": w['axis'].get().lower(),
                    "value": val,
                    "label": w['label'].get(),
                    "color": w['color'].get() or "#888888",
                    "linestyle": w['linestyle'].get(),
                    "linewidth": lw,
                    "visible": w['vis'].get(),
                })
            self._aux_lines = new_lines
            self._refresh_plot()
            dlg.destroy()

        ctk.CTkButton(btn_row, text="OK", width=80, command=_ok).pack(side=tk.LEFT)
        ctk.CTkButton(btn_row, text="Cancel", width=80,
                      command=dlg.destroy).pack(side=tk.LEFT, padx=5)

    def _pin_last_peak_as_vline(self):
        if not self._picked_peaks:
            messagebox.showwarning("No Peaks", "Pick a peak first using Pick Peaks mode.")
            return
        pk = self._picked_peaks[-1]
        freq = pk['freq']
        self._aux_lines.append({
            "axis": "x", "value": freq,
            "label": f"{freq:.1f} Hz",
            "color": "#888888", "linestyle": "--",
            "linewidth": 1.0, "visible": True,
        })
        self._refresh_plot()

    # ── Node management ──────────────────────────────────────────────────────

    def _add_nodes_dialog(self, rt=None):
        if rt is None:
            rt = self._rt_global_var.get()
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("Add Nodes")
        dlg.geometry("380x310")
        dlg.transient(self.frame.winfo_toplevel())
        dlg.grab_set()

        ctk.CTkLabel(
            dlg,
            text="Enter one node/element per line.  Optional label:\n"
                 "  1001        1001 Tip mass        1001, Tip mass",
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(padx=12, pady=(12, 4), fill=tk.X)

        tb = ctk.CTkTextbox(dlg, wrap="none")
        tb.pack(fill=tk.BOTH, expand=True, padx=12)

        def _ok():
            self._parse_and_add_nodes(tb.get("1.0", "end"), rt)
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=12, pady=8)
        ctk.CTkButton(btn_row, text="Add", command=_ok).pack(side=tk.LEFT)
        ctk.CTkButton(btn_row, text="Cancel",
                      command=dlg.destroy).pack(side=tk.LEFT, padx=5)
        dlg.bind("<Return>", lambda _: _ok())

    def _parse_and_add_nodes(self, text, rt=None):
        if rt is None:
            rt = self._rt_global_var.get()
        existing_ids = {n['id'] for n in self._nodes_by_rt.get(rt, [])}
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
        entity = cfg['entity_label']
        added = False
        for raw in text.splitlines():
            line = raw.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(',', 1)] if ',' in line \
                else line.split(None, 1)
            try:
                gid = int(parts[0])
            except (ValueError, IndexError):
                continue
            if gid in existing_ids:
                continue
            existing_ids.add(gid)
            label = parts[1].strip() if len(parts) > 1 and parts[1].strip() \
                else f"{entity} {gid}"
            self._add_node_row(gid, label, rt)
            added = True

        if added:
            self._refresh_plot()

    def _import_nodes(self, rt=None):
        if rt is None:
            rt = self._rt_global_var.get()
        path = filedialog.askopenfilename(
            title="Import Nodes",
            filetypes=[("Text files", "*.txt"),
                       ("CSV files", "*.csv"),
                       ("All files", "*.*")])
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.csv':
                text = self._read_csv_as_text(path)
            else:
                text = self._read_text_node_file(path)
        except Exception as exc:
            messagebox.showerror("Import Error", str(exc))
            return
        self._parse_and_add_nodes(text, rt)

    @staticmethod
    def _read_csv_as_text(path):
        """Parse a CSV node file; return 'gid,label' lines joined by newline."""
        with open(path, newline='', encoding='utf-8-sig') as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
                has_header = csv.Sniffer().has_header(sample)
            except csv.Error:
                dialect = csv.excel
                has_header = False
            reader = csv.reader(f, dialect)
            if has_header:
                next(reader, None)
            lines = []
            for row in reader:
                if not row:
                    continue
                try:
                    gid = int(str(row[0]).strip())
                except (ValueError, IndexError):
                    continue
                label = row[1].strip() if len(row) > 1 and str(row[1]).strip() \
                    else ""
                lines.append(f"{gid},{label}" if label else str(gid))
        return "\n".join(lines)

    @staticmethod
    def _read_text_node_file(path):
        """Parse a loose-format text node file; return 'gid,label' lines."""
        lines = []
        with open(path, encoding='utf-8') as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith('#') or line.startswith('$'):
                    continue
                parts = [p.strip() for p in line.split(',', 1)] if ',' in line \
                    else line.split(None, 1)
                try:
                    gid = int(parts[0])
                except (ValueError, IndexError):
                    continue
                label = parts[1].strip() if len(parts) > 1 and parts[1].strip() \
                    else ""
                lines.append(f"{gid},{label}" if label else str(gid))
        return "\n".join(lines)

    def _add_node_row(self, gid, label, rt=None):
        if rt is None:
            rt = self._rt_global_var.get()
        var = tk.BooleanVar(value=True)
        gid_var = tk.StringVar(value=str(gid))
        label_var = tk.StringVar(value=label)

        rows_frame = self._section_widgets[rt]['rows_frame']
        row_frame = ctk.CTkFrame(rows_frame, fg_color="transparent")
        row_frame.pack(fill=tk.X, pady=1)

        node = {"id": gid, "label": label, "rt": rt,
                "checked": var, "row_frame": row_frame,
                "gid_var": gid_var, "label_var": label_var,
                "style": {}}
        self._nodes_by_rt[rt].append(node)

        ctk.CTkCheckBox(row_frame, text="", variable=var, width=24,
                        command=self._refresh_plot).pack(side=tk.LEFT, padx=(2, 0))

        gid_entry = ctk.CTkEntry(row_frame, textvariable=gid_var, width=68)
        gid_entry.pack(side=tk.LEFT, padx=(2, 2))
        gid_entry.bind("<Return>",   lambda _e, n=node: self._commit_node_gid(n))
        gid_entry.bind("<FocusOut>", lambda _e, n=node: self._commit_node_gid(n))

        lbl_entry = ctk.CTkEntry(row_frame, textvariable=label_var, width=120)
        lbl_entry.pack(side=tk.LEFT, padx=(0, 2))
        lbl_entry.bind("<Return>",   lambda _e, n=node: self._commit_node_label(n))
        lbl_entry.bind("<FocusOut>", lambda _e, n=node: self._commit_node_label(n))

        ctk.CTkButton(row_frame, text="Edit…", width=48,
                      command=lambda n=node: self._curve_style_dialog(n, is_ref=False),
                      ).pack(side=tk.LEFT, padx=(0, 2))

        ctk.CTkButton(row_frame, text="✕", width=22,
                      command=lambda n=node: self._remove_node(n),
                      fg_color="transparent", hover_color=("gray75", "gray30"),
                      text_color=("gray40", "gray60"),
                      ).pack(side=tk.LEFT)

    def _clear_nodes(self, rt=None):
        if rt is None:
            rt = self._rt_global_var.get()
        for n in self._nodes_by_rt.get(rt, []):
            n['row_frame'].destroy()
        self._nodes_by_rt[rt] = []
        self._refresh_plot()

    def _select_all(self, rt=None, state=True):
        if rt is None:
            rt = self._rt_global_var.get()
        for n in self._nodes_by_rt.get(rt, []):
            n['checked'].set(state)
        self._refresh_plot()

    def _remove_node(self, node):
        node['row_frame'].destroy()
        rt = node.get('rt', self._rt_global_var.get())
        if node in self._nodes_by_rt.get(rt, []):
            self._nodes_by_rt[rt].remove(node)
        self._refresh_plot()

    def _commit_node_gid(self, node):
        raw = node['gid_var'].get().strip()
        try:
            new_gid = int(raw)
        except ValueError:
            node['gid_var'].set(str(node['id']))
            return
        if new_gid == node['id']:
            return
        rt = node.get('rt', self._rt_global_var.get())
        if any(n is not node and n['id'] == new_gid for n in self._nodes_by_rt.get(rt, [])):
            node['gid_var'].set(str(node['id']))
            return
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
        auto_lbl = f"{cfg['entity_label']} {node['id']}"
        if node['label'] == auto_lbl:
            node['label'] = f"{cfg['entity_label']} {new_gid}"
            node['label_var'].set(node['label'])
        node['id'] = new_gid
        self._refresh_plot()

    def _commit_node_label(self, node):
        new_label = node['label_var'].get().strip() or f"Node {node['id']}"
        if new_label == node['label']:
            return
        node['label'] = new_label
        node['label_var'].set(new_label)
        self._refresh_plot()

    # ── Title / Environment / Name helpers ───────────────────────────────────

    def _on_title_var_write(self, *_):
        self._refresh_plot()

    def _on_env_var_write(self, *_):
        if self._suppress_env_trace:
            return
        self._env_user_edited = True
        self._refresh_plot()

    def _maybe_autofill_env(self, name):
        if self._env_user_edited or not name:
            return
        self._suppress_env_trace = True
        try:
            self._env_var.set(name)
        finally:
            self._suppress_env_trace = False
        self._refresh_plot()

    def _on_name_var_write(self, slot_idx):
        if self._suppress_name_trace[slot_idx]:
            return
        self._name_user_edited[slot_idx] = True
        self._refresh_plot()

    def _maybe_autofill_name(self, slot_idx, name):
        if self._name_user_edited[slot_idx] or not name:
            return
        self._suppress_name_trace[slot_idx] = True
        try:
            self._name_var[slot_idx].set(name)
        finally:
            self._suppress_name_trace[slot_idx] = False

    # ── RMS helpers ──────────────────────────────────────────────────────────

    @staticmethod
    def _grms_loglog(freqs, asd):
        """Area under an ASD curve using analytical log-log segment integration (FEMCI)."""
        area = 0.0
        for i in range(len(freqs) - 1):
            fl, fh = float(freqs[i]), float(freqs[i + 1])
            al, ah = float(asd[i]), float(asd[i + 1])
            if fl <= 0 or fh <= 0 or al <= 0 or ah <= 0:
                continue
            log_f = np.log(fh / fl)
            b = np.log(ah / al) / log_f if log_f != 0 else 0.0
            if abs(b + 1.0) < 1e-6:
                area += al * fl * log_f
            else:
                area += (ah * fh - al * fl) / (b + 1.0)
        return area

    @staticmethod
    def _cumulative_grms_loglog(freqs, asd):
        """Cumulative RMS array using FEMCI log-log integration. cum[0] = 0."""
        cum_area = np.zeros(len(freqs))
        running = 0.0
        for i in range(len(freqs) - 1):
            fl, fh = float(freqs[i]), float(freqs[i + 1])
            al, ah = float(asd[i]), float(asd[i + 1])
            if fl <= 0 or fh <= 0 or al <= 0 or ah <= 0:
                cum_area[i + 1] = running
                continue
            log_f = np.log(fh / fl)
            b = np.log(ah / al) / log_f if log_f != 0 else 0.0
            if abs(b + 1.0) < 1e-6:
                running += al * fl * log_f
            else:
                running += (ah * fh - al * fl) / (b + 1.0)
            cum_area[i + 1] = running
        return np.sqrt(np.maximum(cum_area, 0.0))

    @staticmethod
    def _interp_input_asd_to_grid(slot, op2_freqs):
        """Log-log interpolate slot's input ASD onto op2_freqs. Out-of-range → 0."""
        f_in = slot['input_asd_freqs']
        a_in = slot['input_asd_g2hz']
        result = np.zeros(len(op2_freqs))
        for i, f in enumerate(op2_freqs):
            if f < f_in[0] or f > f_in[-1]:
                result[i] = 0.0
                continue
            idx = int(np.searchsorted(f_in, f, side='right')) - 1
            idx = min(idx, len(f_in) - 2)
            fl, fh = f_in[idx], f_in[idx + 1]
            al, ah = a_in[idx], a_in[idx + 1]
            if fl <= 0 or fh <= 0 or al <= 0 or ah <= 0 or fl == fh:
                result[i] = al
            else:
                b = np.log(ah / al) / np.log(fh / fl)
                result[i] = al * (f / fl) ** b
        return result

    def _get_response_psd(self, slot_idx, subcase, nid, idof, unit_factor):
        """Return (freqs, data, is_psd) for one node/DOF."""
        slot = self._op2_slots[slot_idx]
        op2 = slot['op2']
        rt = self._rt_global_var.get()
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])

        id_attr = cfg.get('id_attr', 'node_gridtype')

        def _entity_ids(tbl):
            arr = getattr(tbl, id_attr)
            return arr[:, 0] if id_attr == 'node_gridtype' else arr

        if slot['mode'] == "PSD":
            psd_dict = getattr(op2.op2_results.psd, cfg['psd_attr'], None) or {}
            psd_tbl = _lookup_subcase(psd_dict, subcase)
            if psd_tbl is None:
                return None, None, None
            freqs = psd_tbl._times
            hits = np.where(_entity_ids(psd_tbl) == nid)[0]
            if not len(hits):
                return None, None, None
            raw_psd = psd_tbl.data[:, hits[0], idof]
            return freqs, raw_psd / (unit_factor ** 2), True
        else:  # FRF
            frf_dict = getattr(op2, cfg['frf_attr'], None) or {}
            if not frf_dict:
                return None, None, None
            frf_tbl = _lookup_subcase(frf_dict, subcase)
            if frf_tbl is None:
                return None, None, None
            freqs = frf_tbl._times
            hits = np.where(_entity_ids(frf_tbl) == nid)[0]
            if not len(hits):
                return None, None, None
            H_native = frf_tbl.data[:, hits[0], idof]
            H_g = H_native / unit_factor

            if slot_idx == 1 and self._same_input_asd_var.get():
                slot_asd = self._op2_slots[0]
            else:
                slot_asd = slot

            if slot_asd['input_asd_freqs'] is not None:
                H_mag2 = H_g.real ** 2 + H_g.imag ** 2
                S_in = self._interp_input_asd_to_grid(slot_asd, freqs)
                return freqs, H_mag2 * S_in, True
            else:
                return freqs, np.abs(H_g), False

    @staticmethod
    def _get_rms_scalar(op2, subcase, nid, idof, freqs, psd_curve, unit_factor, rt='Acceleration'):
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
        id_attr = cfg.get('id_attr', 'node_gridtype')
        try:
            rms_dict = getattr(op2.op2_results.rms, cfg['rms_attr'], None) or {}
            rms_tbl = _lookup_subcase(rms_dict, subcase)
            if rms_tbl is not None:
                arr = getattr(rms_tbl, id_attr)
                entity_ids = arr[:, 0] if id_attr == 'node_gridtype' else arr
                hits = np.where(entity_ids == nid)[0]
                if len(hits):
                    rms_native = float(rms_tbl.data[0, hits[0], idof])
                    return rms_native / unit_factor
        except Exception:
            pass
        area = AsdOverlayModule._grms_loglog(freqs, psd_curve)
        return float(np.sqrt(area))

    # ── Cycle helpers ────────────────────────────────────────────────────────

    def _cycle_subcase_pool(self):
        """Return [(sc_a_or_None, sc_b_or_None, desc), ...] for lock-step subcase cycling."""
        a_opts = self._op2_slots[0].get('subcase_options', [])
        b_opts = self._op2_slots[1].get('subcase_options', [])
        n = max(len(a_opts), len(b_opts), 1)
        out = []
        for i in range(n):
            sc_a = a_opts[i] if i < len(a_opts) else None
            sc_b = b_opts[i] if i < len(b_opts) else None
            parts = []
            if sc_a:
                parts.append(f"A: {sc_a[1]}")
            if sc_b:
                parts.append(f"B: {sc_b[1]}")
            out.append((sc_a, sc_b, "  /  ".join(parts) or f"Frame {i + 1}"))
        return out

    def _get_plot_frames(self):
        rt = self._rt_global_var.get()
        checked = [(n['id'], n['label'])
                   for n in self._nodes_by_rt.get(rt, [])
                   if n['checked'].get()]
        active_labels = self._active_dof_labels()
        dof_str = self._dof_var.get()
        cur_dof = active_labels.index(dof_str) if dof_str in active_labels else 0
        n_dof = len(active_labels)
        mode = self._view_mode_var.get()

        if mode == "Manual":
            return [[(nid, lbl, cur_dof) for nid, lbl in checked]], [""]

        if mode == "All grids, cycle DOF":
            return (
                [[(nid, lbl, d) for nid, lbl in checked] for d in range(n_dof)],
                [active_labels[d] for d in range(n_dof)],
            )

        if mode == "One grid, cycle DOF×grid":
            frames, descs = [], []
            for nid, lbl in checked:
                for d in range(n_dof):
                    frames.append([(nid, lbl, d)])
                    descs.append(f"{nid} {lbl} — {active_labels[d]}")
            return frames, descs

        if mode == "One grid all DOFs, cycle grid":
            return (
                [[(nid, lbl, d) for d in range(n_dof)] for nid, lbl in checked],
                [f"{nid} {lbl}" for nid, lbl in checked],
            )

        if mode == "Cycle subcases":
            pool = self._cycle_subcase_pool()
            if not pool or (not self._op2_slots[0]['op2'] and
                             not self._op2_slots[1]['op2']):
                return [[(nid, lbl, cur_dof) for nid, lbl in checked]], [""]
            frames = [[(nid, lbl, cur_dof) for nid, lbl in checked]
                      for _ in pool]
            descs = [desc for _a, _b, desc in pool]
            return frames, descs

        return [[(nid, lbl, cur_dof) for nid, lbl in checked]], [""]

    def _update_cycle_controls(self, frames, descs):
        mode = self._view_mode_var.get()
        n = len(frames)
        i = self._cycle_index

        if mode == "Manual" or n <= 1:
            self._cycle_label.configure(text="")
            self._prev_btn.configure(state=tk.DISABLED)
            self._next_btn.configure(state=tk.DISABLED)
            return

        desc = descs[i] if i < len(descs) else ""
        self._cycle_label.configure(text=f"{i + 1} of {n}: {desc}")
        self._prev_btn.configure(state=tk.NORMAL if i > 0 else tk.DISABLED)
        self._next_btn.configure(state=tk.NORMAL if i < n - 1 else tk.DISABLED)

    def _on_view_mode_change(self, _val=None):
        self._cycle_index = 0
        self._refresh_plot()

    def _step_cycle(self, delta):
        self._cycle_index += delta
        self._refresh_plot()

    # ── Peak picking ─────────────────────────────────────────────────────────

    def _toggle_pick_mode(self):
        self._pick_peaks_mode = not self._pick_peaks_mode
        if self._pick_peaks_mode:
            self._pick_btn.configure(fg_color=("#1f6aa5", "#1f6aa5"),
                                     text="Pick Peaks (ON)")
            self._canvas.get_tk_widget().configure(cursor="cross")
        else:
            self._pick_btn.configure(fg_color=("gray75", "gray25"),
                                     text="Pick Peaks")
            self._canvas.get_tk_widget().configure(cursor="")

    def _clear_peaks(self):
        self._picked_peaks.clear()
        self._refresh_plot()

    def _on_canvas_click(self, event):
        if not self._pick_peaks_mode:
            return
        if self._plot_mode_var.get() == "Cumulative RMS":
            return
        if event.inaxes is None or event.xdata is None or event.ydata is None:
            return
        if event.xdata <= 0 or event.ydata <= 0:
            return

        from scipy.signal import find_peaks

        click_lx = np.log10(event.xdata)
        click_ly = np.log10(event.ydata)

        best = None
        for curve in self._last_drawn_curves:
            if not curve["is_psd"]:
                continue
            freqs = curve["freqs"]
            data = curve["data"]
            peak_idx, _ = find_peaks(data)
            if not len(peak_idx):
                continue
            for i in peak_idx:
                f, v = float(freqs[i]), float(data[i])
                if f <= 0 or v <= 0:
                    continue
                d = ((np.log10(f) - click_lx) ** 2
                     + (np.log10(v) - click_ly) ** 2) ** 0.5
                if best is None or d < best[0]:
                    best = (d, curve["slot_idx"], curve["nid"],
                            curve["idof"], f, v)

        if best is None or best[0] > 0.5:
            return
        _pk_curve = next(
            (cv for cv in self._last_drawn_curves
             if cv["slot_idx"] == best[1] and cv["nid"] == best[2]
             and cv["idof"] == best[3]), None)
        self._picked_peaks.append({
            "slot_idx": best[1], "nid": best[2], "idof": best[3],
            "freq": best[4], "value": best[5],
            "label": _pk_curve.get("label", "") if _pk_curve else "",
        })
        self._refresh_plot()

    # ── Plot mode ─────────────────────────────────────────────────────────────

    def _on_plot_mode_change(self, _val=None):
        self._refresh_plot()

    # ── Excel export ─────────────────────────────────────────────────────────

    def _suggested_export_name(self):
        stem = self._title_var.get().strip() or self._env_var.get().strip() or "ASD_Overlay"
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in stem)
        return f"{safe.strip()}.xlsx"

    def _export_excel(self):
        if not self._last_drawn_curves and not any(
                r['checked'].get() for r in self._refs):
            messagebox.showwarning("Nothing to export",
                                   "Plot is empty — load OP2 data or references first.")
            return
        path = filedialog.asksaveasfilename(
            title="Export ASD data to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=self._suggested_export_name(),
        )
        if not path:
            return
        try:
            self._write_excel(path)
        except Exception as exc:
            messagebox.showerror("Export Error", str(exc))
            return
        self._show_export_done_dialog(path)

    def _show_export_done_dialog(self, path):
        import subprocess
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("Export complete")
        dlg.geometry("340x130")
        dlg.resizable(False, False)
        dlg.transient(self.frame.winfo_toplevel())

        ctk.CTkLabel(dlg, text=f"Saved: {os.path.basename(path)}",
                     anchor=tk.W).pack(padx=14, pady=(14, 8), fill=tk.X)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(padx=14, pady=(0, 14), fill=tk.X)

        def _open_file():
            try:
                if os.name == 'nt':
                    os.startfile(path)
                elif os.uname().sysname == 'Darwin':
                    subprocess.run(["open", path], check=False)
                else:
                    subprocess.run(["xdg-open", path], check=False)
            except Exception:
                pass
            dlg.destroy()

        def _open_folder():
            folder = os.path.dirname(path)
            try:
                if os.name == 'nt':
                    os.startfile(folder)
                elif os.uname().sysname == 'Darwin':
                    subprocess.run(["open", folder], check=False)
                else:
                    subprocess.run(["xdg-open", folder], check=False)
            except Exception:
                pass
            dlg.destroy()

        ctk.CTkButton(btn_row, text="Open File", command=_open_file,
                      width=90).pack(side=tk.LEFT, padx=(0, 6))
        ctk.CTkButton(btn_row, text="Open Folder", command=_open_folder,
                      width=100).pack(side=tk.LEFT, padx=(0, 6))
        ctk.CTkButton(btn_row, text="Close", command=dlg.destroy,
                      width=70).pack(side=tk.LEFT)

    def _clear_slot(self, slot_idx):
        self._op2_slots[slot_idx] = self._empty_slot()
        self._file_label[slot_idx].configure(text="(no file)", text_color="gray")
        self._suppress_name_trace[slot_idx] = True
        try:
            self._name_var[slot_idx].set("")
        finally:
            self._suppress_name_trace[slot_idx] = False
        self._name_user_edited[slot_idx] = False
        self._sc_btn[slot_idx].configure(text="(none)")
        self._input_asd_label[slot_idx].configure(text="(no file)", text_color="gray")
        self._input_asd_db_var[slot_idx].set("0")
        self._frf_row[slot_idx].pack_forget()
        self._update_dof_dropdown()
        self._rebuild_sections()
        self._status_label.configure(
            text=f"Slot {_SLOT_TAGS[slot_idx]} cleared", text_color="gray")
        self._refresh_plot()

    def _active_dof_labels(self):
        """DOF labels for the current global response type."""
        rt = self._rt_global_var.get()
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
        return list(cfg['dof_labels'])

    def _update_dof_dropdown(self):
        if self._dof_menu is None:
            return
        labels = self._active_dof_labels()
        self._dof_menu.configure(values=labels)
        if self._dof_var.get() not in labels:
            self._dof_var.set(labels[0])

    def _on_rt_global_change(self, rt):
        cfg = RESPONSE_TYPES[rt]
        for idx in range(2):
            cur_unit = self._unit_var[idx].get()
            self._unit_menu[idx].configure(values=cfg['unit_choices'])
            self._unit_var[idx].set(
                cur_unit if cur_unit in cfg['unit_choices'] else cfg['unit_choices'][0])
            op2 = self._op2_slots[idx].get('op2')
            if op2 is not None:
                psd_dict = getattr(op2.op2_results.psd, cfg['psd_attr'], None) or {}
                frf_dict = getattr(op2, cfg['frf_attr'], None) or {}
                if psd_dict:
                    result_dict = psd_dict
                elif frf_dict:
                    result_dict = frf_dict
                else:
                    messagebox.showwarning(
                        "No Data",
                        f"Slot {_SLOT_TAGS[idx]} OP2 has no {rt} results.\n"
                        "Subcase list cleared for this slot.")
                    self._op2_slots[idx]['subcase'] = None
                    self._op2_slots[idx]['subcases'] = []
                    self._op2_slots[idx]['subcase_options'] = []
                    self._sc_btn[idx].configure(text="(none)")
                    continue
                sc_pairs = _subcase_options(result_dict)
                self._op2_slots[idx]['subcase_options'] = sc_pairs
                # Preserve existing selections that are still valid; otherwise reset to first
                existing = self._op2_slots[idx].get('subcases', [])
                valid = [sc for sc in existing if any(s == sc for s, _ in sc_pairs)]
                if not valid:
                    valid = [sc_pairs[0][0]]
                self._op2_slots[idx]['subcases'] = valid
                self._op2_slots[idx]['subcase'] = valid[0]
                self._sc_btn[idx].configure(text=self._sc_btn_label(idx))
        self._update_dof_dropdown()
        self._rebuild_sections()
        self._refresh_plot()

    def _rebuild_sections(self):
        """Show the section for the current global RT; hide all others."""
        any_op2 = any(s['op2'] is not None for s in self._op2_slots.values())
        self._sections_placeholder.pack_forget()
        for w in self._section_widgets.values():
            w['frame'].pack_forget()
        if not any_op2:
            self._sections_placeholder.pack(pady=20)
            return
        rt = self._rt_global_var.get()
        if rt in self._section_widgets:
            self._section_widgets[rt]['frame'].pack(fill=tk.X)

    def _get_unit_factor(self, slot_idx):
        rt = self._rt_global_var.get()
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
        unit = self._unit_var[slot_idx].get()
        return cfg['unit_factors'].get(unit, 1.0)

    def _load_reference_asd_from_data(self, path, freqs_raw, g2hz_raw, name, checked,
                                       db=0.0, is_manual=False, manual_text="", style=None):
        if path:
            name = name or os.path.splitext(os.path.basename(path))[0]
        name = name or "ASD"
        db = float(db) if db else 0.0
        g2hz_scaled = g2hz_raw * 10.0 ** (db / 10.0) if db != 0.0 else g2hz_raw

        var = tk.BooleanVar(value=checked)
        name_var = tk.StringVar(value=name)
        db_var = tk.StringVar(value=f"{db:.1f}" if db != 0.0 else "0")

        container = ctk.CTkFrame(self._ref_scroll, fg_color="transparent")
        container.pack(fill=tk.X, pady=1)

        row1 = ctk.CTkFrame(container, fg_color="transparent")
        row1.pack(fill=tk.X)
        row2 = ctk.CTkFrame(container, fg_color="transparent")
        row2.pack(fill=tk.X, padx=(26, 0))

        ref = {
            "path": path, "name": name,
            "freqs_raw": freqs_raw, "g2hz_raw": g2hz_raw,
            "freqs": freqs_raw, "g2hz": g2hz_scaled,
            "db": db, "is_manual": is_manual, "manual_text": manual_text,
            "checked": var, "name_var": name_var, "db_var": db_var,
            "row_frame": container, "grms_label": None,
            "style": style if style is not None else {},
        }
        self._refs.append(ref)

        # row1: checkbox, name entry, X
        ctk.CTkCheckBox(row1, text="", variable=var, width=24,
                        command=self._refresh_plot).pack(side=tk.LEFT, padx=(2, 0))
        name_entry = ctk.CTkEntry(row1, textvariable=name_var, width=128)
        name_entry.pack(side=tk.LEFT, padx=(2, 2))
        name_entry.bind("<Return>",   lambda _e, r=ref: self._commit_ref_name(r))
        name_entry.bind("<FocusOut>", lambda _e, r=ref: self._commit_ref_name(r))
        ctk.CTkButton(row1, text="✕", width=22,
                      command=lambda r=ref: self._remove_reference(r),
                      fg_color="transparent", hover_color=("gray75", "gray30"),
                      text_color=("gray40", "gray60"),
                      ).pack(side=tk.LEFT)

        # row2: dB entry, g_RMS readout, Edit button
        ctk.CTkLabel(row2, text="dB:", width=24).pack(side=tk.LEFT)
        db_entry = ctk.CTkEntry(row2, textvariable=db_var, width=46)
        db_entry.pack(side=tk.LEFT, padx=(0, 4))
        db_entry.bind("<Return>",   lambda _e, r=ref: self._on_ref_db_change(r))
        db_entry.bind("<FocusOut>", lambda _e, r=ref: self._on_ref_db_change(r))

        grms_init = float(np.sqrt(max(self._grms_loglog(freqs_raw, g2hz_scaled), 0.0)))
        grms_lbl = ctk.CTkLabel(row2, text=f"{grms_init:.3g} g", width=50, anchor=tk.W,
                                text_color="gray")
        grms_lbl.pack(side=tk.LEFT, padx=(0, 4))
        ref['grms_label'] = grms_lbl

        ctk.CTkButton(row2, text="Edit…", width=46,
                      command=lambda r=ref: self._curve_style_dialog(r, is_ref=True),
                      ).pack(side=tk.LEFT)

    def _save_session(self):
        import json, datetime
        if not any(s['op2'] is not None for s in self._op2_slots.values()):
            messagebox.showwarning("Nothing to save", "Load at least one OP2 first.")
            return
        stem = (self._title_var.get().strip()
                or self._env_var.get().strip()
                or "asd_session")
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in stem)
        path = filedialog.asksaveasfilename(
            title="Save Session",
            defaultextension=".asdsession.json",
            filetypes=[("ASD session", "*.asdsession.json"), ("All files", "*.*")],
            initialfile=f"{safe.strip()}.asdsession.json",
        )
        if not path:
            return

        nodes_by_rt = {
            rt: [{'id': n['id'], 'label': n['label'],
                  'style': n.get('style', {})} for n in nodes]
            for rt, nodes in self._nodes_by_rt.items()
            if nodes
        }
        checked_by_rt = {
            rt: [n['id'] for n in nodes if n['checked'].get()]
            for rt, nodes in self._nodes_by_rt.items()
            if nodes
        }

        slots_data = []
        for idx, slot in self._op2_slots.items():
            slots_data.append({
                "tag": _SLOT_TAGS[idx],
                "name": self._name_var[idx].get(),
                "op2_path": slot.get('path'),
                "mode": slot.get('mode', 'PSD'),
                "units": self._unit_var[idx].get(),
                "subcase": slot.get('subcase'),
                "subcases": slot.get('subcases', []),
                "input_asd_path": slot.get('input_asd_path'),
                "input_asd_db": slot.get('input_asd_db', 0.0),
            })

        refs_data = []
        for r in self._refs:
            rd = {
                "path": r['path'],
                "name": r['name_var'].get() or r['name'],
                "checked": r['checked'].get(),
                "db": r.get('db', 0.0),
                "is_manual": r.get('is_manual', False),
                "style": r.get('style', {}),
            }
            if r.get('is_manual'):
                rd['manual_text'] = r.get('manual_text', '')
                rd['freqs_raw'] = r['freqs_raw'].tolist()
                rd['g2hz_raw'] = r['g2hz_raw'].tolist()
            refs_data.append(rd)

        data = {
            "version": 2,
            "tool": "ASD Overlay",
            "saved_at": datetime.datetime.now().isoformat(timespec='seconds'),
            "response_type": self._rt_global_var.get(),
            "slots": slots_data,
            "nodes_by_rt": nodes_by_rt,
            "checked_by_rt": checked_by_rt,
            "references": refs_data,
            "aux_lines": list(self._aux_lines),
            "view": {
                "view_mode": self._view_mode_var.get(),
                "plot_mode": self._plot_mode_var.get(),
                "y_axis": self._yscale_var.get(),
                "title": self._title_var.get(),
                "env": self._env_var.get(),
                "label_style": self._peak_label_style.get(),
            },
            "picked_peaks": [
                {"slot_idx": pk['slot_idx'], "nid": pk['nid'],
                 "idof": pk['idof'], "freq": pk['freq'], "value": pk['value'],
                 "label": pk.get('label', '')}
                for pk in self._picked_peaks
            ],
        }
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc))
            return
        self._show_export_done_dialog(path)

    def _open_session(self):
        import json
        path = filedialog.askopenfilename(
            title="Open Session",
            filetypes=[("ASD session", "*.asdsession.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
        except Exception as exc:
            messagebox.showerror("Load Error", f"Could not read session:\n{exc}")
            return
        if data.get('version') not in (1, 2):
            messagebox.showerror("Unsupported Version",
                                 "Session file version not supported.")
            return

        view = data.get('view', {})
        self._view_mode_var.set(view.get('view_mode', 'Manual'))
        _pm = view.get('plot_mode', 'ASD')
        if _pm == 'Cumulative GRMS':
            _pm = 'Cumulative RMS'
        self._plot_mode_var.set(_pm)
        self._yscale_var.set(view.get('y_axis', 'Log'))
        self._suppress_env_trace = True
        try:
            self._env_var.set(view.get('env', ''))
        finally:
            self._suppress_env_trace = False
        self._env_user_edited = bool(view.get('env', ''))
        self._title_var.set(view.get('title', ''))
        self._peak_label_style.set(view.get('label_style', 'Freq only'))

        # Global response type (new format) or infer from first slot (legacy)
        _loaded_rt = data.get('response_type')
        if _loaded_rt is None:
            for _sd in data.get('slots', []):
                _loaded_rt = _sd.get('response_type', 'Acceleration')
                break
        if _loaded_rt not in RESPONSE_TYPES:
            _loaded_rt = 'Acceleration'
        self._rt_global_var.set(_loaded_rt)
        self._on_rt_global_change(_loaded_rt)

        # Nodes — new per-RT format or legacy flat list
        for _rt_key in RESPONSE_TYPES:
            for _n in self._nodes_by_rt.get(_rt_key, []):
                _n['row_frame'].destroy()
            self._nodes_by_rt[_rt_key] = []

        nodes_by_rt = data.get('nodes_by_rt')
        if nodes_by_rt is not None:
            checked_by_rt = data.get('checked_by_rt', {})
            for _rt_key, _nlist in nodes_by_rt.items():
                if _rt_key not in RESPONSE_TYPES:
                    continue
                _checked_ids = set(checked_by_rt.get(_rt_key, []))
                for _n in _nlist:
                    self._add_node_row(_n['id'], _n.get('label', f"{_rt_key} {_n['id']}"), _rt_key)
                    if _n.get('style'):
                        self._nodes_by_rt[_rt_key][-1]['style'] = _n['style']
                for _node in self._nodes_by_rt[_rt_key]:
                    _node['checked'].set(_node['id'] in _checked_ids)
        else:
            # Legacy: single flat list → put in global RT bucket
            _checked_ids = set(data.get('checked_nodes', []))
            for _n in data.get('nodes', []):
                self._add_node_row(_n['id'], _n.get('label', f"Node {_n['id']}"), _loaded_rt)
            for _node in self._nodes_by_rt[_loaded_rt]:
                _node['checked'].set(_node['id'] in _checked_ids)

        self._picked_peaks = [
            {"slot_idx": pk['slot_idx'], "nid": pk['nid'],
             "idof": pk['idof'], "freq": pk['freq'], "value": pk['value'],
             "label": pk.get('label', '')}
            for pk in data.get('picked_peaks', [])
        ]

        self._aux_lines = list(data.get('aux_lines', []))

        self._clear_references()
        missing_refs = []
        for ref in data.get('references', []):
            rpath = ref.get('path', '')
            _is_manual = ref.get('is_manual', False)
            _db = ref.get('db', 0.0)
            _style = ref.get('style', {})
            if _is_manual:
                _mt = ref.get('manual_text', '')
                _fr = ref.get('freqs_raw')
                _gr = ref.get('g2hz_raw')
                if _fr and _gr:
                    self._load_reference_asd_from_data(
                        path=None,
                        freqs_raw=np.array(_fr), g2hz_raw=np.array(_gr),
                        name=ref.get('name', ''), checked=ref.get('checked', True),
                        db=_db, is_manual=True, manual_text=_mt, style=_style)
                else:
                    missing_refs.append(f"  Manual ASD '{ref.get('name', '?')}' (data missing from session)")
            elif rpath:
                if os.path.isfile(rpath):
                    freqs, asds = self._parse_asd_text_file(rpath)
                    if freqs is not None:
                        self._load_reference_asd_from_data(
                            rpath, freqs, asds, ref.get('name', ''),
                            ref.get('checked', True), db=_db, style=_style)
                else:
                    missing_refs.append(f"  Ref ASD: {rpath}")
        if missing_refs:
            messagebox.showwarning(
                "Missing Reference Files",
                "These reference ASD files could not be found:\n\n"
                + "\n".join(missing_refs))

        self._rebuild_sections()

        missing = []
        for slot_data in data.get('slots', []):
            idx = _SLOT_TAGS.index(slot_data.get('tag', 'A'))
            cfg = RESPONSE_TYPES.get(_loaded_rt, RESPONSE_TYPES['Acceleration'])
            unit = slot_data.get('units', cfg['unit_choices'][0])
            if unit in cfg['unit_choices']:
                self._unit_var[idx].set(unit)
            else:
                self._unit_var[idx].set(cfg['unit_choices'][0])

            op2_path = slot_data.get('op2_path')
            if not op2_path or not os.path.isfile(op2_path):
                if op2_path:
                    missing.append(f"Slot {slot_data.get('tag')}: {op2_path}")
                saved_name = slot_data.get('name', '')
                if saved_name:
                    self._suppress_name_trace[idx] = True
                    try:
                        self._name_var[idx].set(saved_name)
                    finally:
                        self._suppress_name_trace[idx] = False
                continue

            _target_subcase = slot_data.get('subcase')
            _target_subcases = slot_data.get('subcases') or (
                [_target_subcase] if _target_subcase is not None else [])
            _target_name = slot_data.get('name', '')
            _target_input_asd = slot_data.get('input_asd_path')
            _target_input_asd_db = float(slot_data.get('input_asd_db', 0.0) or 0.0)
            _target_mode = slot_data.get('mode', 'PSD')
            _slot_idx = idx

            def _work(p=op2_path):
                from pyNastran.op2.op2 import OP2
                op2 = OP2(mode='nx', debug=False)
                op2.read_op2(p)
                return op2

            def _done(op2, error,
                      si=_slot_idx, cfg_=cfg,
                      tsc=_target_subcase, tscs=_target_subcases, tname=_target_name,
                      tasd=_target_input_asd, tasd_db=_target_input_asd_db,
                      tmode=_target_mode, op2_path_=op2_path):
                if error is not None or op2 is None:
                    return
                psd_dict = getattr(op2.op2_results.psd, cfg_['psd_attr'], None) or {}
                frf_dict = getattr(op2, cfg_['frf_attr'], None) or {}
                if psd_dict:
                    result_dict = psd_dict
                    mode = "PSD"
                elif frf_dict:
                    result_dict = frf_dict
                    mode = "FRF"
                else:
                    return
                if tmode == "FRF" and frf_dict:
                    mode = "FRF"
                    result_dict = frf_dict
                self._op2_slots[si]['op2'] = op2
                self._op2_slots[si]['path'] = op2_path_
                self._op2_slots[si]['mode'] = mode
                mode_label = "PSD (RANDOM)" if mode == "PSD" else "FRF + Input ASD"
                self._mode_var[si].set(mode_label)
                if mode == "FRF":
                    self._frf_row[si].pack(fill=tk.X, pady=1)
                else:
                    self._frf_row[si].pack_forget()
                sc_pairs = _subcase_options(result_dict)
                self._op2_slots[si]['subcase_options'] = sc_pairs
                valid_ids = {sc for sc, _ in sc_pairs}
                restored = [sc for sc in tscs if sc in valid_ids]
                if not restored:
                    # Fall back to legacy single subcase, then first available
                    if tsc in valid_ids:
                        restored = [tsc]
                    elif sc_pairs:
                        restored = [sc_pairs[0][0]]
                self._op2_slots[si]['subcases'] = restored
                self._op2_slots[si]['subcase'] = restored[0] if restored else None
                self._sc_btn[si].configure(text=self._sc_btn_label(si))
                if tname:
                    self._suppress_name_trace[si] = True
                    try:
                        self._name_var[si].set(tname)
                    finally:
                        self._suppress_name_trace[si] = False
                self._file_label[si].configure(
                    text=os.path.basename(op2_path_),
                    text_color=("gray10", "gray90"))
                if tasd:
                    if os.path.isfile(tasd):
                        freqs, asds = self._parse_asd_text_file(tasd)
                        if freqs is not None:
                            self._op2_slots[si]['input_asd_path'] = tasd
                            self._op2_slots[si]['input_asd_freqs'] = freqs
                            self._op2_slots[si]['input_asd_g2hz_raw'] = asds
                            self._op2_slots[si]['input_asd_db'] = tasd_db
                            self._op2_slots[si]['input_asd_g2hz'] = (
                                asds * 10.0 ** (tasd_db / 10.0) if tasd_db else asds)
                            self._input_asd_db_var[si].set(
                                f"{tasd_db:.1f}" if tasd_db else "0")
                            self._input_asd_label[si].configure(
                                text=os.path.basename(tasd),
                                text_color=("gray10", "gray90"))
                    else:
                        self._input_asd_label[si].configure(
                            text=f"⚠ not found: {os.path.basename(tasd)}",
                            text_color="orange")
                self._rebuild_sections()
                self._refresh_plot()

            self._run_in_background(
                f"Loading OP2 {slot_data.get('tag')}…", _work, _done)

        if missing:
            messagebox.showwarning(
                "Missing Files",
                "The following files could not be found:\n\n"
                + "\n".join(missing))

    def _fmt_entity_label(self, nid, idof, node_label=""):
        """Format 'Entity NID (user label) DOF' matching the plot legend."""
        rt = self._rt_global_var.get()
        cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
        entity = cfg['entity_label']
        dofs = cfg['dof_labels']
        dof = dofs[idof] if idof < len(dofs) else str(idof)
        tail = f" ({node_label.strip()})" if node_label and node_label.strip() else ""
        return f"{entity} {nid}{tail} {dof}"

    def _build_sheet_blocks(self, is_cum):
        """Build column blocks for data sheets. One block per slot, one per reference.

        Each block: {'freq_label', 'freqs', 'columns': [{'label', 'values'}, ...]}
        Curves within a slot all share the same freq vector.
        """
        blocks = []
        by_slot_sc = {}
        for c in self._last_drawn_curves:
            key = (c['slot_idx'], c.get('subcase'))
            by_slot_sc.setdefault(key, []).append(c)

        for (slot_idx, sc) in sorted(by_slot_sc):
            curves = by_slot_sc[(slot_idx, sc)]
            tag = _SLOT_TAGS[slot_idx]
            name = self._name_var[slot_idx].get().strip() or tag
            slot_cfg = RESPONSE_TYPES.get(
                self._rt_global_var.get(), RESPONSE_TYPES['Acceleration'])
            freqs = curves[0]['freqs']
            cols = []
            for c in curves:
                base = f"{name}: {self._fmt_entity_label(c['nid'], c['idof'], c.get('label', ''))}"
                if c['is_psd']:
                    unit = slot_cfg['rms_units'] if is_cum else slot_cfg['psd_units']
                    vals = (self._cumulative_grms_loglog(c['freqs'], c['data'])
                            if is_cum else c['data'])
                else:
                    if is_cum:
                        continue  # skip FRF magnitude in cumulative sheet
                    unit = slot_cfg['frf_units']
                    vals = c['data']
                cols.append({'label': f"{base}  [{unit}]",
                             'values': np.asarray(vals)})
            if cols:
                blocks.append({
                    'freq_label': f"Frequency (Hz) — {tag} (sc {sc})",
                    'freqs': np.asarray(freqs),
                    'columns': cols,
                })

        for ref in self._refs:
            if not ref['checked'].get():
                continue
            rname = ref['name_var'].get().strip() or ref['name']
            unit = "Cumulative RMS" if is_cum else "PSD"
            vals = (self._cumulative_grms_loglog(ref['freqs'], ref['g2hz'])
                    if is_cum else ref['g2hz'])
            blocks.append({
                'freq_label': f"Frequency (Hz) — [Ref] {rname}",
                'freqs': np.asarray(ref['freqs']),
                'columns': [{'label': f"[Ref] {rname}  [{unit}]",
                              'values': np.asarray(vals)}],
            })
        return blocks

    def _write_data_sheet(self, ws, blocks):
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        if not blocks:
            ws.cell(1, 1, "(no data)")
            return

        # Header row
        col = 1
        for block in blocks:
            ws.cell(1, col, block['freq_label'])
            for k, c in enumerate(block['columns']):
                ws.cell(1, col + 1 + k, c['label'])
            col += 1 + len(block['columns'])

        # Data rows — pad shorter blocks with blanks
        col = 1
        for block in blocks:
            for r, f in enumerate(block['freqs'], start=2):
                ws.cell(r, col, float(f))
                for k, c in enumerate(block['columns']):
                    if r - 2 < len(c['values']):
                        ws.cell(r, col + 1 + k, float(c['values'][r - 2]))
            col += 1 + len(block['columns'])

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "A2"
        for ci in range(1, col):
            ws.column_dimensions[get_column_letter(ci)].width = 22

    def _write_summary_sheet(self, ws):
        import datetime
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        bold = Font(bold=True)

        def h(row, col, val):
            c = ws.cell(row, col, val)
            c.font = bold

        def v(row, col, val):
            ws.cell(row, col, val)

        # Fixed preamble (rows 1–8)
        h(1, 1, "Random Response Overlay Export")
        h(2, 1, "Title:"),       v(2, 2, self._title_var.get().strip())
        h(3, 1, "Environment:"), v(3, 2, self._env_var.get().strip())
        h(4, 1, "Plot mode:"),   v(4, 2, self._plot_mode_var.get())
        h(5, 1, "Y-axis:"),      v(5, 2, self._yscale_var.get())
        h(6, 1, "View mode:"),   v(6, 2, self._view_mode_var.get())
        h(7, 1, "Exported:"),    v(7, 2, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        # Row 8 intentionally blank

        # Curve summary table
        row = 9
        for col, hdr in enumerate(["Slot", "Name", "OP2 file", "Subcase",
                                    "Curve (Entity / DOF)", "RMS"], start=1):
            h(row, col, hdr)
        row += 1

        for c in self._last_drawn_curves:
            slot = self._op2_slots[c['slot_idx']]
            tag = _SLOT_TAGS[c['slot_idx']]
            name = self._name_var[c['slot_idx']].get().strip() or tag
            op2_file = os.path.basename(slot['path']) if slot['path'] else ""
            sc = c.get('subcase', slot['subcase'])
            curve_lbl = self._fmt_entity_label(c['nid'], c['idof'], c.get('label', ''))
            if c['is_psd']:
                rms = self._get_rms_scalar(
                    slot['op2'], sc, c['nid'], c['idof'],
                    c['freqs'], c['data'],
                    self._get_unit_factor(c['slot_idx']),
                    rt=self._rt_global_var.get())
                grms_str = f"{rms:.4g}"
            else:
                grms_str = "n/a (FRF)"
            for col, val in enumerate([tag, name, op2_file, sc, curve_lbl, grms_str],
                                       start=1):
                ws.cell(row, col, val)
            row += 1

        # References
        if any(r['checked'].get() for r in self._refs):
            row += 1
            h(row, 1, "[Ref] Name")
            h(row, 2, "File")
            h(row, 3, "Freq range")
            h(row, 4, "RMS")
            row += 1
            for ref in self._refs:
                if not ref['checked'].get():
                    continue
                rname = ref['name_var'].get().strip() or ref['name']
                grms = float(np.sqrt(max(self._grms_loglog(ref['freqs'], ref['g2hz']), 0.0)))
                freq_range = f"{ref['freqs'][0]:.1f}–{ref['freqs'][-1]:.1f} Hz"
                _ref_src = "(manual)" if ref.get('is_manual') else os.path.basename(ref['path'] or '')
                for col, val in enumerate(
                        [rname, _ref_src, freq_range, f"{grms:.4g}"],
                        start=1):
                    ws.cell(row, col, val)
                row += 1

        # Picked peaks
        row += 1
        h(row, 1, "Picked Peaks")
        row += 1
        if self._picked_peaks:
            for col, hdr in enumerate(["Slot", "Entity", "DOF", "Label",
                                        "Freq (Hz)", "Value"], start=1):
                h(row, col, hdr)
            row += 1
            for pk in self._picked_peaks:
                tag = _SLOT_TAGS[pk['slot_idx']]
                _rt_pk = self._rt_global_var.get()
                _cfg_pk = RESPONSE_TYPES.get(_rt_pk, RESPONSE_TYPES['Acceleration'])
                _dofs_pk = _cfg_pk['dof_labels']
                _dof_pk = _dofs_pk[pk['idof']] if pk['idof'] < len(_dofs_pk) else str(pk['idof'])
                for col, val in enumerate(
                        [tag, pk['nid'], _dof_pk,
                         pk.get('label', ''), pk['freq'], pk['value']], start=1):
                    ws.cell(row, col, val)
                row += 1
        else:
            ws.cell(row, 1, "(none)")

        for ci in range(1, 7):
            ws.column_dimensions[get_column_letter(ci)].width = 20
        ws.freeze_panes = "A9"

    def _write_excel(self, path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary)

        ws_asd = wb.create_sheet("ASD")
        self._write_data_sheet(ws_asd, self._build_sheet_blocks(is_cum=False))

        ws_cum = wb.create_sheet("Cumulative RMS")
        self._write_data_sheet(ws_cum, self._build_sheet_blocks(is_cum=True))

        wb.save(path)

    # ── Plot ─────────────────────────────────────────────────────────────────

    def _refresh_plot(self):
        t = _THEMES[self._plot_theme]
        ax = self._ax
        ax.clear()
        ax.set_facecolor(t["plot_bg"])

        plot_mode = self._plot_mode_var.get()   # "ASD" | "Cumulative RMS"
        yscale = self._yscale_var.get().lower() # "log" | "linear"
        is_cum = (plot_mode == "Cumulative RMS")

        ax.set_xscale("log")
        ax.set_yscale(yscale)

        frames, descs = self._get_plot_frames()
        if not frames:
            frames, descs = [[]], [""]
        self._cycle_index = max(0, min(self._cycle_index, len(frames) - 1))
        curves = frames[self._cycle_index]

        view_mode = self._view_mode_var.get()
        color_by_dof = (view_mode == "One grid all DOFs, cycle grid")

        if view_mode in ("All grids, cycle DOF", "One grid, cycle DOF×grid"):
            if curves:
                active_dof = curves[0][2]
                _al = self._active_dof_labels()
                if active_dof < len(_al):
                    self._dof_var.set(_al[active_dof])

        # Subcase override for "Cycle subcases" mode
        sc_override = {0: None, 1: None}  # None means use slot's current selection
        if view_mode == "Cycle subcases":
            pool = self._cycle_subcase_pool()
            if pool and 0 <= self._cycle_index < len(pool):
                sc_a, sc_b, _ = pool[self._cycle_index]
                sc_override[0] = sc_a[0] if sc_a else "SKIP"
                sc_override[1] = sc_b[0] if sc_b else "SKIP"
                # Sync buttons so user sees the active subcase
                if sc_a:
                    self._sc_btn[0].configure(text=sc_a[1][:35])
                if sc_b:
                    self._sc_btn[1].configure(text=sc_b[1][:35])

        has_curves = False
        has_psd = False
        has_frf_mag = False
        self._last_drawn_curves = []

        for slot_idx, slot in self._op2_slots.items():
            op2 = slot['op2']
            ov = sc_override[slot_idx]
            if ov == "SKIP":
                continue
            if ov is not None:
                subcases = [ov]
            else:
                subcases = slot.get('subcases') or (
                    [slot['subcase']] if slot['subcase'] is not None else [])
            if op2 is None or not subcases:
                continue

            unit_factor = self._get_unit_factor(slot_idx)
            rt = self._rt_global_var.get()
            cfg = RESPONSE_TYPES.get(rt, RESPONSE_TYPES['Acceleration'])
            tag = _SLOT_TAGS[slot_idx]
            name = self._name_var[slot_idx].get().strip() or tag
            sc_opts = slot.get('subcase_options', [])
            multi_sc = len(subcases) > 1

            for sc_pos, subcase in enumerate(subcases):
                sc_ls = _SC_LINES[sc_pos % len(_SC_LINES)] if multi_sc else _SLOT_LINES[slot_idx]
                sc_suffix = ""
                if multi_sc:
                    sc_lbl = next((l for sc, l in sc_opts if sc == subcase), str(subcase))
                    sc_suffix = f" [{sc_lbl}]"

                for curve_idx, (nid, lbl, idof) in enumerate(curves):
                    color = (_NODE_COLORS[idof % len(_NODE_COLORS)] if color_by_dof
                             else _NODE_COLORS[curve_idx % len(_NODE_COLORS)])

                    freqs, data, is_psd = self._get_response_psd(
                        slot_idx, subcase, nid, idof, unit_factor)
                    if freqs is None:
                        continue

                    dof_labels = cfg.get('dof_labels', self.DOF_LABELS)
                    dof_label = dof_labels[idof] if idof < len(dof_labels) else str(idof)

                    if is_psd:
                        if is_cum:
                            plot_data = self._cumulative_grms_loglog(freqs, data)
                            final_g = float(plot_data[-1]) if len(plot_data) else 0.0
                            _fmt = cfg.get('rms_fmt', '.3g')
                            label = f"{name}{sc_suffix}: {lbl} {dof_label}  (final RMS = {final_g:{_fmt}} {cfg['rms_units']})"
                        else:
                            plot_data = data
                            rms_g = self._get_rms_scalar(
                                op2, subcase, nid, idof, freqs, data, unit_factor,
                                rt=rt)
                            _fmt = cfg.get('rms_fmt', '.3g')
                            label = f"{name}{sc_suffix}: {lbl} {dof_label}  (RMS = {rms_g:{_fmt}} {cfg['rms_units']})"
                        has_psd = True
                    else:
                        if is_cum:
                            continue  # cumulative only for PSD
                        plot_data = data
                        label = f"{name}{sc_suffix}: {lbl} {dof_label}  (FRF magnitude, {cfg['frf_units']})"
                        has_frf_mag = True

                    _node = next((n for n in self._nodes_by_rt.get(rt, [])
                                  if n['id'] == nid), None)
                    _st = _node.get('style', {}) if _node else {}
                    ax.plot(freqs, plot_data,
                            label=_st.get('label_override') or label,
                            color=_st.get('color') or color,
                            linestyle=_st.get('linestyle') or sc_ls,
                            linewidth=_st.get('linewidth') or 1.5)
                    has_curves = True
                    self._last_drawn_curves.append({
                        "slot_idx": slot_idx, "subcase": subcase,
                        "nid": nid, "idof": idof, "label": lbl,
                        "freqs": np.asarray(freqs), "data": np.asarray(data),
                        "is_psd": is_psd, "color": _st.get('color') or color,
                    })

        # ── Reference ASD overlays ────────────────────────────────────────────
        for ref_idx, ref in enumerate(self._refs):
            if not ref['checked'].get():
                continue
            color = self._REF_COLORS[ref_idx % len(self._REF_COLORS)]
            rname = ref['name_var'].get().strip() or ref['name']
            if is_cum:
                ref_plot = self._cumulative_grms_loglog(ref['freqs'], ref['g2hz'])
                final = float(ref_plot[-1]) if len(ref_plot) else 0.0
                ref_label = f"{rname}  (final RMS = {final:.3g})"
            else:
                ref_plot = ref['g2hz']
                grms = float(np.sqrt(max(self._grms_loglog(ref['freqs'], ref['g2hz']), 0.0)))
                ref_label = f"{rname}  (RMS = {grms:.3g})"
            _rst = ref.get('style', {})
            ax.plot(ref['freqs'], ref_plot,
                    label=_rst.get('label_override') or ref_label,
                    color=_rst.get('color') or color,
                    linestyle=_rst.get('linestyle') or ":",
                    linewidth=_rst.get('linewidth') or 2.0,
                    alpha=0.65)
            has_curves = True

        # ── X/Y reference lines ───────────────────────────────────────────────
        for _ln in self._aux_lines:
            if not _ln.get('visible', True):
                continue
            _lv = _ln.get('value', 0.0)
            _lc = _ln.get('color', '#888888')
            _lls = _ln.get('linestyle', '--')
            _llw = _ln.get('linewidth', 1.0)
            if _ln.get('axis', 'x') == 'x':
                ax.axvline(_lv, color=_lc, linestyle=_lls, linewidth=_llw,
                           alpha=0.7, zorder=2)
                if _ln.get('label'):
                    ax.text(_lv, 1.0, _ln['label'],
                            transform=ax.get_xaxis_transform(),
                            rotation=90, va='top', ha='right',
                            color=t['text'], fontsize=8,
                            bbox=dict(facecolor=t['legend_bg'], edgecolor=_lc,
                                      alpha=0.75, pad=2))
            else:
                ax.axhline(_lv, color=_lc, linestyle=_lls, linewidth=_llw,
                           alpha=0.7, zorder=2)
                if _ln.get('label'):
                    ax.text(1.0, _lv, _ln['label'],
                            transform=ax.get_yaxis_transform(),
                            va='bottom', ha='right',
                            color=t['text'], fontsize=8,
                            bbox=dict(facecolor=t['legend_bg'], edgecolor=_lc,
                                      alpha=0.75, pad=2))

        ax.set_xlabel("Frequency (Hz)", color=t["text"])
        _rt_now = self._rt_global_var.get()
        _cfg_now = RESPONSE_TYPES.get(_rt_now, RESPONSE_TYPES['Acceleration'])
        if is_cum:
            ax.set_ylabel(f"Cumulative RMS ({_cfg_now['rms_units']})", color=t["text"])
        elif has_psd and has_frf_mag:
            ax.set_ylabel("PSD / FRF Magnitude", color=t["text"])
        elif has_frf_mag:
            ax.set_ylabel(f"FRF Magnitude ({_cfg_now['frf_units']})", color=t["text"])
        else:
            ax.set_ylabel(f"PSD ({_cfg_now['psd_units']})", color=t["text"])

        ax.tick_params(colors=t["text"], which="both")
        for spine in ax.spines.values():
            spine.set_edgecolor(t["spine"])
        ax.grid(True, which="both", alpha=0.3, color=t["grid"])
        self._fig.set_facecolor(t["fig_bg"])

        if has_curves:
            ax.legend(loc="best", fontsize=8,
                      facecolor=t["legend_bg"], labelcolor=t["text"],
                      edgecolor=t["spine"])
        else:
            ax.text(0.5, 0.5,
                    "No data — check OP2 loaded, nodes added, and boxes checked",
                    transform=ax.transAxes,
                    ha="center", va="center", color="gray", fontsize=10)

        # ── Picked peak markers and labels (ASD mode only) ────────────────────
        if not is_cum:
            label_style = self._peak_label_style.get()
            for pk in self._picked_peaks:
                curve = next((c for c in self._last_drawn_curves
                              if c["slot_idx"] == pk["slot_idx"]
                              and c["nid"] == pk["nid"]
                              and c["idof"] == pk["idof"]), None)
                if curve is None:
                    continue
                f, v = pk["freq"], pk["value"]
                ax.plot([f], [v], marker="o", markersize=6,
                        markerfacecolor=curve["color"],
                        markeredgecolor=t["text"], linestyle="none", zorder=5)
                if label_style == "Freq + value":
                    _pk_cfg = RESPONSE_TYPES.get(
                        self._rt_global_var.get(), RESPONSE_TYPES['Acceleration'])
                    ann_text = f"{f:.0f} Hz\n{v:.3g} {_pk_cfg['psd_units']}"
                else:
                    ann_text = f"{f:.0f} Hz"
                ax.annotate(ann_text, xy=(f, v),
                            xytext=(6, 6), textcoords="offset points",
                            color=t["text"], fontsize=9,
                            bbox=dict(boxstyle="round,pad=0.2",
                                      fc=t["legend_bg"], ec=t["spine"], alpha=0.85))

        # ── Title and environment ─────────────────────────────────────────────
        plot_title = self._title_var.get().strip()
        env_name = self._env_var.get().strip()
        ax.set_title(
            plot_title or "",
            color=t["text"],
            fontsize=13,
            weight="bold",
            pad=20 if env_name else 6,
        )
        if env_name:
            ax.text(0.5, 1.01, env_name,
                    transform=ax.transAxes,
                    ha="center", va="bottom",
                    color=t["text"], fontsize=10,
                    style="italic", alpha=0.75)

        self._canvas.draw_idle()
        self._update_cycle_controls(frames, descs)
