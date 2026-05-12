"""CBUSH element forces module.

Reads CBUSH force results (Fx, Fy, Fz, Mx, My, Mz) from a Nastran OP2
file across all subcases/load cases.  For random analysis the values are
RMS.  Each load case can be named by the user and exported to its own
Excel sheet.

Optional BDF loading adds a Property column (comment names from PBUSH
cards) and enables sorting by property group.  Per-case scale factors
are applied on export only.
"""
import os
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from tksheet import Sheet

import numpy as np

_HEADERS_BASE = ['Axial Dir', 'EID', 'Fx', 'Fy', 'Fz', 'Mx', 'My', 'Mz',
                 'Axial', 'Comb Shear', 'Comb Moment']
_HEADERS_BASE_WITH_PROP = ['Property', 'Axial Dir', 'EID', 'Fx', 'Fy', 'Fz',
                           'Mx', 'My', 'Mz', 'Axial', 'Comb Shear',
                           'Comb Moment']
_FORCE_COLS = ['fx', 'fy', 'fz', 'mx', 'my', 'mz']

_AXIAL_MAP = {
    'X': {'axial': 0, 'shear': (1, 2), 'moment': (4, 5)},
    'Y': {'axial': 1, 'shear': (0, 2), 'moment': (3, 5)},
    'Z': {'axial': 2, 'shear': (0, 1), 'moment': (3, 4)},
}


def _compute_derived(forces_row, axial_dir):
    """Return (axial_force, combined_shear, combined_moment).

    Parameters
    ----------
    forces_row : array-like, length 6 — [Fx, Fy, Fz, Mx, My, Mz]
    axial_dir : str — 'X', 'Y', or 'Z'
    """
    m = _AXIAL_MAP[axial_dir]
    axial = float(forces_row[m['axial']])
    s1, s2 = m['shear']
    shear = float(np.sqrt(forces_row[s1] ** 2 + forces_row[s2] ** 2))
    m1, m2 = m['moment']
    moment = float(np.sqrt(forces_row[m1] ** 2 + forces_row[m2] ** 2))
    return axial, shear, moment


def _build_headers(has_bdf):
    """Return header list based on whether BDF is loaded."""
    return list(_HEADERS_BASE_WITH_PROP if has_bdf else _HEADERS_BASE)


# --------------------------------------------------------- Excel helpers

def make_cbush_styles():
    """Return a dict of openpyxl style objects for CBUSH Forces sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        'dark_fill': PatternFill("solid", fgColor="1F4E79"),
        'mid_fill': PatternFill("solid", fgColor="2E75B6"),
        'white_bold': Font(bold=True, color="FFFFFF", size=11),
        'sub_font': Font(bold=True, color="FFFFFF", size=10),
        'center': Alignment(horizontal="center", vertical="center"),
        'left': Alignment(horizontal="left", vertical="center"),
        'cell_border': Border(bottom=Side(style='thin', color="B4C6E7")),
        'sci_fmt': '0.00E+00',
    }


def write_cbush_sheet(ws, eids, forces, styles, op2_name=None,
                       title=None, sheet_label=None, prop_names=None,
                       scale_factor=1.0, start_row=0,
                       axial_dirs=None, derived=None):
    """Write one CBUSH forces block to an openpyxl worksheet.

    Parameters
    ----------
    ws : openpyxl Worksheet
    eids : array-like, shape (nelems,)
    forces : ndarray, shape (nelems, 6)
    styles : dict from ``make_cbush_styles()``
    op2_name : str or None -- OP2 filename for the header
    title : str or None -- optional user title (row 1)
    sheet_label : str or None -- load case label
    prop_names : list[str] or None -- per-element property names (same order as eids)
    scale_factor : float -- multiplier applied to force values
    start_row : int -- 0-based row offset (for stacking multiple blocks)
    axial_dirs : list[str] or None -- per-element axial direction ('X','Y','Z')
    derived : list[tuple] or None -- per-element (axial, shear, moment)

    Returns
    -------
    int : next available row (0-based) after this block
    """
    from openpyxl.utils import get_column_letter

    s = styles
    has_prop = prop_names is not None
    headers = _build_headers(has_prop)
    total_cols = len(headers)
    cur_row = start_row

    # Row 1 -- Title (always written; blank if no title)
    cur_row += 1
    cell = ws.cell(row=cur_row, column=1, value=title or "")
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # Row 2 -- OP2 filename (always written; blank if no op2_name)
    cur_row += 1
    cell = ws.cell(row=cur_row, column=1, value=op2_name or "")
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # Row 3 -- Load case label (always written; blank if no label)
    cur_row += 1
    cell = ws.cell(row=cur_row, column=1, value=sheet_label or "")
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # Column headers
    header_row = cur_row + 1
    cur_row = header_row
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data rows
    data_start = cur_row + 1
    col = 1  # running 1-based column

    for i in range(len(eids)):
        row = data_start + i
        col = 1
        if has_prop:
            c = ws.cell(row=row, column=col, value=prop_names[i])
            c.alignment = s['left']
            col += 1
        # Axial Dir
        ad = axial_dirs[i] if axial_dirs else 'X'
        ws.cell(row=row, column=col, value=ad).alignment = s['center']
        col += 1
        # EID
        ws.cell(row=row, column=col, value=int(eids[i])).alignment = s['center']
        col += 1
        # 6 force components
        for j in range(6):
            c = ws.cell(row=row, column=col + j,
                        value=float(forces[i, j]) * scale_factor)
            c.number_format = s['sci_fmt']
            c.alignment = s['center']
        col += 6
        # 3 derived columns
        if derived:
            for dval in derived[i]:
                c = ws.cell(row=row, column=col, value=dval * scale_factor)
                c.number_format = s['sci_fmt']
                c.alignment = s['center']
                col += 1
        else:
            col += 3
        # Border all cells
        for ci in range(1, total_cols + 1):
            ws.cell(row=row, column=ci).border = s['cell_border']

    # Column widths (only set on first block — start_row == 0)
    if start_row == 0:
        col = 1
        if has_prop:
            ws.column_dimensions[get_column_letter(col)].width = 18
            col += 1
        ws.column_dimensions[get_column_letter(col)].width = 10  # Axial Dir
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 10  # EID
        col += 1
        for ci in range(col, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = f'A{data_start}'

    return data_start + len(eids), header_row


def _write_cbush_block_combined(ws, eids, forces, styles, sheet_label=None,
                                prop_names=None, scale_factor=1.0,
                                start_row=0, axial_dirs=None, derived=None):
    """Write a subcase block for combined-sheet mode (no title/OP2 rows).

    Returns (next_row, header_row).
    """
    from openpyxl.utils import get_column_letter

    s = styles
    has_prop = prop_names is not None
    headers = _build_headers(has_prop)
    total_cols = len(headers)
    cur_row = start_row

    # Load case label row
    cur_row += 1
    cell = ws.cell(row=cur_row, column=1, value=sheet_label or "")
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # Column headers
    header_row = cur_row + 1
    cur_row = header_row
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data rows
    data_start = cur_row + 1

    for i in range(len(eids)):
        row = data_start + i
        col = 1
        if has_prop:
            c = ws.cell(row=row, column=col, value=prop_names[i])
            c.alignment = s['left']
            col += 1
        # Axial Dir
        ad = axial_dirs[i] if axial_dirs else 'X'
        ws.cell(row=row, column=col, value=ad).alignment = s['center']
        col += 1
        # EID
        ws.cell(row=row, column=col, value=int(eids[i])).alignment = s['center']
        col += 1
        # 6 force components
        for j in range(6):
            c = ws.cell(row=row, column=col + j,
                        value=float(forces[i, j]) * scale_factor)
            c.number_format = s['sci_fmt']
            c.alignment = s['center']
        col += 6
        # 3 derived columns
        if derived:
            for dval in derived[i]:
                c = ws.cell(row=row, column=col, value=dval * scale_factor)
                c.number_format = s['sci_fmt']
                c.alignment = s['center']
                col += 1
        # Border
        for ci in range(1, total_cols + 1):
            ws.cell(row=row, column=ci).border = s['cell_border']

    return data_start + len(eids), header_row


def _write_joint_summary_block(ws, joint_data, styles, header_row, start_col):
    """Write joint summary mini-table to the right of main data.

    Parameters
    ----------
    ws : openpyxl Worksheet
    joint_data : list of (joint_name, shear_total) tuples
    styles : dict from make_cbush_styles()
    header_row : int (1-based) — row of the column headers in the main block
    start_col : int (1-based) — first column for the joint summary
    """
    s = styles
    # Headers
    for ci, h in enumerate(['Joint', 'Shear Total']):
        cell = ws.cell(row=header_row, column=start_col + ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data rows
    from openpyxl.utils import get_column_letter
    for i, (name, total) in enumerate(joint_data):
        row = header_row + 1 + i
        ws.cell(row=row, column=start_col, value=name).alignment = s['left']
        c = ws.cell(row=row, column=start_col + 1, value=total)
        c.number_format = s['sci_fmt']
        c.alignment = s['center']
        for ci in range(2):
            ws.cell(row=row, column=start_col + ci).border = s['cell_border']

    # Column widths
    ws.column_dimensions[get_column_letter(start_col)].width = 16
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 14


# ---------------------------------------------------------- post-export dialog

def _open_path(path):
    """Open a file or directory with the platform default handler."""
    if sys.platform == 'darwin':
        subprocess.Popen(['open', path])
    elif sys.platform == 'win32':
        os.startfile(path)
    else:
        subprocess.Popen(['xdg-open', path])


class _ExportDoneDialog(tk.Toplevel):
    """Modal dialog shown after a successful Excel export."""

    def __init__(self, parent, message, file_path):
        super().__init__(parent)
        self.title("Exported")
        self.resizable(False, False)
        self._file_path = file_path

        tk.Label(self, text=message, justify='left').pack(
            padx=16, pady=(16, 8))

        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=(0, 16))

        tk.Button(btn_frame, text="Open File",
                  command=self._open_file).pack(side='left', padx=4)
        tk.Button(btn_frame, text="Open Folder",
                  command=self._open_folder).pack(side='left', padx=4)
        tk.Button(btn_frame, text="Close",
                  command=self.destroy).pack(side='left', padx=4)

        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _open_file(self):
        _open_path(self._file_path)
        self.destroy()

    def _open_folder(self):
        _open_path(os.path.dirname(self._file_path))
        self.destroy()


# --------------------------------------------------------- Manage Joints Dialog

class ManageJointsDialog(ctk.CTkToplevel):
    """Dialog for creating/editing joint definitions (property groups)."""

    def __init__(self, parent, available_pids, existing_joints, on_apply,
                 id_labels=None, bdf_path=None):
        super().__init__(parent)
        self.title("Manage Joints")
        self.geometry("600x450")
        self.resizable(True, True)
        self.transient(parent)

        self._available_pids = sorted(available_pids)
        self._id_labels = id_labels or {}
        self._joints = {k: set(v) for k, v in existing_joints.items()}
        self._on_apply = on_apply
        self._editing_key = None
        self._bdf_path = bdf_path

        self._build_ui()

    def _build_ui(self):
        dark = ctk.get_appearance_mode() == "Dark"
        self._dark = dark
        lb_bg = "#2b2b2b" if dark else "white"
        lb_fg = "#dce4ee" if dark else "black"
        lb_sel_bg = "#1f6aa5" if dark else "#0078d4"

        main = ctk.CTkFrame(self)
        main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left: available properties
        left = ctk.CTkFrame(main)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        ctk.CTkLabel(left, text="Available Properties",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W)
        self._pid_listbox = tk.Listbox(left, selectmode=tk.EXTENDED,
                                       exportselection=False,
                                       bg=lb_bg, fg=lb_fg,
                                       selectbackground=lb_sel_bg,
                                       selectforeground="white")
        self._pid_listbox.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        for pid in self._available_pids:
            self._pid_listbox.insert(tk.END, self._id_labels.get(pid, str(pid)))

        # Middle: controls
        mid = ctk.CTkFrame(main, width=150)
        mid.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        mid.pack_propagate(False)

        ctk.CTkLabel(mid, text="Joint Name:").pack(anchor=tk.W, pady=(20, 2))
        self._name_var = tk.StringVar()
        ctk.CTkEntry(mid, textvariable=self._name_var, width=130).pack()

        ctk.CTkButton(mid, text="Save Joint \u2192", width=130,
                      command=self._save_joint).pack(pady=(10, 2))
        ctk.CTkButton(mid, text="Clear", width=130,
                      fg_color="gray50",
                      command=self._clear_form).pack(pady=2)
        ctk.CTkButton(mid, text="Delete Joint", width=130,
                      fg_color="firebrick",
                      command=self._delete_joint).pack(pady=2)

        ctk.CTkButton(mid, text="Import CSV\u2026", width=130,
                      command=self._import_csv).pack(pady=(12, 2))
        ctk.CTkButton(mid, text="Export Template\u2026", width=130,
                      command=self._export_template).pack(pady=2)

        ctk.CTkButton(mid, text="Save Joints\u2026", width=130,
                      command=self._save_joints_json).pack(pady=(12, 2))
        ctk.CTkButton(mid, text="Load Joints\u2026", width=130,
                      command=self._load_joints_json).pack(pady=2)

        # Right: existing joints
        right = ctk.CTkFrame(main)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        ctk.CTkLabel(right, text="Joints",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W)
        self._joint_listbox = tk.Listbox(right, selectmode=tk.SINGLE,
                                         exportselection=False,
                                         bg=lb_bg, fg=lb_fg,
                                         selectbackground=lb_sel_bg,
                                         selectforeground="white")
        self._joint_listbox.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self._joint_listbox.bind("<<ListboxSelect>>", self._on_joint_select)
        self._refresh_joint_list()

        reorder = ctk.CTkFrame(right, fg_color="transparent")
        reorder.pack(fill=tk.X, pady=(4, 0))
        ctk.CTkButton(reorder, text="\u25b2 Up", width=60,
                      command=self._move_up).pack(side=tk.LEFT, padx=(0, 4))
        ctk.CTkButton(reorder, text="\u25bc Down", width=60,
                      command=self._move_down).pack(side=tk.LEFT)

        # Bottom buttons
        bottom = ctk.CTkFrame(self, fg_color="transparent")
        bottom.pack(fill=tk.X, padx=10, pady=(0, 10))
        ctk.CTkButton(bottom, text="Apply", width=80,
                      command=self._apply).pack(side=tk.RIGHT, padx=(5, 0))
        ctk.CTkButton(bottom, text="Cancel", width=80,
                      fg_color="gray50",
                      command=self.destroy).pack(side=tk.RIGHT)

    def _refresh_joint_list(self):
        self._joint_listbox.delete(0, tk.END)
        for name, pids in self._joints.items():
            preview = ', '.join(str(p) for p in sorted(pids)[:5])
            if len(pids) > 5:
                preview += f'... ({len(pids)} total)'
            self._joint_listbox.insert(tk.END, f"{name}: {preview}")
        self._update_consumed_styling()

    def _update_consumed_styling(self):
        fg_consumed = "gray50" if self._dark else "gray"
        fg_available = "#dce4ee" if self._dark else "black"
        consumed = set()
        for pids in self._joints.values():
            consumed.update(pids)
        for i, pid in enumerate(self._available_pids):
            if pid in consumed:
                self._pid_listbox.itemconfig(i, fg=fg_consumed)
            else:
                self._pid_listbox.itemconfig(i, fg=fg_available)

    def _on_joint_select(self, event):
        sel = self._joint_listbox.curselection()
        if not sel:
            return
        keys = list(self._joints.keys())
        name = keys[sel[0]]
        self._editing_key = name
        self._name_var.set(name)

        members = self._joints[name]
        self._pid_listbox.selection_clear(0, tk.END)
        first_idx = None
        for i, pid in enumerate(self._available_pids):
            if pid in members:
                self._pid_listbox.selection_set(i)
                if first_idx is None:
                    first_idx = i
        if first_idx is not None:
            self._pid_listbox.see(first_idx)

    def _save_joint(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showwarning("No Name", "Enter a joint name.", parent=self)
            return
        sel = self._pid_listbox.curselection()
        if not sel:
            messagebox.showwarning("No Selection",
                                   "Select properties from the left list.",
                                   parent=self)
            return
        pids = {self._available_pids[i] for i in sel}

        if self._editing_key is None:
            if name in self._joints:
                if not messagebox.askyesno(
                        "Overwrite Joint",
                        f"Joint '{name}' already exists. Overwrite it?",
                        parent=self):
                    return
            self._joints[name] = pids
        else:
            if name == self._editing_key:
                self._joints[name] = pids
            else:
                if name in self._joints and name != self._editing_key:
                    if not messagebox.askyesno(
                            "Overwrite Joint",
                            f"Joint '{name}' already exists. Merge into it?",
                            parent=self):
                        return
                self._joints = {
                    (name if k == self._editing_key else k):
                    (pids if k == self._editing_key else v)
                    for k, v in self._joints.items()
                    if k != name or k == self._editing_key
                }

        self._clear_form()
        self._refresh_joint_list()

    def _clear_form(self):
        self._editing_key = None
        self._name_var.set('')
        self._pid_listbox.selection_clear(0, tk.END)
        self._joint_listbox.selection_clear(0, tk.END)

    def _delete_joint(self):
        sel = self._joint_listbox.curselection()
        if not sel:
            return
        name = list(self._joints.keys())[sel[0]]
        del self._joints[name]
        if self._editing_key == name:
            self._clear_form()
        self._refresh_joint_list()

    def _move_up(self):
        sel = self._joint_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        idx = sel[0]
        keys = list(self._joints.keys())
        keys[idx - 1], keys[idx] = keys[idx], keys[idx - 1]
        self._joints = {k: self._joints[k] for k in keys}
        self._refresh_joint_list()
        self._joint_listbox.selection_set(idx - 1)
        self._on_joint_select(None)

    def _move_down(self):
        sel = self._joint_listbox.curselection()
        if not sel or sel[0] >= len(self._joints) - 1:
            return
        idx = sel[0]
        keys = list(self._joints.keys())
        keys[idx], keys[idx + 1] = keys[idx + 1], keys[idx]
        self._joints = {k: self._joints[k] for k in keys}
        self._refresh_joint_list()
        self._joint_listbox.selection_set(idx + 1)
        self._on_joint_select(None)

    def _import_csv(self):
        import csv
        path = filedialog.askopenfilename(
            title="Import Joints CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            parent=self)
        if not path:
            return
        available_set = set(self._available_pids)
        try:
            with open(path, newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header and header[0].strip().lower() in ('joint name',
                                                             'joint_name',
                                                             'name'):
                    pass
                else:
                    if header:
                        self._process_csv_row(header, available_set)
                for row in reader:
                    self._process_csv_row(row, available_set)
        except Exception as exc:
            messagebox.showerror("Import Error", str(exc), parent=self)
            return
        self._refresh_joint_list()

    def _process_csv_row(self, row, available_set):
        if len(row) < 3:
            return
        name = row[0].strip()
        if not name:
            return
        all_pids = set()
        i = 1
        while i + 1 < len(row):
            try:
                id_start = int(row[i])
                id_end = int(row[i + 1])
            except (ValueError, TypeError):
                i += 2
                continue
            if id_start > id_end:
                id_start, id_end = id_end, id_start
            all_pids.update(j for j in range(id_start, id_end + 1)
                            if j in available_set)
            i += 2
        if not all_pids:
            return
        if name in self._joints:
            self._joints[name].update(all_pids)
        else:
            self._joints[name] = all_pids

    def _export_template(self):
        import csv, os
        path = filedialog.asksaveasfilename(
            title="Export Joints Template",
            defaultextension='.csv',
            filetypes=[("CSV files", "*.csv")],
            parent=self)
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['Joint Name', 'ID Start', 'ID End'])
                for pid in self._available_pids:
                    writer.writerow(['', pid, pid])
            messagebox.showinfo("Exported", f"Template saved to:\n{path}", parent=self)
        except Exception as exc:
            messagebox.showerror("Export Error", str(exc), parent=self)

    def _save_joints_json(self):
        import json, os
        initial_dir = os.path.dirname(self._bdf_path) if self._bdf_path else None
        initial_file = (
            os.path.splitext(os.path.basename(self._bdf_path))[0] + ".cbush_joints.json"
            if self._bdf_path else "joints.cbush_joints.json"
        )
        path = filedialog.asksaveasfilename(
            title="Save Joints",
            defaultextension=".json",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            parent=self)
        if not path:
            return
        payload = {
            "format": "cbush_joints",
            "version": 1,
            "joints": [
                {"name": name, "ids": sorted(pids)}
                for name, pids in self._joints.items()
            ],
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2)
            messagebox.showinfo("Saved", f"Joints saved to:\n{path}", parent=self)
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc), parent=self)

    def _load_joints_json(self):
        import json, os
        initial_dir = os.path.dirname(self._bdf_path) if self._bdf_path else None
        path = filedialog.askopenfilename(
            title="Load Joints",
            initialdir=initial_dir,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            parent=self)
        if not path:
            return
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as exc:
            messagebox.showerror("Load Error", str(exc), parent=self)
            return
        if data.get("format") != "cbush_joints":
            messagebox.showerror("Load Error",
                                 "File is not a cbush_joints JSON.", parent=self)
            return
        available_set = set(self._available_pids)
        new_joints = {}
        skipped_ids = []
        skipped_joints = []
        for entry in data.get("joints", []):
            name = entry.get("name", "").strip()
            if not name:
                continue
            pids = set()
            for pid in entry.get("ids", []):
                if pid in available_set:
                    pids.add(pid)
                else:
                    skipped_ids.append(pid)
            if pids:
                new_joints[name] = pids
            else:
                skipped_joints.append(name)
        if skipped_ids or skipped_joints:
            msg_parts = []
            if skipped_ids:
                preview = ", ".join(str(x) for x in skipped_ids[:5])
                if len(skipped_ids) > 5:
                    preview += f"\u2026 ({len(skipped_ids)} total)"
                msg_parts.append(f"{len(skipped_ids)} PID(s) not found: {preview}")
            if skipped_joints:
                msg_parts.append(
                    f"{len(skipped_joints)} empty joint(s) skipped: "
                    + ", ".join(f"'{g}'" for g in skipped_joints[:3])
                    + ("\u2026" if len(skipped_joints) > 3 else "")
                )
            messagebox.showwarning("Partial Load", "\n".join(msg_parts), parent=self)
        self._joints = new_joints
        self._clear_form()
        self._refresh_joint_list()

    def _apply(self):
        self._on_apply(self._joints)
        self.destroy()


# ---------------------------------------------------------------- GUI module

class CbushForcesModule:
    name = "CBUSH Forces"

    def __init__(self, parent):
        self.frame = ctk.CTkFrame(parent)
        self._op2_path = None
        self._bdf_path = None
        self._title_var = tk.StringVar(value='')
        self._name_var = tk.StringVar(value='')
        self._scale_var = tk.StringVar(value='1.0')
        self._combined_var = tk.BooleanVar(value=False)

        # Per-subcase state
        self._subcase_data = {}    # {sc_id: ndarray(nelems, 6)}
        self._subcase_eids = {}    # {sc_id: ndarray}
        self._subcase_titles = {}  # {sc_id: str}  (OP2 subtitle)
        self._subcase_names = {}   # {sc_id: str}  (user-editable)
        self._subcase_scales = {}  # {sc_id: str}  (scale factor per subcase)
        self._subcase_order = []   # sorted subcase IDs
        self._active_subcase = None

        # BDF mappings
        self._pid_names = {}       # {pid: comment name}
        self._eid_to_pid = {}      # {eid: pid}

        # Axial direction state
        self._axial_var = tk.StringVar(value='X')
        self._eid_axial = {}       # {eid: 'X'|'Y'|'Z'} per-element overrides

        # Joint state (property groups for shear totals)
        self._joints = {}          # {joint_name: set(pid, ...)}
        self._joint_order = []     # display order

        self._build_ui()

    # ------------------------------------------------------------------ UI
    _GUIDE_TEXT = """\
CBUSH Forces Tool -- Quick Guide

PURPOSE
Extract CBUSH element force results (Fx, Fy, Fz, Mx, My, Mz) from a
Nastran OP2 file.  Displays forces per element for each subcase / load
case.  For random analysis the values represent RMS.

WORKFLOW
1. Open OP2 -- select a Nastran OP2 containing CBUSH force output.
   (Requires FORCE(PLOT) = ALL or FORCE = ALL in case control.)
2. Open BDF (optional) -- load the corresponding BDF to add a Property
   column showing comment names from PBUSH cards.  Rows are sorted by
   property group when a BDF is loaded.
3. Select Load Case -- use the dropdown to switch between subcases.
4. Name Load Cases -- type a descriptive name in the Name field.
   Names persist when switching between subcases.
5. Scale Factor -- enter a multiplier per load case (applied on export
   only; the table always shows raw OP2 values).  Noted in the Excel
   header when not 1.0.
6. Export to Excel -- choose separate sheets (one per subcase) or
   combined (all subcases stacked on a single sheet).

AXIAL DIRECTION
The Axial Dir dropdown sets the global axial direction (X, Y, or Z)
for computing derived force quantities:
  Axial      -- force along the axial direction
  Comb Shear -- sqrt of sum-of-squares of the two transverse forces
  Comb Moment-- sqrt of sum-of-squares of the two transverse moments

Per-element overrides: click the Axial Dir cell for any row to change
that element's axial direction independently.  Overrides persist across
subcase switches and are included in the Excel export.

JOINTS
Joints are named groups of properties.  When joints are defined, a
summary table appears to the right showing the total combined shear
summed across all elements belonging to each joint's member properties.
  - Manage Joints -- opens a dialog to create/delete joints (requires BDF)
  - Joint totals update automatically when switching subcases or changing
    axial directions
  - Joint summary is included in the Excel export

TITLE FIELD
Optional title text that appears as a header row in every Excel sheet.
Leave blank to omit.

EXPORT MODES
  Separate sheets (default) -- one worksheet per load case
  All cases on one sheet -- subcases stacked vertically with a blank
    row between blocks

EXCEL OUTPUT
  - Scientific notation for force values (0.000E+00)
  - Frozen header rows for easy scrolling
  - Color-coded header bands
  - Property column included when BDF is loaded
  - Axial Dir, Axial, Comb Shear, Comb Moment columns always present
  - Joint summary table to the right of main data (when joints defined)

REQUIREMENTS
  - pyNastran (for OP2/BDF reading)
  - numpy
  - openpyxl (for Excel export only)\
"""

    def _build_ui(self):
        # Row 1: Open OP2 | Open BDF | Load Case | Name | ? | Export
        row1 = ctk.CTkFrame(self.frame, fg_color="transparent")
        row1.pack(fill=tk.X, padx=5, pady=(5, 0))

        self._op2_btn = ctk.CTkButton(row1, text="Open OP2\u2026", width=100,
                                      command=self._open_op2)
        self._op2_btn.pack(side=tk.LEFT)

        self._bdf_btn = ctk.CTkButton(row1, text="Open BDF\u2026", width=100,
                                      command=self._open_bdf)
        self._bdf_btn.pack(side=tk.LEFT, padx=(5, 0))

        # Load case dropdown
        ctk.CTkLabel(row1, text="Load Case:").pack(
            side=tk.LEFT, padx=(10, 2))
        self._lc_var = tk.StringVar(value="(none)")
        self._lc_menu = ctk.CTkOptionMenu(
            row1, variable=self._lc_var, values=["(none)"],
            command=self._on_lc_select, width=220)
        self._lc_menu.pack(side=tk.LEFT, padx=(0, 4))

        # Per-subcase name entry
        ctk.CTkLabel(row1, text="Name:").pack(side=tk.LEFT, padx=(10, 2))
        name_entry = ctk.CTkEntry(row1, textvariable=self._name_var,
                                  width=180)
        name_entry.pack(side=tk.LEFT, padx=(0, 4))
        self._name_var.trace_add('write', lambda *_: self._on_lc_name_change())

        # Right-side buttons (row 1)
        ctk.CTkButton(
            row1, text="?", width=30, font=ctk.CTkFont(weight="bold"),
            command=self._show_guide,
        ).pack(side=tk.RIGHT, padx=(5, 0))

        ctk.CTkButton(row1, text="Export to Excel\u2026", width=130,
                       command=self._export_excel).pack(side=tk.RIGHT)

        # Row 2: Title | Scale Factor | Combined checkbox
        row2 = ctk.CTkFrame(self.frame, fg_color="transparent")
        row2.pack(fill=tk.X, padx=5, pady=(2, 0))

        ctk.CTkLabel(row2, text="Title:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkEntry(row2, textvariable=self._title_var, width=160).pack(
            side=tk.LEFT, padx=(0, 4))

        ctk.CTkLabel(row2, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        ctk.CTkLabel(row2, text="Scale Factor:").pack(
            side=tk.LEFT, padx=(0, 2))
        self._scale_entry = ctk.CTkEntry(
            row2, textvariable=self._scale_var, width=60)
        self._scale_entry.pack(side=tk.LEFT, padx=(0, 4))
        self._scale_var.trace_add('write', lambda *_: self._on_scale_change())

        ctk.CTkLabel(row2, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        ctk.CTkCheckBox(row2, text="All cases on one sheet",
                        variable=self._combined_var).pack(
            side=tk.LEFT, padx=(0, 4))

        ctk.CTkLabel(row2, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        ctk.CTkLabel(row2, text="Axial Dir:").pack(
            side=tk.LEFT, padx=(0, 2))
        self._axial_menu = ctk.CTkOptionMenu(
            row2, variable=self._axial_var, values=['X', 'Y', 'Z'],
            command=self._on_axial_change, width=60)
        self._axial_menu.pack(side=tk.LEFT, padx=(0, 4))

        ctk.CTkLabel(row2, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        self._joints_btn = ctk.CTkButton(
            row2, text="Manage Joints\u2026", width=120,
            command=self._open_joints_dialog, state=tk.DISABLED)
        self._joints_btn.pack(side=tk.LEFT, padx=(0, 4))

        # Status label
        self._status_label = ctk.CTkLabel(
            self.frame, text="No OP2 loaded", text_color="gray")
        self._status_label.pack(anchor=tk.W, padx=10, pady=(2, 0))

        # Table area: main sheet + joint summary side by side
        table_frame = ctk.CTkFrame(self.frame, fg_color="transparent")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self._sheet = Sheet(
            table_frame,
            headers=_build_headers(False),
            show_top_left=False,
            show_row_index=False,
        )
        self._sheet.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._sheet.disable_bindings()
        self._sheet.enable_bindings(
            "single_select", "drag_select", "row_select",
            "column_select", "copy", "arrowkeys",
            "column_width_resize", "row_height_resize",
        )

        # Joint summary sheet (hidden until joints exist)
        self._joint_sheet = Sheet(
            table_frame,
            headers=['Joint', 'Shear Total'],
            show_top_left=False,
            show_row_index=False,
            width=250,
        )
        self._joint_sheet.disable_bindings()
        self._joint_sheet.enable_bindings(
            "single_select", "copy", "arrowkeys",
            "column_width_resize",
        )
        # Hidden by default — shown when joints exist
        self._joint_sheet_visible = False

    # ---------------------------------------------------------- Guide
    def _show_guide(self):
        try:
            from structures_tools import show_guide
        except ImportError:
            return
        show_guide(self.frame.winfo_toplevel(), "CBUSH Forces Guide",
                   self._GUIDE_TEXT)

    # ---------------------------------------------------------- background work
    def _run_in_background(self, label, work_fn, done_fn):
        """Run *work_fn* in a background thread, keeping the UI responsive."""
        self._status_label.configure(text=label, text_color="gray")
        self._op2_btn.configure(state=tk.DISABLED)
        self._bdf_btn.configure(state=tk.DISABLED)

        container = {}

        def _worker():
            try:
                container['result'] = work_fn()
            except Exception as exc:
                container['error'] = exc

        def _poll():
            if thread.is_alive():
                self.frame.after(50, _poll)
            else:
                self._op2_btn.configure(state=tk.NORMAL)
                self._bdf_btn.configure(state=tk.NORMAL)
                done_fn(container.get('result'), container.get('error'))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()
        self.frame.after(50, _poll)

    # ---------------------------------------------------------- OP2 loading
    def _open_op2(self):
        path = filedialog.askopenfilename(
            title="Open OP2 File",
            filetypes=[("OP2 files", "*.op2"), ("All files", "*.*")])
        if not path:
            return

        def _work():
            from pyNastran.op2.op2 import OP2
            op2 = OP2(mode='nx')
            op2.read_op2(path)
            return op2

        def _done(op2, error):
            if error is not None:
                messagebox.showerror("Error",
                                     f"Could not read OP2:\n{error}")
                self._status_label.configure(text="Load failed",
                                             text_color="red")
                return

            self._op2_path = path
            self._status_label.configure(
                text=os.path.basename(path),
                text_color=("gray10", "gray90"))
            self.load(op2)

        self._run_in_background("Loading\u2026", _work, _done)

    # ---------------------------------------------------------- BDF loading
    def _open_bdf(self):
        path = filedialog.askopenfilename(
            title="Open BDF File",
            filetypes=[("BDF files", "*.bdf *.dat *.nas *.bulk"),
                       ("All files", "*.*")])
        if not path:
            return

        def _work():
            self._build_bdf_mappings(path)

        def _done(_result, error):
            if error is not None:
                messagebox.showerror("Error",
                                     f"Could not read BDF:\n{error}")
                self._status_label.configure(text="BDF load failed",
                                             text_color="red")
                return

            self._bdf_path = path
            self._joints_btn.configure(state=tk.NORMAL)

            status = ""
            if self._op2_path:
                status += f"OP2: {os.path.basename(self._op2_path)}  |  "
            status += (f"BDF: {os.path.basename(path)} "
                       f"({len(self._pid_names)} properties, "
                       f"{len(self._eid_to_pid)} elements)")
            self._status_label.configure(text=status,
                                         text_color=("gray10", "gray90"))

            if self._subcase_data:
                self._show_subcase()

        self._run_in_background("Loading BDF\u2026", _work, _done)

    @staticmethod
    def _extract_comment_name(comment):
        """Extract a descriptive name from a BDF card comment string.

        Takes the last non-empty comment line (directly above the card).
        Strips any prefix before a colon (e.g. ``$ Skin: Wing Upper``
        becomes ``Wing Upper``).
        """
        if not comment:
            return None
        result = None
        for line in comment.splitlines():
            line = line.strip().lstrip('$').strip()
            if line:
                if ':' in line:
                    line = line.split(':', 1)[1].strip()
                if line:
                    result = line
        return result

    def _build_bdf_mappings(self, bdf_path):
        """Build pid_names and eid_to_pid mappings from BDF."""
        from bdf_utils import make_model, read_bdf_safe

        model = make_model()
        read_bdf_safe(model, bdf_path)

        # Extract comment names from property cards
        self._pid_names = {}
        for pid, prop in model.properties.items():
            name = self._extract_comment_name(getattr(prop, 'comment', ''))
            if name:
                self._pid_names[pid] = name

        self._eid_to_pid = {}
        for eid, elem in model.elements.items():
            pid = getattr(elem, 'pid', None)
            if pid is not None:
                try:
                    pid_int = int(pid)
                except (ValueError, TypeError):
                    continue
                if pid_int != 0:
                    self._eid_to_pid[eid] = pid_int

    def _get_prop_name(self, eid):
        """Return property display name for an element ID."""
        pid = self._eid_to_pid.get(eid)
        if pid is None:
            return ""
        return self._pid_names.get(pid, f"PID {pid}")

    def _has_bdf(self):
        """True if BDF mappings are loaded."""
        return bool(self._eid_to_pid)

    # -------------------------------------------------------------- load
    def load(self, op2):
        """Extract CBUSH force data from all subcases."""
        self._subcase_data.clear()
        self._subcase_eids.clear()
        self._subcase_titles.clear()
        self._subcase_names.clear()
        self._subcase_scales.clear()
        self._subcase_order.clear()
        self._active_subcase = None
        self._eid_axial.clear()  # reset per-element overrides on new OP2
        self._sheet.set_sheet_data([])

        if not hasattr(op2, 'cbush_force') or not op2.cbush_force:
            messagebox.showwarning(
                "No CBUSH Forces",
                "No CBUSH element force results found in this OP2.\n\n"
                "Ensure your case control includes:\n"
                "  FORCE(PLOT) = ALL  (or FORCE = ALL)")
            self._status_label.configure(text="No CBUSH forces",
                                         text_color="red")
            return

        for sc_id, result in op2.cbush_force.items():
            # Element IDs -- may be 2D (ntimes, nelems)
            eids = result.element
            if eids.ndim == 2:
                eids = eids[0]
            self._subcase_eids[sc_id] = eids

            # Force data -- first time step: result.data shape (ntimes, nelems, 6)
            data = result.data
            if data.ndim == 3:
                forces = data[0]  # (nelems, 6)
            else:
                forces = data
            self._subcase_data[sc_id] = forces

            subtitle = getattr(result, 'subtitle', '').strip()
            self._subcase_titles[sc_id] = subtitle
            self._subcase_names[sc_id] = ''
            self._subcase_scales[sc_id] = '1.0'

        self._subcase_order = sorted(self._subcase_data.keys())

        if not self._subcase_order:
            return

        # Build dropdown labels
        labels = []
        for sc_id in self._subcase_order:
            sub = self._subcase_titles[sc_id]
            lbl = f"Subcase {sc_id}"
            if sub:
                lbl += f" - {sub}"
            labels.append(lbl)

        self._lc_menu.configure(values=labels)
        self._lc_var.set(labels[0])
        self._active_subcase = self._subcase_order[0]
        self._name_var.set('')
        self._scale_var.set('1.0')
        self._show_subcase()

    # ---------------------------------------------------------- subcase switching
    def _on_lc_select(self, choice):
        """Handle load case dropdown selection."""
        # Save current name and scale
        if self._active_subcase is not None:
            self._subcase_names[self._active_subcase] = self._name_var.get()
            self._subcase_scales[self._active_subcase] = self._scale_var.get()

        # Parse subcase ID from label
        idx = self._lc_menu.cget("values").index(choice)
        self._active_subcase = self._subcase_order[idx]

        # Restore name and scale for new subcase
        self._name_var.set(self._subcase_names.get(self._active_subcase, ''))
        self._scale_var.set(self._subcase_scales.get(self._active_subcase, '1.0'))
        self._show_subcase()

    def _on_lc_name_change(self):
        """Store name entry value for the active subcase."""
        if self._active_subcase is not None:
            self._subcase_names[self._active_subcase] = self._name_var.get()

    def _on_scale_change(self):
        """Store scale entry value for the active subcase."""
        if self._active_subcase is not None:
            self._subcase_scales[self._active_subcase] = self._scale_var.get()

    def _on_axial_change(self, _value=None):
        """Global axial direction dropdown changed — refresh table."""
        if self._subcase_data:
            self._show_subcase()

    def _on_cell_axial_change(self, event=None):
        """Per-cell axial direction dropdown changed."""
        if event is None:
            return
        try:
            r = event[1]  # row index
        except (IndexError, TypeError):
            return

        has_bdf = self._has_bdf()
        axial_col = 1 if has_bdf else 0
        new_dir = self._sheet.get_cell_data(r, axial_col)
        if new_dir not in ('X', 'Y', 'Z'):
            return

        # Determine EID for this row
        eid_col = 2 if has_bdf else 1
        eid = int(self._sheet.get_cell_data(r, eid_col))

        # Store or remove override
        default_dir = self._axial_var.get()
        if new_dir == default_dir:
            self._eid_axial.pop(eid, None)
        else:
            self._eid_axial[eid] = new_dir

        # Recompute derived values for this row
        sc = self._active_subcase
        if sc is None:
            return
        eids = self._subcase_eids[sc]
        forces = self._subcase_data[sc]
        idx = None
        for i in range(len(eids)):
            if int(eids[i]) == eid:
                idx = i
                break
        if idx is None:
            return

        axial, shear, moment = _compute_derived(forces[idx], new_dir)
        force_offset = 3 if has_bdf else 2
        derived_offset = force_offset + 6
        self._sheet.set_cell_data(r, derived_offset, f"{axial:.2E}")
        self._sheet.set_cell_data(r, derived_offset + 1, f"{shear:.2E}")
        self._sheet.set_cell_data(r, derived_offset + 2, f"{moment:.2E}")
        self._sheet.redraw()

        # Refresh joint summary if joints exist
        if self._joints:
            self._refresh_joint_summary()

    def _show_subcase(self):
        """Populate sheet with the active subcase data."""
        sc = self._active_subcase
        if sc is None or sc not in self._subcase_data:
            self._sheet.set_sheet_data([])
            return

        eids = self._subcase_eids[sc]
        forces = self._subcase_data[sc]
        has_bdf = self._has_bdf()
        headers = _build_headers(has_bdf)
        default_dir = self._axial_var.get()

        # Build sorted row indices by EID
        order = sorted(range(len(eids)), key=lambda i: int(eids[i]))

        rows = []
        for i in order:
            eid = int(eids[i])
            axial_dir = self._eid_axial.get(eid, default_dir)
            axial, shear, moment = _compute_derived(forces[i], axial_dir)

            if has_bdf:
                prop = self._get_prop_name(eid)
                row = [prop, axial_dir, eid]
            else:
                row = [axial_dir, eid]
            for j in range(6):
                row.append(f"{forces[i, j]:.2E}")
            row.append(f"{axial:.2E}")
            row.append(f"{shear:.2E}")
            row.append(f"{moment:.2E}")
            rows.append(row)

        self._sheet.headers(headers)
        self._sheet.set_sheet_data(rows)
        ncols = len(headers)
        self._sheet.set_all_column_widths(90)
        if has_bdf:
            self._sheet.column_width(column=0, width=140)  # Property
            self._sheet.column_width(column=1, width=70)   # Axial Dir
            self._sheet.column_width(column=2, width=80)   # EID
        else:
            self._sheet.column_width(column=0, width=70)   # Axial Dir
            self._sheet.column_width(column=1, width=80)   # EID
        self._sheet.align_columns(
            list(range(ncols)), align="center", align_header=True)

        # All columns readonly, then enable Axial Dir dropdown per cell
        self._sheet.readonly_columns(columns=list(range(ncols)))
        axial_col = 1 if has_bdf else 0
        for r in range(len(rows)):
            self._sheet.create_dropdown(
                r, axial_col, values=['X', 'Y', 'Z'],
                set_value=rows[r][axial_col],
            )

        # Bind dropdown edit events
        self._sheet.extra_bindings([
            ("end_edit_cell", self._on_cell_axial_change),
        ])

        # Refresh joint summary
        if self._joints:
            self._refresh_joint_summary()

    # ---------------------------------------------------------- joints

    def _open_joints_dialog(self):
        """Open the Manage Joints dialog."""
        if not self._has_bdf():
            return
        available_pids = sorted(set(self._eid_to_pid.values()))
        id_labels = {}
        for pid in available_pids:
            name = self._pid_names.get(pid)
            if name:
                id_labels[pid] = f"{pid} - {name}"
            else:
                id_labels[pid] = str(pid)
        ManageJointsDialog(
            self.frame.winfo_toplevel(),
            available_pids,
            {n: self._joints[n] for n in self._joint_order if n in self._joints},
            self._on_joints_apply,
            id_labels=id_labels,
            bdf_path=self._bdf_path,
        )

    def _on_joints_apply(self, joints):
        """Callback from ManageJointsDialog."""
        self._joints = {k: set(v) for k, v in joints.items()}
        self._joint_order = list(joints.keys())

        if self._joints:
            if not self._joint_sheet_visible:
                self._joint_sheet.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 0))
                self._joint_sheet_visible = True
            self._refresh_joint_summary()
        else:
            if self._joint_sheet_visible:
                self._joint_sheet.pack_forget()
                self._joint_sheet_visible = False

    def _refresh_joint_summary(self):
        """Recompute joint shear totals for the active subcase."""
        sc = self._active_subcase
        if sc is None or not self._joints:
            self._joint_sheet.set_sheet_data([])
            return

        eids = self._subcase_eids[sc]
        forces = self._subcase_data[sc]
        default_dir = self._axial_var.get()

        # Build eid→force index
        eid_idx = {int(eids[i]): i for i in range(len(eids))}

        rows = []
        for name in self._joint_order:
            if name not in self._joints:
                continue
            pids = self._joints[name]
            shear_total = 0.0
            for eid, pid in self._eid_to_pid.items():
                if pid in pids and eid in eid_idx:
                    axial_dir = self._eid_axial.get(eid, default_dir)
                    _, shear, _ = _compute_derived(
                        forces[eid_idx[eid]], axial_dir)
                    shear_total += shear
            rows.append([name, f"{shear_total:.2E}"])

        self._joint_sheet.headers(['Joint', 'Shear Total'])
        self._joint_sheet.set_sheet_data(rows)
        self._joint_sheet.set_all_column_widths(110)
        self._joint_sheet.align_columns(
            [0, 1], align="center", align_header=True)
        self._joint_sheet.readonly_columns(columns=[0, 1])

    # ------------------------------------------------------------ export helpers

    def _get_scale_factor(self, sc_id):
        """Parse and return scale factor for a subcase, defaulting to 1.0."""
        try:
            return float(self._subcase_scales.get(sc_id, '1.0'))
        except (ValueError, TypeError):
            return 1.0

    def _prepare_export_data(self, sc_id):
        """Build sorted eids, forces, prop_names, axial_dirs, derived, scale, label.

        Returns (eids, forces, prop_names_or_None, axial_dirs, derived, scale, lc_label).
        """
        eids = self._subcase_eids[sc_id]
        forces = self._subcase_data[sc_id]
        has_bdf = self._has_bdf()
        scale = self._get_scale_factor(sc_id)
        default_dir = self._axial_var.get()

        # Sort by EID
        order = sorted(range(len(eids)), key=lambda i: int(eids[i]))
        eids = eids[order]
        forces = forces[order]

        if has_bdf:
            prop_names = [self._get_prop_name(int(eids[i]))
                          for i in range(len(eids))]
        else:
            prop_names = None

        # Axial dirs and derived values
        axial_dirs = []
        derived = []
        for i in range(len(eids)):
            eid = int(eids[i])
            ad = self._eid_axial.get(eid, default_dir)
            axial_dirs.append(ad)
            derived.append(_compute_derived(forces[i], ad))

        # Build label
        user_name = self._subcase_names.get(sc_id, '').strip()
        effective_name = user_name or self._subcase_titles.get(sc_id, '')
        lc_label = f"Subcase {sc_id}"
        if effective_name:
            lc_label += f" \u2014 {effective_name}"
        if scale != 1.0:
            lc_label += f" (SF: {scale:g})"

        return eids, forces, prop_names, axial_dirs, derived, scale, lc_label

    def _make_sheet_name(self, sc_id, used_names):
        """Build a unique 31-char Excel sheet name."""
        user_name = self._subcase_names.get(sc_id, '').strip()
        effective_name = user_name or self._subcase_titles.get(sc_id, '')
        if effective_name:
            base = effective_name[:31]
        else:
            base = f"Subcase {sc_id}"[:31]

        sheet_name = base
        counter = 2
        while sheet_name.lower() in used_names:
            suffix = f" ({counter})"
            sheet_name = base[:31 - len(suffix)] + suffix
            counter += 1
        used_names.add(sheet_name.lower())
        return sheet_name

    # ------------------------------------------------------------ export
    def _compute_joint_data(self, sc_id, scale):
        """Compute joint summary data for export.

        Returns list of (joint_name, scaled_shear_total) or empty list.
        """
        if not self._joints or not self._has_bdf():
            return []

        eids = self._subcase_eids[sc_id]
        forces = self._subcase_data[sc_id]
        default_dir = self._axial_var.get()
        eid_idx = {int(eids[i]): i for i in range(len(eids))}

        result = []
        for name in self._joint_order:
            if name not in self._joints:
                continue
            pids = self._joints[name]
            shear_total = 0.0
            for eid, pid in self._eid_to_pid.items():
                if pid in pids and eid in eid_idx:
                    ad = self._eid_axial.get(eid, default_dir)
                    _, shear, _ = _compute_derived(forces[eid_idx[eid]], ad)
                    shear_total += shear
            result.append((name, shear_total * scale))
        return result

    def _export_excel(self):
        if not self._subcase_data:
            messagebox.showinfo("Nothing to export",
                                "Open an OP2 file first.")
            return

        # Save current name and scale before export
        if self._active_subcase is not None:
            self._subcase_names[self._active_subcase] = self._name_var.get()
            self._subcase_scales[self._active_subcase] = self._scale_var.get()

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return

        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for Excel export.\n\n"
                "pip install openpyxl")
            return

        op2_name = os.path.basename(self._op2_path) if self._op2_path else None
        title = self._title_var.get().strip() or None
        styles = make_cbush_styles()
        combined = self._combined_var.get()
        has_bdf = self._has_bdf()
        headers = _build_headers(has_bdf)
        joint_col = len(headers) + 2  # gap column + start

        wb = Workbook()

        if combined:
            # All subcases on one sheet
            ws = wb.active
            ws.title = "CBUSH Forces"
            cur_row = 0

            for i, sc_id in enumerate(self._subcase_order):
                eids, forces, prop_names, axial_dirs, derived, scale, \
                    lc_label = self._prepare_export_data(sc_id)

                if i == 0:
                    cur_row, header_row = write_cbush_sheet(
                        ws, eids, forces, styles,
                        op2_name=op2_name, title=title,
                        sheet_label=lc_label, prop_names=prop_names,
                        scale_factor=scale, start_row=cur_row,
                        axial_dirs=axial_dirs, derived=derived)
                else:
                    cur_row += 1  # blank row
                    cur_row, header_row = _write_cbush_block_combined(
                        ws, eids, forces, styles,
                        sheet_label=lc_label, prop_names=prop_names,
                        scale_factor=scale, start_row=cur_row,
                        axial_dirs=axial_dirs, derived=derived)

                # Joint summary per block
                jdata = self._compute_joint_data(sc_id, scale)
                if jdata:
                    _write_joint_summary_block(
                        ws, jdata, styles, header_row, joint_col)

            n = len(self._subcase_order)
            msg = f"Saved {n} case(s) to 1 sheet:\n{path}"
        else:
            # Separate sheets
            used_names = set()

            for i, sc_id in enumerate(self._subcase_order):
                eids, forces, prop_names, axial_dirs, derived, scale, \
                    lc_label = self._prepare_export_data(sc_id)
                sheet_name = self._make_sheet_name(sc_id, used_names)

                if i == 0:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)

                _next_row, header_row = write_cbush_sheet(
                    ws, eids, forces, styles,
                    op2_name=op2_name, title=title,
                    sheet_label=lc_label, prop_names=prop_names,
                    scale_factor=scale,
                    axial_dirs=axial_dirs, derived=derived,
                )

                # Joint summary
                jdata = self._compute_joint_data(sc_id, scale)
                if jdata:
                    _write_joint_summary_block(
                        ws, jdata, styles, header_row, joint_col)

            n = len(self._subcase_order)
            msg = f"Saved {n} sheet(s) to:\n{path}"

        try:
            wb.save(path)
            _ExportDoneDialog(
                self.frame.winfo_toplevel(), msg, path)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))


def main():
    import logging
    logging.getLogger("customtkinter").setLevel(logging.ERROR)
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    root.title("CBUSH Forces")
    root.geometry("1000x600")
    mod = CbushForcesModule(root)
    mod.frame.pack(fill=tk.BOTH, expand=True)
    root.mainloop()


if __name__ == '__main__':
    main()
