"""Element strain energy breakdown module.

Reads element strain energy (ESE%) from an OP2 file and displays
per-mode percentages grouped by include file, property ID, or material ID.
"""
import os
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from tksheet import Sheet

import numpy as np
if not hasattr(np, 'in1d'):
    np.in1d = np.isin


# --------------------------------------------------------- Excel helpers

def make_energy_styles():
    """Return a dict of openpyxl style objects for Energy Breakdown sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        'dark_fill': PatternFill("solid", fgColor="1F4E79"),
        'mid_fill': PatternFill("solid", fgColor="2E75B6"),
        'white_bold': Font(bold=True, color="FFFFFF", size=11),
        'sub_font': Font(bold=True, color="FFFFFF", size=10),
        'center': Alignment(horizontal="center", vertical="center"),
        'right': Alignment(horizontal="right", vertical="center"),
        'cell_border': Border(bottom=Side(style='thin', color="B4C6E7")),
        'bold_font': Font(bold=True),
        'red_font': Font(color="FF0000"),
        'bold_red_font': Font(bold=True, color="FF0000"),
        'num1': '0.0',
        'num2': '0',
    }


def write_energy_sheet(ws, data, styles, op2_name=None, threshold=5.0,
                       title=None):
    """Write an energy breakdown sheet to an openpyxl worksheet.

    data: dict with keys 'modes', 'freqs', 'headers', 'table' (list of rows).
    """
    from openpyxl.utils import get_column_letter

    s = styles
    headers = data['headers']
    table = data['table']
    total_cols = len(headers)

    row_offset = 1 if title else 0

    # Row 1: custom title (only when title provided)
    if title:
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=total_cols)
        for ci in range(2, total_cols + 1):
            ws.cell(row=1, column=ci).fill = s['dark_fill']

    # OP2 filename row
    name_row = 1 + row_offset
    name_text = op2_name if op2_name else ""
    cell = ws.cell(row=name_row, column=1, value=name_text)
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=name_row, start_column=1,
                   end_row=name_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=name_row, column=ci).fill = s['dark_fill']

    # Sub-headers row
    sub_row = 2 + row_offset
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=sub_row, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data rows
    data_start = 3 + row_offset
    total_col_idx = total_cols - 1  # 0-based index of Total column
    for i, row_data in enumerate(table):
        row = i + data_start
        for ci, val in enumerate(row_data):
            cell = ws.cell(row=row, column=ci + 1, value=val)
            cell.alignment = s['center']
            if ci == 0:
                pass  # Mode number
            elif ci == 1:
                cell.number_format = s['num1']
            else:
                cell.number_format = s['num2']
                # Bold for values >= threshold
                if isinstance(val, (int, float)) and ci < total_col_idx:
                    if val >= threshold:
                        cell.font = s['bold_font']
                # Red for total deviating from 100%
                if ci == total_col_idx and isinstance(val, (int, float)):
                    if abs(val - 100.0) > 0.5:
                        cell.font = s['bold_red_font'] if val >= threshold else s['red_font']
            cell.border = s['cell_border']

    # Column widths
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 12
    for ci in range(3, total_cols + 1):
        col_letter = get_column_letter(ci)
        header_len = len(str(headers[ci - 1]))
        ws.column_dimensions[col_letter].width = max(10, header_len + 3)
    freeze_row = data_start
    ws.freeze_panes = f'A{freeze_row}'


# --------------------------------------------------------- Manage Groups Dialog

class ManageGroupsDialog(ctk.CTkToplevel):
    """Dialog for creating/editing custom group merges."""

    def __init__(self, parent, available_ids, existing_groups, show_ungrouped,
                 on_apply):
        super().__init__(parent)
        self.title("Manage Groups")
        self.geometry("600x450")
        self.resizable(True, True)
        self.transient(parent)

        self._available_ids = sorted(available_ids)
        self._groups = {k: set(v) for k, v in existing_groups.items()}
        self._show_ungrouped = tk.BooleanVar(value=show_ungrouped)
        self._on_apply = on_apply

        self._build_ui()

    def _build_ui(self):
        # Main paned layout
        main = ctk.CTkFrame(self)
        main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left: available IDs
        left = ctk.CTkFrame(main)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        ctk.CTkLabel(left, text="Available IDs",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W)

        self._id_listbox = tk.Listbox(left, selectmode=tk.EXTENDED,
                                      exportselection=False)
        self._id_listbox.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        for id_val in self._available_ids:
            self._id_listbox.insert(tk.END, str(id_val))
        # Initial consumed-ID styling applied after _refresh_group_list below

        # Middle: controls
        mid = ctk.CTkFrame(main, width=160)
        mid.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        mid.pack_propagate(False)

        ctk.CTkLabel(mid, text="Group Name:").pack(anchor=tk.W, pady=(20, 2))
        self._name_var = tk.StringVar()
        ctk.CTkEntry(mid, textvariable=self._name_var, width=140).pack()

        ctk.CTkButton(mid, text="Create Group \u2192", width=140,
                      command=self._create_group).pack(pady=(10, 2))
        ctk.CTkButton(mid, text="Delete Group", width=140,
                      fg_color="firebrick",
                      command=self._delete_group).pack(pady=2)

        ctk.CTkCheckBox(
            mid, text="Show ungrouped\nas individual",
            variable=self._show_ungrouped,
        ).pack(pady=(20, 0))

        # Right: existing groups
        right = ctk.CTkFrame(main)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

        ctk.CTkLabel(right, text="Custom Groups",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W)

        self._group_listbox = tk.Listbox(right, selectmode=tk.SINGLE,
                                         exportselection=False)
        self._group_listbox.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self._refresh_group_list()

        # Bottom buttons
        bottom = ctk.CTkFrame(self, fg_color="transparent")
        bottom.pack(fill=tk.X, padx=10, pady=(0, 10))

        ctk.CTkButton(bottom, text="Apply", width=80,
                      command=self._apply).pack(side=tk.RIGHT, padx=(5, 0))
        ctk.CTkButton(bottom, text="Cancel", width=80,
                      fg_color="gray50",
                      command=self.destroy).pack(side=tk.RIGHT)

    def _refresh_group_list(self):
        self._group_listbox.delete(0, tk.END)
        for name, ids in self._groups.items():
            id_preview = ', '.join(str(x) for x in sorted(ids)[:5])
            if len(ids) > 5:
                id_preview += f'... ({len(ids)} total)'
            self._group_listbox.insert(tk.END, f"{name}: {id_preview}")
        self._update_consumed_styling()

    def _update_consumed_styling(self):
        """Grey out IDs in the available list that are already in a group."""
        consumed = set()
        for ids in self._groups.values():
            consumed.update(ids)
        for i, id_val in enumerate(self._available_ids):
            if id_val in consumed:
                self._id_listbox.itemconfig(i, fg="gray")
            else:
                self._id_listbox.itemconfig(i, fg="black")

    def _create_group(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showwarning("No Name", "Enter a group name.",
                                   parent=self)
            return

        sel = self._id_listbox.curselection()
        if not sel:
            messagebox.showwarning("No Selection",
                                   "Select IDs from the left list.",
                                   parent=self)
            return

        ids = set()
        for idx in sel:
            ids.add(self._available_ids[idx])
        self._groups[name] = ids
        self._name_var.set('')
        self._refresh_group_list()

    def _delete_group(self):
        sel = self._group_listbox.curselection()
        if not sel:
            return
        group_names = list(self._groups.keys())
        name = group_names[sel[0]]
        del self._groups[name]
        self._refresh_group_list()

    def _apply(self):
        self._on_apply(self._groups, self._show_ungrouped.get())
        self.destroy()


# ---------------------------------------------------------------- GUI module

class EnergyBreakdownModule:
    name = "ESE Breakdown"

    def __init__(self, parent):
        self.frame = ctk.CTkFrame(parent)
        self._op2_path = None
        self._bdf_path = None
        self._threshold_var = tk.StringVar(value='5.0')
        self._title_var = tk.StringVar(value='')
        self._group_by_var = tk.StringVar(value='Property ID')

        # Raw data from OP2
        self._modes = None        # array of mode numbers
        self._freqs = None        # array of frequencies
        self._ese_by_eid = None   # {eid: array[nmodes] of percent}

        # Mappings from BDF
        self._eid_to_pid = {}
        self._eid_to_mid = {}
        self._eid_to_file = {}
        self._pid_names = {}      # {pid: comment name}
        self._mid_names = {}      # {mid: comment name}
        self._file_order = []     # include files in BDF encounter order
        self._bdf_loaded = False

        # Custom group merges
        self._custom_groups = {}       # {name: set(ids)}
        self._show_ungrouped = True

        self._build_ui()

    # ------------------------------------------------------------------ UI
    _GUIDE_TEXT = """\
ESE Breakdown Tool — Quick Guide

PURPOSE
Display element strain energy percentages (ESE%) from a Nastran OP2 file,
broken down by include file, property ID, or material ID. Useful for modal
analysis (SOL 103) to understand energy distribution per mode.

WORKFLOW
1. Open OP2 — select an OP2 file containing strain energy data.
   (Requires ESE(PLOT) = ALL in your case control deck.)
2. Open BDF — load the corresponding BDF to enable grouping by
   property ID, material ID, or include file.
3. Select grouping — use the "Group by" dropdown to choose how
   elements are aggregated.
4. Review — the table shows modes as rows and groups as columns,
   with a Total column that should sum to ~100%.
5. Manage Groups — combine multiple IDs into named groups.
6. Export to Excel — save the table as a formatted .xlsx workbook.

GROUPING MODES
  Property ID — group by element property ID (PID)
  Material ID — group by material ID (MID), resolved through properties
  Include File — group by source BDF include file

MANAGE GROUPS
Opens a dialog to combine multiple IDs/files into named groups.
Select IDs from the left list, enter a group name, and create.
Use "Show ungrouped as individual" to control whether unmapped
IDs appear as separate columns or are lumped into "Other".

THRESHOLD
Values >= threshold are displayed in bold/blue. Default is 5.0%.
The Total column highlights red when deviating from 100% by >0.5%.

EDGE CASES
  - CONROD elements have no PID — shown as "CONROD (no PID)" group
  - Elements in OP2 but not in BDF — shown as "Unmapped"
  - DMIG strain energy entries are skipped (non-numeric element IDs)

REQUIREMENTS
  - pyNastran (for OP2/BDF reading)
  - numpy
  - openpyxl (for Excel export only)\
"""

    def _build_ui(self):
        toolbar = ctk.CTkFrame(self.frame, fg_color="transparent")
        toolbar.pack(fill=tk.X, padx=5, pady=(5, 0))

        ctk.CTkButton(toolbar, text="Open OP2\u2026", width=100,
                      command=self._open_op2).pack(side=tk.LEFT)

        ctk.CTkButton(toolbar, text="Open BDF\u2026", width=100,
                      command=self._open_bdf).pack(side=tk.LEFT, padx=(5, 0))

        # Separator
        ctk.CTkLabel(toolbar, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        # Group by dropdown
        ctk.CTkLabel(toolbar, text="Group by:").pack(side=tk.LEFT, padx=(0, 2))
        self._group_by_menu = ctk.CTkOptionMenu(
            toolbar, variable=self._group_by_var,
            values=["Property ID", "Material ID", "Include File"],
            width=140, command=self._on_group_by_change,
        )
        self._group_by_menu.pack(side=tk.LEFT)
        self._group_by_menu.configure(state=tk.DISABLED)

        # Manage Groups button
        self._manage_btn = ctk.CTkButton(
            toolbar, text="Manage Groups\u2026", width=120,
            command=self._manage_groups)
        self._manage_btn.pack(side=tk.LEFT, padx=(5, 0))
        self._manage_btn.configure(state=tk.DISABLED)

        # Separator
        ctk.CTkLabel(toolbar, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        # Title field
        ctk.CTkLabel(toolbar, text="Title:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkEntry(toolbar, textvariable=self._title_var, width=160).pack(
            side=tk.LEFT, padx=(0, 4))

        # Right side: ?, Export, Threshold
        ctk.CTkButton(
            toolbar, text="?", width=30, font=ctk.CTkFont(weight="bold"),
            command=self._show_guide,
        ).pack(side=tk.RIGHT, padx=(5, 0))

        ctk.CTkButton(toolbar, text="Export to Excel\u2026", width=130,
                      command=self._export_excel).pack(side=tk.RIGHT)

        ctk.CTkEntry(toolbar, width=50,
                     textvariable=self._threshold_var).pack(
            side=tk.RIGHT, padx=(0, 4))
        ctk.CTkLabel(toolbar, text="Threshold:").pack(
            side=tk.RIGHT, padx=(10, 2))

        self._threshold_var.trace_add('write', self._on_threshold_change)

        # Status label
        self._status_label = ctk.CTkLabel(
            self.frame, text="No OP2 loaded", text_color="gray",
            anchor=tk.W)
        self._status_label.pack(fill=tk.X, padx=10, pady=(2, 0))

        # Table (tksheet)
        self._sheet = Sheet(
            self.frame,
            headers=["Mode", "Freq (Hz)", "Total"],
            show_top_left=False,
            show_row_index=False,
        )
        self._sheet.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._sheet.disable_bindings()
        self._sheet.enable_bindings(
            "single_select", "copy", "arrowkeys",
            "column_width_resize", "row_height_resize",
        )

    # ---------------------------------------------------------- Guide
    def _show_guide(self):
        try:
            from nastran_tools import show_guide
        except ImportError:
            return
        show_guide(self.frame.winfo_toplevel(), "Energy Breakdown Guide",
                   self._GUIDE_TEXT)

    # ---------------------------------------------------------- threshold
    def _get_threshold(self):
        try:
            return float(self._threshold_var.get())
        except (ValueError, tk.TclError):
            return 5.0

    def _on_threshold_change(self, *args):
        if self._ese_by_eid is not None:
            self._refresh_table()

    def _on_group_by_change(self, *args):
        # Clear custom groups when switching grouping type
        self._custom_groups = {}
        self._show_ungrouped = True
        if self._ese_by_eid is not None:
            self._refresh_table()

    # ---------------------------------------------------------- OP2 loading
    def _open_op2(self):
        path = filedialog.askopenfilename(
            title="Open OP2 File",
            filetypes=[("OP2 files", "*.op2"), ("All files", "*.*")])
        if not path:
            return

        self._status_label.configure(text="Loading OP2\u2026",
                                     text_color="gray")
        self.frame.update_idletasks()

        try:
            from pyNastran.op2.op2 import OP2
            op2 = OP2(mode='nx')
            op2.read_op2(path)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not read OP2:\n{exc}")
            self._status_label.configure(text="Load failed",
                                         text_color="red")
            return

        self._op2_path = path

        # Extract eigenvalues
        if not op2.eigenvalues:
            messagebox.showwarning(
                "No Eigenvalues",
                "No eigenvalue data found in this OP2.")
            self._status_label.configure(text="No eigenvalues",
                                         text_color="red")
            return

        eigval_table = next(iter(op2.eigenvalues.values()))
        self._modes = np.array(eigval_table.mode)
        self._freqs = np.array(eigval_table.cycles)

        # Collect strain energy
        self._ese_by_eid = self._collect_strain_energy(op2)
        if not self._ese_by_eid:
            messagebox.showwarning(
                "No Strain Energy Data",
                "No element strain energy data found in this OP2.\n\n"
                "Add to your Nastran case control:\n"
                "  ESE(PLOT) = ALL")
            self._status_label.configure(text="No ESE data",
                                         text_color="red")
            return

        n_elems = len(self._ese_by_eid)
        status = f"OP2: {os.path.basename(path)} ({n_elems} elements)"
        if self._bdf_path:
            status += f"  |  BDF: {os.path.basename(self._bdf_path)}"
        self._status_label.configure(text=status,
                                     text_color=("gray10", "gray90"))

        self._refresh_table()

    def _collect_strain_energy(self, op2):
        """Discover all *_strain_energy attributes and collect ESE% per element.

        Returns {eid: array[nmodes]} of percent values.
        """
        nmodes = len(self._modes)
        ese_by_eid = {}

        # Modern pyNastran (1.4+): data lives in op2.op2_results.strain_energy
        se = getattr(getattr(op2, 'op2_results', None), 'strain_energy', None)
        if se is None:
            se = op2  # fallback for older pyNastran

        for attr_name in dir(se):
            if not attr_name.endswith('_strain_energy'):
                continue
            result_dict = getattr(se, attr_name, None)
            if not isinstance(result_dict, dict) or not result_dict:
                continue

            for subcase_id, result in result_dict.items():
                if not hasattr(result, 'data') or not hasattr(result, 'element'):
                    continue

                eids = result.element
                # element array is 2D (ntimes, nelems) in pyNastran 1.4+
                # Use first time step — element IDs are the same across all modes
                if eids.ndim == 2:
                    eids = eids[0]
                # data shape: (nmodes, nelems, ncols)
                # column index 1 = percent of total
                data = result.data
                if data.ndim != 3 or data.shape[2] < 2:
                    continue

                percent_data = data[:, :, 1]  # (nmodes, nelems)

                for j, eid in enumerate(eids):
                    # Skip DMIG sentinel entries (eid=100000000 or non-int)
                    try:
                        eid_int = int(eid)
                    except (ValueError, TypeError):
                        continue
                    if eid_int >= 100000000:
                        continue

                    pct = percent_data[:, j]
                    n = min(len(pct), nmodes)
                    if eid_int in ese_by_eid:
                        # Sum contributions if element appears in multiple tables
                        ese_by_eid[eid_int][:n] += pct[:n]
                    else:
                        arr = np.zeros(nmodes)
                        arr[:n] = pct[:n]
                        ese_by_eid[eid_int] = arr

                break  # Only process first valid result per element type

        return ese_by_eid

    # ---------------------------------------------------------- BDF loading
    def _open_bdf(self):
        path = filedialog.askopenfilename(
            title="Open BDF File",
            filetypes=[("BDF files", "*.bdf *.dat *.nas *.bulk"),
                       ("All files", "*.*")])
        if not path:
            return

        self._status_label.configure(text="Loading BDF\u2026",
                                     text_color="gray")
        self.frame.update_idletasks()

        try:
            self._build_mappings(path)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not read BDF:\n{exc}")
            self._status_label.configure(text="BDF load failed",
                                         text_color="red")
            return

        self._bdf_path = path
        self._bdf_loaded = True

        # Enable grouping controls
        self._group_by_menu.configure(state=tk.NORMAL)
        self._manage_btn.configure(state=tk.NORMAL)

        # Update status
        n_mapped = len(self._eid_to_pid)
        status = ""
        if self._op2_path:
            n_elems = len(self._ese_by_eid) if self._ese_by_eid else 0
            status += f"OP2: {os.path.basename(self._op2_path)} ({n_elems} elements)  |  "
        status += f"BDF: {os.path.basename(path)} ({n_mapped} elements)"

        # Check for unmapped elements
        if self._ese_by_eid:
            unmapped = set(self._ese_by_eid.keys()) - set(self._eid_to_pid.keys())
            if unmapped:
                status += f"  [{len(unmapped)} unmapped]"

        self._status_label.configure(text=status,
                                     text_color=("gray10", "gray90"))

        # Clear custom groups when loading new BDF
        self._custom_groups = {}
        self._show_ungrouped = True

        if self._ese_by_eid is not None:
            self._refresh_table()

    @staticmethod
    def _extract_comment_name(comment):
        """Extract a descriptive name from a BDF card comment string.

        pyNastran stores ``$ Wing Skin\\n`` style comments on cards.
        Returns the first non-empty line with the leading ``$`` stripped,
        or *None* if there is nothing useful.
        """
        if not comment:
            return None
        for line in comment.splitlines():
            line = line.strip().lstrip('$').strip()
            if line:
                return line
        return None

    def _build_mappings(self, bdf_path):
        """Build eid→pid, eid→mid, eid→file mappings from BDF."""
        from bdf_utils import IncludeFileParser, make_model

        # Parse include structure for file mapping
        parser = IncludeFileParser()
        parser.parse(bdf_path)

        # Build eid→file mapping
        self._eid_to_file = {}
        main_dir = os.path.dirname(os.path.abspath(bdf_path))
        for filepath, ids_by_type in parser.file_ids.items():
            eids = ids_by_type.get('eid', set())
            # Use relative path from BDF directory for shorter labels
            try:
                rel = os.path.relpath(filepath, main_dir)
            except ValueError:
                rel = os.path.basename(filepath)
            for eid in eids:
                self._eid_to_file[eid] = rel

        # Save include file encounter order (relative paths)
        self._file_order = []
        for fp in parser.all_files:
            try:
                rel = os.path.relpath(fp, main_dir)
            except ValueError:
                rel = os.path.basename(fp)
            self._file_order.append(rel)

        # Read BDF model for PID/MID mappings
        model = make_model()
        model.read_bdf(bdf_path)

        # Extract comment names from property and material cards
        self._pid_names = {}
        for pid, prop in model.properties.items():
            name = self._extract_comment_name(getattr(prop, 'comment', ''))
            if name:
                self._pid_names[pid] = name

        self._mid_names = {}
        for mid, mat in model.materials.items():
            name = self._extract_comment_name(getattr(mat, 'comment', ''))
            if name:
                self._mid_names[mid] = name

        self._eid_to_pid = {}
        self._eid_to_mid = {}

        for eid, elem in model.elements.items():
            # PID mapping
            pid = getattr(elem, 'pid', None)
            if pid is not None:
                try:
                    self._eid_to_pid[eid] = int(pid)
                except (ValueError, TypeError):
                    self._eid_to_pid[eid] = pid
            elif elem.type == 'CONROD':
                self._eid_to_pid[eid] = 'CONROD (no PID)'

            # MID mapping
            mid = None
            if elem.type == 'CONROD':
                # CONROD has mid directly on the element
                mid = getattr(elem, 'mid', None)
            elif pid is not None:
                # Resolve through property
                try:
                    pid_int = int(pid)
                    prop = model.properties.get(pid_int)
                    if prop is not None:
                        if hasattr(prop, 'Mid'):
                            mid = prop.Mid()
                        elif hasattr(prop, 'mid'):
                            mid = prop.mid
                except (ValueError, TypeError, AttributeError):
                    pass

            if mid is not None:
                try:
                    self._eid_to_mid[eid] = int(mid)
                except (ValueError, TypeError):
                    self._eid_to_mid[eid] = mid

        # Also include mass elements that may not have PID/MID
        for eid, elem in model.masses.items():
            if eid not in self._eid_to_pid:
                pid = getattr(elem, 'pid', None)
                if pid is not None:
                    try:
                        self._eid_to_pid[eid] = int(pid)
                    except (ValueError, TypeError):
                        self._eid_to_pid[eid] = pid

    # ---------------------------------------------------------- aggregation
    def _aggregate_by_group(self):
        """Aggregate ESE% by the selected grouping type.

        Returns (group_labels, group_data) where:
          group_labels: list of group label strings
          group_data: dict {label: array[nmodes]}
        """
        if self._ese_by_eid is None:
            return [], {}

        nmodes = len(self._modes)
        group_by = self._group_by_var.get()

        # Select the appropriate mapping
        if group_by == "Property ID" and self._bdf_loaded:
            mapping = self._eid_to_pid
            pid_names = self._pid_names
            label_fn = lambda gid: (f"PID {gid} \u2014 {pid_names[gid]}"
                                    if gid in pid_names else f"PID {gid}")
        elif group_by == "Material ID" and self._bdf_loaded:
            mapping = self._eid_to_mid
            mid_names = self._mid_names
            label_fn = lambda gid: (f"MID {gid} \u2014 {mid_names[gid]}"
                                    if gid in mid_names else f"MID {gid}")
        elif group_by == "Include File" and self._bdf_loaded:
            mapping = self._eid_to_file
            label_fn = lambda gid: str(gid)
        else:
            # No BDF — flat ungrouped view
            mapping = None
            label_fn = None

        if mapping is None:
            # No grouping: show total only
            total = np.zeros(nmodes)
            for pct in self._ese_by_eid.values():
                total += pct
            return ['All Elements'], {'All Elements': total}

        # Group elements
        raw_groups = {}   # {group_id: array[nmodes]}
        unmapped = np.zeros(nmodes)
        has_unmapped = False

        for eid, pct in self._ese_by_eid.items():
            gid = mapping.get(eid)
            if gid is None:
                unmapped += pct
                has_unmapped = True
            else:
                if gid not in raw_groups:
                    raw_groups[gid] = np.zeros(nmodes)
                raw_groups[gid] += pct

        # Apply custom group merges
        if self._custom_groups:
            merged_groups = {}
            consumed_ids = set()

            for group_name, member_ids in self._custom_groups.items():
                merged = np.zeros(nmodes)
                for gid in member_ids:
                    if gid in raw_groups:
                        merged += raw_groups[gid]
                        consumed_ids.add(gid)
                merged_groups[group_name] = merged

            # Handle ungrouped IDs
            remaining = {gid: arr for gid, arr in raw_groups.items()
                         if gid not in consumed_ids}

            if self._show_ungrouped:
                # Show remaining as individual columns
                for gid, arr in remaining.items():
                    merged_groups[label_fn(gid)] = arr
            else:
                # Lump remaining into "Other"
                if remaining:
                    other = np.zeros(nmodes)
                    for arr in remaining.values():
                        other += arr
                    merged_groups['Other'] = other

            final_groups = merged_groups
        else:
            # No custom groups — label each raw group
            final_groups = {label_fn(gid): arr
                            for gid, arr in raw_groups.items()}

        if has_unmapped:
            final_groups['Unmapped'] = unmapped

        # Sort labels
        if group_by == "Include File" and self._file_order:
            # Preserve BDF encounter order for include files
            order_map = {f: i for i, f in enumerate(self._file_order)}
            labels = sorted(final_groups.keys(),
                            key=lambda lbl: order_map.get(lbl, 999999))
        else:
            labels = sorted(final_groups.keys(), key=self._group_sort_key)

        return labels, final_groups

    @staticmethod
    def _group_sort_key(label):
        """Sort key: numeric groups by number, text groups alphabetically."""
        # Try to extract number from "PID 123" or "PID 123 — Name" patterns
        parts = label.split()
        if len(parts) >= 2 and parts[0] in ('PID', 'MID'):
            try:
                return (0, int(parts[1]), '')
            except ValueError:
                pass
        # "Unmapped" and "Other" sort last
        if label in ('Unmapped', 'Other'):
            return (2, 0, label)
        # Custom group names or filenames sort in the middle
        return (1, 0, label)

    # ---------------------------------------------------------- display
    def _refresh_table(self):
        """Rebuild the table from current data and grouping settings."""
        if self._ese_by_eid is None:
            return

        labels, group_data = self._aggregate_by_group()
        nmodes = len(self._modes)

        # Build headers
        headers = ['Mode', 'Freq (Hz)'] + list(labels) + ['Total']

        # Build table data
        table_data = []
        for i in range(nmodes):
            row = [int(self._modes[i]), f"{self._freqs[i]:.1f}"]
            total = 0.0
            for lbl in labels:
                val = group_data[lbl][i]
                total += val
                row.append(f"{val:.0f}")
            row.append(f"{total:.0f}")
            table_data.append(row)

        # Update sheet
        self._sheet.headers(headers)
        self._sheet.set_sheet_data(table_data)
        self._sheet.readonly_columns(columns=list(range(len(headers))))
        self._sheet.align_columns(
            list(range(len(headers))), align="center", align_header=True)

        self._apply_highlights()

    def _apply_highlights(self):
        """Apply threshold and total-deviation highlighting."""
        self._sheet.dehighlight_all(redraw=False)

        if self._ese_by_eid is None:
            return

        threshold = self._get_threshold()
        labels, group_data = self._aggregate_by_group()
        nmodes = len(self._modes)
        n_groups = len(labels)
        total_col = 2 + n_groups  # 0-based index of Total column

        for i in range(nmodes):
            # Highlight group columns above threshold
            for j, lbl in enumerate(labels):
                val = group_data[lbl][i]
                if val >= threshold:
                    self._sheet.highlight_cells(row=i, column=2 + j, fg="blue")

            # Highlight total column if deviating from 100%
            total = sum(group_data[lbl][i] for lbl in labels)
            if abs(total - 100.0) > 0.5:
                self._sheet.highlight_cells(row=i, column=total_col, fg="red")

    # ---------------------------------------------------------- manage groups
    def _manage_groups(self):
        """Open the Manage Groups dialog."""
        if self._ese_by_eid is None:
            messagebox.showinfo("No Data", "Load an OP2 file first.")
            return

        group_by = self._group_by_var.get()
        if group_by == "Property ID":
            mapping = self._eid_to_pid
        elif group_by == "Material ID":
            mapping = self._eid_to_mid
        else:
            mapping = self._eid_to_file

        # Collect all unique group IDs that appear in the energy data
        available = set()
        for eid in self._ese_by_eid:
            gid = mapping.get(eid)
            if gid is not None:
                available.add(gid)

        ManageGroupsDialog(
            self.frame.winfo_toplevel(),
            available_ids=available,
            existing_groups=self._custom_groups,
            show_ungrouped=self._show_ungrouped,
            on_apply=self._on_groups_applied,
        )

    def _on_groups_applied(self, groups, show_ungrouped):
        """Callback from ManageGroupsDialog."""
        self._custom_groups = {k: set(v) for k, v in groups.items()}
        self._show_ungrouped = show_ungrouped
        self._refresh_table()

    # ------------------------------------------------------------ export
    def _export_excel(self):
        if self._ese_by_eid is None:
            messagebox.showinfo("Nothing to export",
                                "Open an OP2 file first.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return

        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for Excel export.\n\n"
                "pip install openpyxl")
            return

        labels, group_data = self._aggregate_by_group()
        nmodes = len(self._modes)
        threshold = self._get_threshold()
        title = self._title_var.get().strip() or None
        op2_name = os.path.basename(self._op2_path) if self._op2_path else None

        # Build export data
        headers = ['Mode', 'Freq (Hz)'] + list(labels) + ['Total']
        table = []
        for i in range(nmodes):
            row = [int(self._modes[i]), float(self._freqs[i])]
            total = 0.0
            for lbl in labels:
                val = float(group_data[lbl][i])
                total += val
                row.append(val)
            row.append(total)
            table.append(row)

        export_data = {
            'headers': headers,
            'table': table,
        }

        wb = Workbook()
        styles = make_energy_styles()
        ws = wb.active
        ws.title = "ESE Breakdown"
        write_energy_sheet(ws, export_data, styles, op2_name=op2_name,
                           threshold=threshold, title=title)

        try:
            wb.save(path)
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))


def main():
    import logging
    logging.getLogger("customtkinter").setLevel(logging.ERROR)
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    root.title("ESE Breakdown")
    root.geometry("1400x600")
    mod = EnergyBreakdownModule(root)
    mod.frame.pack(fill=tk.BOTH, expand=True)
    root.mainloop()


if __name__ == '__main__':
    main()
