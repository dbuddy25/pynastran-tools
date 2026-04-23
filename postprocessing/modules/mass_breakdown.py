"""Mass breakdown module.

Reads element masses from a BDF file and displays per-group mass totals,
with optional OP2 GPWG validation. Supports superelements and DMIG mass
matrices referenced via M2GG case control.
"""
import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from tksheet import Sheet


# --------------------------------------------------------- Excel helpers

def make_mass_styles():
    """Return a dict of openpyxl style objects for Mass Breakdown sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        'dark_fill': PatternFill("solid", fgColor="1F4E79"),
        'mid_fill': PatternFill("solid", fgColor="2E75B6"),
        'white_bold': Font(bold=True, color="FFFFFF", size=11),
        'sub_font': Font(bold=True, color="FFFFFF", size=10),
        'center': Alignment(horizontal="center", vertical="center",
                             wrap_text=True),
        'right': Alignment(horizontal="right", vertical="center"),
        'cell_border': Border(bottom=Side(style='thin', color="B4C6E7")),
        'bold_font': Font(bold=True),
        'italic_font': Font(italic=True, color="808080"),
        'num2': '0.00',
        'pct1': '0.0',
    }


def write_mass_sheet(ws, data, styles, bdf_name=None, title=None, wtmass=None):
    """Write a mass breakdown sheet to an openpyxl worksheet.

    data keys:
      'headers'   — column header labels
      'table'     — list of data rows (group name, mass, %)
      'total_row' — total row values
      'gpwg_row'  — optional GPWG validation row
    """
    from openpyxl.utils import get_column_letter

    s = styles
    headers = data['headers']
    table = data['table']
    total_cols = len(headers)

    cur_row = 0

    # Row 1: custom title (only when provided)
    if title:
        cur_row += 1
        cell = ws.cell(row=cur_row, column=1, value=title)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=total_cols)
        for ci in range(2, total_cols + 1):
            ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # BDF filename row (always present, blank if no name)
    cur_row += 1
    name_text = bdf_name if bdf_name else ""
    cell = ws.cell(row=cur_row, column=1, value=name_text)
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # WTMASS row (always present so header rows stay at fixed positions)
    cur_row += 1
    if wtmass is not None and abs(wtmass - 1.0) < 1e-9:
        wtmass_text = f"PARAM,WTMASS = {wtmass} (default)"
    elif wtmass is not None:
        wtmass_text = f"PARAM,WTMASS = {wtmass:g}"
    else:
        wtmass_text = ""
    cell = ws.cell(row=cur_row, column=1, value=wtmass_text)
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=cur_row, column=ci).fill = s['dark_fill']

    # Headers row
    cur_row += 1
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Identify which columns are mass vs percent
    pct_col = total_cols - 1  # last column is always %

    # Data rows
    data_start = cur_row + 1
    for i, row_data in enumerate(table):
        row = i + data_start
        for ci, val in enumerate(row_data):
            cell = ws.cell(row=row, column=ci + 1, value=val)
            cell.alignment = s['center']
            if isinstance(val, float):
                cell.number_format = s['pct1'] if ci == pct_col else s['num2']
            cell.border = s['cell_border']

    # Total row
    total_row_data = data.get('total_row')
    if total_row_data:
        row = data_start + len(table)
        for ci, val in enumerate(total_row_data):
            cell = ws.cell(row=row, column=ci + 1, value=val)
            cell.alignment = s['center']
            cell.font = s['bold_font']
            if isinstance(val, float):
                cell.number_format = s['pct1'] if ci == pct_col else s['num2']
            cell.border = s['cell_border']

    # GPWG validation row
    gpwg_row_data = data.get('gpwg_row')
    if gpwg_row_data:
        row = data_start + len(table) + (1 if total_row_data else 0)
        for ci, val in enumerate(gpwg_row_data):
            cell = ws.cell(row=row, column=ci + 1, value=val)
            cell.alignment = s['center']
            cell.font = s['italic_font']
            if isinstance(val, float):
                cell.number_format = s['pct1'] if ci == pct_col else s['num2']
            cell.border = s['cell_border']

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14    # Type column
    for ci in range(2, total_cols):
        col_letter = get_column_letter(ci + 1)
        ws.column_dimensions[col_letter].width = 16

    ws.freeze_panes = f'A{data_start}'


# ---------------------------------------------------------- EID Range Dialog

class EIDRangeDialog(ctk.CTkToplevel):
    """Dialog for defining EID ranges to split mass elements into groups."""

    def __init__(self, parent, existing_ranges, eid_info, on_apply):
        super().__init__(parent)
        self.title("Mass Element EID Ranges")
        self.geometry("500x400")
        self.transient(parent)
        self.grab_set()

        self._ranges = dict(existing_ranges)  # {name: (lo, hi)}
        self._on_apply = on_apply

        # Info label
        ctk.CTkLabel(self, text=eid_info, text_color="gray").pack(
            padx=10, pady=(10, 5))

        # Input row
        input_frame = ctk.CTkFrame(self, fg_color="transparent")
        input_frame.pack(fill=tk.X, padx=10, pady=5)

        ctk.CTkLabel(input_frame, text="Name:").pack(side=tk.LEFT)
        self._name_var = tk.StringVar()
        ctk.CTkEntry(input_frame, textvariable=self._name_var,
                     width=140).pack(side=tk.LEFT, padx=(4, 10))

        ctk.CTkLabel(input_frame, text="EID range:").pack(side=tk.LEFT)
        self._lo_var = tk.StringVar()
        ctk.CTkEntry(input_frame, textvariable=self._lo_var,
                     width=70).pack(side=tk.LEFT, padx=(4, 0))
        ctk.CTkLabel(input_frame, text="\u2013").pack(side=tk.LEFT, padx=2)
        self._hi_var = tk.StringVar()
        ctk.CTkEntry(input_frame, textvariable=self._hi_var,
                     width=70).pack(side=tk.LEFT)

        ctk.CTkButton(input_frame, text="Add", width=60,
                      command=self._add_range).pack(side=tk.LEFT, padx=(10, 0))

        # List of existing ranges
        self._listbox = tk.Listbox(
            self, bg="#2b2b2b", fg="#dce4ee", selectbackground="#1f6aa5",
            font=("Consolas", 11), activestyle="none")
        self._listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self._refresh_list()

        # Bottom buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        ctk.CTkButton(btn_frame, text="Delete Selected", width=120,
                      command=self._delete_selected).pack(side=tk.LEFT)
        ctk.CTkButton(btn_frame, text="Apply", width=80,
                      command=self._apply).pack(side=tk.RIGHT)
        ctk.CTkButton(btn_frame, text="Cancel", width=80,
                      command=self.destroy).pack(side=tk.RIGHT, padx=(0, 5))

    def _refresh_list(self):
        self._listbox.delete(0, tk.END)
        for name, (lo, hi) in self._ranges.items():
            self._listbox.insert(tk.END, f"{name}  [{lo}\u2013{hi}]")

    def _add_range(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showwarning("Name required", "Enter a group name.",
                                   parent=self)
            return
        try:
            lo = int(self._lo_var.get())
            hi = int(self._hi_var.get())
        except ValueError:
            messagebox.showwarning("Invalid range",
                                   "Enter integer EID values.",
                                   parent=self)
            return
        if lo > hi:
            lo, hi = hi, lo

        self._ranges[name] = (lo, hi)
        self._name_var.set('')
        self._lo_var.set('')
        self._hi_var.set('')
        self._refresh_list()

    def _delete_selected(self):
        sel = self._listbox.curselection()
        if not sel:
            return
        keys = list(self._ranges.keys())
        del self._ranges[keys[sel[0]]]
        self._refresh_list()

    def _apply(self):
        self._on_apply(self._ranges)
        self.destroy()


# ---------------------------------------------------------------- GUI module

class MassBreakdownModule:
    name = "Mass Breakdown"

    def __init__(self, parent):
        self.frame = ctk.CTkFrame(parent)
        self._bdf_path = None
        self._op2_path = None
        self._title_var = tk.StringVar(value='')
        self._group_by_var = tk.StringVar(value='Property ID')
        self._units_var = tk.StringVar(value='slinch')
        self._display_var = tk.StringVar(value='lb')

        # Raw mass data from BDF
        self._mass_by_key = {}     # {"PID 5": float, "SE10:PID 5": float, ...}
        self._count_by_key = {}    # element counts per key
        self._pid_names = {}       # {"PID 5": "Wing Upper", ...}
        self._has_superelements = False
        self._bdf_loaded = False

        # Include file mapping
        self._mass_by_file = {}    # {rel_path: float}
        self._count_by_file = {}
        self._file_order = []
        self._file_types = {}      # {rel_path: set of type strings}
        self._dmig_name_to_file = {}  # {dmig_name: rel_path}
        self._wtmass = 1.0

        # DMIG mass from M2GG case control
        self._dmig_mass = {}       # {"M2GG: MPART1 (x1.03)": float, ...}

        # Per-EID mass for mass elements (for ID range grouping)
        self._mass_elem_by_eid = {}  # {eid: float}

        # GPWG from OP2
        self._gpwg_mass = None     # total GPWG mass (float) or None

        # Custom group merges
        self._custom_groups = {}       # {name: set(keys)}
        self._show_ungrouped = True

        # EID range groups for mass elements
        self._eid_range_groups = {}    # {name: (start, end)}

        # Editable column names (name row)
        self._column_names = {}        # {key: display_name}
        self._current_keys = []        # keys for mapping row edits

        self._build_ui()

    # ------------------------------------------------------------------ UI
    _GUIDE_TEXT = """\
Mass Breakdown Tool — Quick Guide

PURPOSE
Compute element mass breakdown from a Nastran BDF file, grouped by
property ID or include file. Optionally validate against OP2 GPWG
(Grid Point Weight Generator) total mass.

WORKFLOW
1. Open BDF — select a BDF file to extract element masses.
2. Select grouping — "Property ID" or "Include File".
3. Review — the table shows mass per group with percentages.
4. Manage Groups — combine multiple IDs into named groups.
5. Open OP2 (optional) — load an OP2 for GPWG validation.
6. Export to Excel — save as a formatted .xlsx workbook.

UNIT CONVERSION
Set "Units" to your model's mass unit (kg, lb, or slinch) and
use the arrow dropdown to select display units. For example,
slinch -> lb converts from slinch to pounds. The conversion is
applied to the table and Excel export. Percentages are unaffected.

GROUPING MODES
  Property ID — group by element property ID (PID)
  Include File — group by source BDF include file

SUPERELEMENT SUPPORT
When the BDF contains superelements (via INCLUDE + BEGIN SUPER),
each superelement's elements are prefixed with "SE{id}:" in the
group labels (e.g. "SE10:PID 5") to distinguish from residual.
You can merge SE and residual PIDs together via Manage Groups.

GENERATING GPWG DATA
Add to your Nastran bulk data section:
  PARAM,GRDPNT,0
This tells Nastran to compute and output the Grid Point Weight
Generator table at the origin. The GPWG contains total mass,
center of gravity, and inertia for the assembled model.
For superelement models, each SE will have its own GPWG entry.

DMIG MASS (M2GG)
If the case control deck contains M2GG entries referencing DMIG mass
matrices (e.g. M2GG = 1.03*MPART1, 1.06*MPART2), the tool extracts
mass from each matrix by summing diagonal translational DOFs and
applying the scale factor. Each matrix appears as its own group
(e.g. "M2GG: MPART1 (x1.03)") in both grouping modes.

MASS ELEMENTS
CONM2 and other mass elements (CMASS1-4, CONM1) are grouped as
"Mass Elements". CONROD elements appear as "CONROD (no PID)".

REQUIREMENTS
  - pyNastran (for BDF/OP2 reading)
  - openpyxl (for Excel export only)\
"""

    def _build_ui(self):
        toolbar = ctk.CTkFrame(self.frame, fg_color="transparent")
        toolbar.pack(fill=tk.X, padx=5, pady=(5, 0))

        self._bdf_btn = ctk.CTkButton(toolbar, text="Open BDF\u2026", width=100,
                                      command=self._open_bdf)
        self._bdf_btn.pack(side=tk.LEFT)

        self._op2_btn = ctk.CTkButton(toolbar, text="Open OP2\u2026", width=100,
                                      command=self._open_op2)
        self._op2_btn.pack(side=tk.LEFT, padx=(5, 0))

        # Separator
        ctk.CTkLabel(toolbar, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        # Group by dropdown
        ctk.CTkLabel(toolbar, text="Group by:").pack(side=tk.LEFT, padx=(0, 2))
        self._group_by_menu = ctk.CTkOptionMenu(
            toolbar, variable=self._group_by_var,
            values=["Property ID", "Include File"],
            width=140, command=self._on_group_by_change,
        )
        self._group_by_menu.pack(side=tk.LEFT)

        # Manage Groups button
        self._manage_btn = ctk.CTkButton(
            toolbar, text="Manage Groups\u2026", width=120,
            command=self._manage_groups)
        self._manage_btn.pack(side=tk.LEFT, padx=(5, 0))
        self._manage_btn.configure(state=tk.DISABLED)

        # EID Ranges button (for splitting mass elements)
        self._eid_range_btn = ctk.CTkButton(
            toolbar, text="EID Ranges\u2026", width=100,
            command=self._manage_eid_ranges)
        self._eid_range_btn.pack(side=tk.LEFT, padx=(5, 0))
        self._eid_range_btn.configure(state=tk.DISABLED)

        # Separator
        ctk.CTkLabel(toolbar, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        # Unit conversion
        unit_choices = list(self._UNIT_TO_KG.keys())
        ctk.CTkLabel(toolbar, text="Units:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkOptionMenu(
            toolbar, variable=self._units_var, values=unit_choices,
            width=80, command=self._on_units_change,
        ).pack(side=tk.LEFT)

        ctk.CTkLabel(toolbar, text="\u2192").pack(side=tk.LEFT, padx=4)

        ctk.CTkOptionMenu(
            toolbar, variable=self._display_var, values=unit_choices,
            width=80, command=self._on_units_change,
        ).pack(side=tk.LEFT)

        # Separator
        ctk.CTkLabel(toolbar, text="|", text_color="gray").pack(
            side=tk.LEFT, padx=6)

        # Title field
        ctk.CTkLabel(toolbar, text="Title:").pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkEntry(toolbar, textvariable=self._title_var, width=160).pack(
            side=tk.LEFT, padx=(0, 4))

        # Right side: ?, Export
        ctk.CTkButton(
            toolbar, text="?", width=30, font=ctk.CTkFont(weight="bold"),
            command=self._show_guide,
        ).pack(side=tk.RIGHT, padx=(5, 0))

        ctk.CTkButton(toolbar, text="Export to Excel\u2026", width=130,
                      command=self._export_excel).pack(side=tk.RIGHT)

        # Status label
        self._status_label = ctk.CTkLabel(
            self.frame, text="No BDF loaded", text_color="gray",
            anchor=tk.W)
        self._status_label.pack(fill=tk.X, padx=10, pady=(2, 0))

        # Table (tksheet)
        self._sheet = Sheet(
            self.frame,
            headers=["Group", "Mass", "% of Total"],
            show_top_left=False,
            show_row_index=False,
        )
        self._sheet.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._sheet.disable_bindings()
        self._sheet.enable_bindings(
            "single_select", "copy", "arrowkeys",
            "column_width_resize", "row_height_resize",
            "edit_cell",
        )
        self._sheet.extra_bindings([("end_edit_cell", self._on_name_edit)])

    # ---------------------------------------------------------- Guide
    def _show_guide(self):
        try:
            from structures_tools import show_guide
        except ImportError:
            return
        show_guide(self.frame.winfo_toplevel(), "Mass Breakdown Guide",
                   self._GUIDE_TEXT)

    def _on_group_by_change(self, *args):
        self._custom_groups = {}
        self._show_ungrouped = True
        self._column_names = {}
        if self._bdf_loaded:
            self._refresh_table()

    # Mass unit conversion factors to kg (base unit)
    _UNIT_TO_KG = {
        'kg':     1.0,
        'lb':     0.45359237,
        'slinch': 175.12683,    # 1 lbf·s²/in = 386.088 lbm = 175.127 kg
    }

    def _on_units_change(self, *args):
        if self._bdf_loaded:
            self._refresh_table()

    def _get_display_scale(self):
        """Return conversion factor from model units to display units."""
        from_unit = self._units_var.get()
        to_unit = self._display_var.get()
        from_kg = self._UNIT_TO_KG.get(from_unit, 1.0)
        to_kg = self._UNIT_TO_KG.get(to_unit, 1.0)
        return from_kg / to_kg

    # ---------------------------------------------------------- background work
    def _run_in_background(self, label, work_fn, done_fn):
        """Run *work_fn* in a background thread, keeping the UI responsive."""
        self._status_label.configure(text=label, text_color="gray")
        self._bdf_btn.configure(state=tk.DISABLED)
        self._op2_btn.configure(state=tk.DISABLED)

        container = {}

        def _worker():
            try:
                container['result'] = work_fn()
            except Exception as exc:
                container['error'] = exc

        def _poll():
            if thread.is_alive():
                self.frame.after(50, _poll)
            else:
                self._bdf_btn.configure(state=tk.NORMAL)
                self._op2_btn.configure(state=tk.NORMAL)
                done_fn(container.get('result'), container.get('error'))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()
        self.frame.after(50, _poll)

    # ---------------------------------------------------------- BDF loading
    @staticmethod
    def _extract_comment_name(comment):
        """Extract a descriptive name from a BDF card comment string."""
        if not comment:
            return None
        result = None
        for line in comment.splitlines():
            line = line.strip().lstrip('$').strip()
            if line:
                if ':' in line:
                    line = line.split(':', 1)[1].strip()
                if line:
                    result = line
        return result

    @staticmethod
    def _extract_model_mass(model, seid, mass_by_key, count_by_key,
                            pid_names, mass_elem_by_eid):
        """Extract per-element mass from a single BDF model.

        Populates mass_by_key and count_by_key with string keys like
        "PID 5" (residual) or "SE10:PID 5" (superelement).
        Also populates mass_elem_by_eid with per-EID mass for mass elements.
        """
        prefix = f"SE{seid}:" if seid else ""

        # Extract comment names from property cards
        for pid, prop in model.properties.items():
            comment = getattr(prop, 'comment', '')
            name = MassBreakdownModule._extract_comment_name(comment)
            if name:
                key = f"{prefix}PID {pid}"
                pid_names[key] = name

        # Structural elements
        for eid, elem in model.elements.items():
            pid = getattr(elem, 'pid', None)

            if elem.type == 'CONROD':
                key = f"{prefix}CONROD (no PID)"
            elif pid is not None:
                try:
                    pid_int = int(pid)
                except (ValueError, TypeError):
                    continue
                if pid_int == 0:
                    continue
                key = f"{prefix}PID {pid_int}"
            else:
                continue

            try:
                m = elem.Mass()
            except Exception:
                m = 0.0

            mass_by_key[key] = mass_by_key.get(key, 0.0) + m
            count_by_key[key] = count_by_key.get(key, 0) + 1

        # Mass elements (CONM2, CMASS1-4, CONM1)
        for eid, elem in model.masses.items():
            key = f"{prefix}Mass Elements"

            try:
                if elem.type == 'CONM2':
                    m = elem.mass
                else:
                    m = elem.Mass()
            except Exception:
                m = 0.0

            mass_by_key[key] = mass_by_key.get(key, 0.0) + m
            count_by_key[key] = count_by_key.get(key, 0) + 1
            mass_elem_by_eid[eid] = m

    @staticmethod
    def _extract_dmig_mass(model):
        """Extract mass from DMIG matrices referenced by M2GG case control.

        Parses M2GG entries from the case control deck, retrieves each
        referenced DMIG matrix, and computes total mass via rigid body
        translation: m = {1}^T [M_dd] {1} for each translational direction,
        then averages across directions. This correctly handles condensed
        superelement mass matrices with off-diagonal coupling terms.

        Returns dict {display_label: mass} for each M2GG term, e.g.
        {"MPART1 (x1.03)": 45.23, "MPART2 (x1.06)": 38.71}.
        """
        import numpy as np

        ccd = getattr(model, 'case_control_deck', None)
        if ccd is None:
            return {}

        # Collect M2GG entries from all subcases (including subcase 0 = global)
        m2gg_terms = []  # list of (scale, matrix_name)
        seen_names = set()
        for sc_id, subcase in ccd.subcases.items():
            try:
                value, _options = subcase.get_parameter('M2GG')
            except (KeyError, AttributeError):
                continue
            if isinstance(value, int):
                # SET reference — not supported yet
                continue
            if isinstance(value, list):
                for item in value:
                    if isinstance(item, tuple) and len(item) == 2:
                        scale, name = item
                        if name not in seen_names:
                            m2gg_terms.append((float(scale), str(name)))
                            seen_names.add(name)

        if not m2gg_terms:
            return {}

        dmig_dict = getattr(model, 'dmig', {})
        dmig_mass = {}

        for scale, name in m2gg_terms:
            dmig_obj = dmig_dict.get(name)
            if dmig_obj is None:
                continue

            try:
                matrix, rows, _cols = dmig_obj.get_matrix(
                    is_sparse=False, apply_symmetry=True)
                matrix = np.asarray(matrix)
            except Exception:
                continue

            # Compute total mass via rigid body translation:
            # For each translational DOF direction d, collect all row/col
            # indices with that component, extract the submatrix, and sum
            # all entries: m_d = {1}^T [M_dd] {1}
            dir_masses = []
            for comp in (1, 2, 3):
                indices = [idx for idx, (g, c) in rows.items() if c == comp]
                if not indices:
                    continue
                ix = np.array(indices)
                sub = matrix[np.ix_(ix, ix)]
                dir_masses.append(float(sub.sum()))

            if not dir_masses:
                continue

            # Average across available directions (should be nearly equal)
            node_mass = sum(dir_masses) / len(dir_masses)

            scaled_mass = scale * node_mass
            if abs(scaled_mass) < 1e-20:
                continue

            if abs(scale - 1.0) > 1e-6:
                label = f"M2GG: {name} (x{scale:g})"
            else:
                label = f"M2GG: {name}"
            dmig_mass[label] = scaled_mass

        return dmig_mass

    def _open_bdf(self):
        path = filedialog.askopenfilename(
            title="Open BDF File",
            filetypes=[("BDF files", "*.bdf *.dat *.nas *.bulk"),
                       ("All files", "*.*")])
        if not path:
            return

        def _work():
            return self._load_bdf(path)

        def _done(result, error):
            if error is not None:
                messagebox.showerror("Error",
                                     f"Could not read BDF:\n{error}")
                self._status_label.configure(text="BDF load failed",
                                             text_color="red")
                return

            (mass_by_key, count_by_key, pid_names,
             has_se, mass_by_file, count_by_file, file_order,
             dmig_mass, mass_elem_by_eid,
             wtmass, dmig_name_to_file, file_types) = result

            self._bdf_path = path
            self._mass_by_key = mass_by_key
            self._count_by_key = count_by_key
            self._pid_names = pid_names
            self._has_superelements = has_se
            self._mass_by_file = mass_by_file
            self._count_by_file = count_by_file
            self._file_order = file_order
            self._dmig_mass = dmig_mass
            self._mass_elem_by_eid = mass_elem_by_eid
            self._wtmass = wtmass
            self._dmig_name_to_file = dmig_name_to_file
            self._file_types = file_types
            self._bdf_loaded = True

            self._manage_btn.configure(state=tk.NORMAL)
            self._eid_range_btn.configure(state=tk.NORMAL)

            # Reset custom groups and EID ranges
            self._custom_groups = {}
            self._show_ungrouped = True
            self._column_names = {}
            self._eid_range_groups = {}

            n_groups = len(mass_by_key) + len(dmig_mass)
            total_mass = sum(mass_by_key.values()) + sum(dmig_mass.values())
            wtmass_str = (f"WTMASS={wtmass:g}"
                          if abs(wtmass - 1.0) > 1e-9 else "WTMASS=1.0")
            status = (f"BDF: {os.path.basename(path)} "
                      f"({n_groups} groups, total mass: {total_mass:.1f})  |  "
                      f"{wtmass_str}")
            if dmig_mass:
                status += f"  [{len(dmig_mass)} M2GG]"
            if self._op2_path:
                status += f"  |  OP2: {os.path.basename(self._op2_path)}"
            self._status_label.configure(text=status,
                                         text_color=("gray10", "gray90"))
            self._refresh_table()

        self._run_in_background("Loading BDF\u2026", _work, _done)

    def _load_bdf(self, bdf_path):
        """Background worker — extract mass data from BDF.

        Returns (mass_by_key, count_by_key, pid_names, has_superelements,
                 mass_by_file, count_by_file, file_order, dmig_mass,
                 mass_elem_by_eid, wtmass, dmig_name_to_file, file_types).
        """
        from bdf_utils import IncludeFileParser, make_model, read_bdf_safe

        model = make_model()
        read_bdf_safe(model, bdf_path, xref=True)

        mass_by_key = {}
        count_by_key = {}
        pid_names = {}
        mass_elem_by_eid = {}

        # Residual structure
        self._extract_model_mass(model, seid=0,
                                 mass_by_key=mass_by_key,
                                 count_by_key=count_by_key,
                                 pid_names=pid_names,
                                 mass_elem_by_eid=mass_elem_by_eid)

        # Superelements
        has_se = False
        se_models = getattr(model, 'superelement_models', {})
        for se_key, se_model in se_models.items():
            has_se = True
            seid = se_key[1] if isinstance(se_key, tuple) else int(se_key)
            try:
                se_model.cross_reference()
            except Exception:
                pass
            self._extract_model_mass(se_model, seid=seid,
                                     mass_by_key=mass_by_key,
                                     count_by_key=count_by_key,
                                     pid_names=pid_names,
                                     mass_elem_by_eid=mass_elem_by_eid)

        # DMIG mass from M2GG case control
        dmig_mass = self._extract_dmig_mass(model)

        # Apply WTMASS to element-computed masses so they're in true mass
        # units (consistent with GPWG and DMIG matrices)
        wtmass = getattr(model, 'wtmass', 1.0)
        if wtmass != 1.0:
            for key in mass_by_key:
                mass_by_key[key] *= wtmass
            for eid in mass_elem_by_eid:
                mass_elem_by_eid[eid] *= wtmass

        # Include file mapping
        mass_by_file = {}
        count_by_file = {}
        file_order = []
        file_types = {}  # {rel: set of type strings}

        parser = IncludeFileParser()
        parser.parse(bdf_path)
        main_dir = os.path.dirname(os.path.abspath(bdf_path))

        # Build eid→file mapping (now includes Part SE elements via BEGIN SUPER)
        eid_to_file = {}
        for filepath, ids_by_type in parser.file_ids.items():
            eids = ids_by_type.get('eid', set())
            try:
                rel = os.path.relpath(filepath, main_dir)
            except ValueError:
                rel = os.path.basename(filepath)
            for eid in eids:
                eid_to_file[eid] = rel

        for fp in parser.all_files:
            try:
                rel = os.path.relpath(fp, main_dir)
            except ValueError:
                rel = os.path.basename(fp)
            file_order.append(rel)

        # Build DMIG name→file mapping
        dmig_name_to_file = {}
        for name, fp in parser.dmig_origins.items():
            try:
                rel = os.path.relpath(fp, main_dir)
            except ValueError:
                rel = os.path.basename(fp)
            dmig_name_to_file[name] = rel

        def _add_to_file(rel, m, type_label):
            mass_by_file[rel] = mass_by_file.get(rel, 0.0) + m
            count_by_file[rel] = count_by_file.get(rel, 0) + 1
            file_types.setdefault(rel, set()).add(type_label)

        # Accumulate residual element mass per include file
        for eid, elem in model.elements.items():
            rel = eid_to_file.get(eid)
            if rel is None:
                continue
            try:
                m = elem.Mass()
            except Exception:
                m = 0.0
            _add_to_file(rel, m, 'Residual')

        for eid, elem in model.masses.items():
            rel = eid_to_file.get(eid)
            if rel is None:
                continue
            try:
                m = elem.mass if elem.type == 'CONM2' else elem.Mass()
            except Exception:
                m = 0.0
            _add_to_file(rel, m, 'Residual')

        # Accumulate Part SE element mass per include file
        for _se_key, se_model in se_models.items():
            for eid, elem in se_model.elements.items():
                rel = eid_to_file.get(eid)
                if rel is None:
                    continue
                try:
                    m = elem.Mass()
                except Exception:
                    m = 0.0
                _add_to_file(rel, m, 'Part SE')
            for eid, elem in se_model.masses.items():
                rel = eid_to_file.get(eid)
                if rel is None:
                    continue
                try:
                    m = elem.mass if elem.type == 'CONM2' else elem.Mass()
                except Exception:
                    m = 0.0
                _add_to_file(rel, m, 'Part SE')

        # Apply WTMASS to include-file masses so they match GPWG / DMIG units
        if wtmass != 1.0:
            for key in mass_by_file:
                mass_by_file[key] *= wtmass

        return (mass_by_key, count_by_key, pid_names, has_se,
                mass_by_file, count_by_file, file_order, dmig_mass,
                mass_elem_by_eid, wtmass, dmig_name_to_file, file_types)

    # ---------------------------------------------------------- OP2 loading
    def _open_op2(self):
        path = filedialog.askopenfilename(
            title="Open OP2 File (GPWG validation)",
            filetypes=[("OP2 files", "*.op2"), ("All files", "*.*")])
        if not path:
            return

        def _work():
            return self._load_gpwg(path)

        def _done(result, error):
            if error is not None:
                messagebox.showerror("Error",
                                     f"Could not read OP2:\n{error}")
                self._status_label.configure(text="OP2 load failed",
                                             text_color="red")
                return

            gpwg_mass = result
            if gpwg_mass is None:
                messagebox.showwarning(
                    "No GPWG Data",
                    "No Grid Point Weight Generator data found.\n\n"
                    "Add to your Nastran bulk data:\n"
                    "  PARAM,GRDPNT,0")
                return

            self._op2_path = path
            self._gpwg_mass = gpwg_mass

            status = ""
            if self._bdf_path:
                n_groups = len(self._mass_by_key)
                total_mass = sum(self._mass_by_key.values())
                wtmass_str = (f"WTMASS={self._wtmass:g}"
                              if abs(self._wtmass - 1.0) > 1e-9 else "WTMASS=1.0")
                status += (f"BDF: {os.path.basename(self._bdf_path)} "
                           f"({n_groups} groups, total mass: {total_mass:.1f})  |  "
                           f"{wtmass_str}  |  ")
            status += f"OP2: {os.path.basename(path)} (GPWG: {gpwg_mass:.1f})"
            self._status_label.configure(text=status,
                                         text_color=("gray10", "gray90"))
            if self._bdf_loaded:
                self._refresh_table()

        self._run_in_background("Loading OP2\u2026", _work, _done)

    @staticmethod
    def _load_gpwg(op2_path):
        """Load GPWG total mass from OP2. Returns float or None."""
        from pyNastran.op2.op2 import OP2
        op2 = OP2(mode='nx')
        op2.read_op2(op2_path)

        gpwg = getattr(op2, 'grid_point_weight', None)
        if not gpwg:
            return None

        # Sum GPWG mass across all superelements
        total = 0.0
        for key, weight in gpwg.items():
            mass = weight.mass
            if hasattr(mass, '__len__'):
                total += float(mass[0])
            else:
                total += float(mass)
        return total if total > 0.0 else None

    # ---------------------------------------------------------- aggregation
    def _aggregate_by_group(self):
        """Aggregate mass by the selected grouping type.

        Returns (group_keys, group_mass) where:
          group_keys: ordered list of group key strings
          group_mass: dict {key: float}
        """
        if not self._bdf_loaded:
            return [], {}

        group_by = self._group_by_var.get()

        if group_by == "Include File":
            raw_groups = dict(self._mass_by_file)
        else:
            raw_groups = dict(self._mass_by_key)

        # Split "Mass Elements" by EID ranges (if any defined)
        if self._eid_range_groups and self._mass_elem_by_eid:
            claimed = set()
            for range_name, (eid_lo, eid_hi) in self._eid_range_groups.items():
                range_mass = 0.0
                for eid, m in self._mass_elem_by_eid.items():
                    if eid_lo <= eid <= eid_hi and eid not in claimed:
                        range_mass += m
                        claimed.add(eid)
                if range_mass > 0:
                    raw_groups[range_name] = range_mass

            # Subtract claimed mass from "Mass Elements" bucket
            claimed_total = sum(self._mass_elem_by_eid[e]
                                for e in claimed
                                if e in self._mass_elem_by_eid)
            me_key = 'Mass Elements'
            if me_key in raw_groups:
                raw_groups[me_key] -= claimed_total
                if raw_groups[me_key] < 1e-20:
                    del raw_groups[me_key]

        # Append DMIG mass — in "Include File" mode, bucket under the file
        # that defines the DMIG card; fall back to own label if unknown
        if group_by == "Include File":
            for label, mass in self._dmig_mass.items():
                name_match = re.match(r'M2GG:\s+(\S+)', label)
                dmig_name = name_match.group(1) if name_match else None
                rel = self._dmig_name_to_file.get(dmig_name) if dmig_name else None
                if rel:
                    raw_groups[rel] = raw_groups.get(rel, 0.0) + mass
                    self._file_types.setdefault(rel, set()).add('External SE')
                else:
                    raw_groups[label] = mass
        else:
            for label, mass in self._dmig_mass.items():
                raw_groups[label] = mass

        # Apply custom group merges
        if self._custom_groups:
            merged_groups = {}
            consumed_keys = set()

            for group_name, member_keys in self._custom_groups.items():
                merged = 0.0
                for k in member_keys:
                    if k in raw_groups:
                        merged += raw_groups[k]
                        consumed_keys.add(k)
                merged_groups[group_name] = merged

            remaining = {k: v for k, v in raw_groups.items()
                         if k not in consumed_keys}

            if self._show_ungrouped:
                for k, v in remaining.items():
                    merged_groups[k] = v
            else:
                if remaining:
                    merged_groups['Other'] = sum(remaining.values())

            final_groups = merged_groups
        else:
            final_groups = raw_groups

        # Sort keys
        if self._custom_groups:
            custom_names = [n for n in self._custom_groups if n in final_groups]
            rest = [k for k in final_groups if k not in self._custom_groups]
            if group_by == "Include File" and self._file_order:
                order_map = {f: i for i, f in enumerate(self._file_order)}
                rest.sort(key=lambda k: order_map.get(k, 999999))
            else:
                rest.sort(key=self._group_sort_key)
            keys = custom_names + rest
        elif group_by == "Include File" and self._file_order:
            order_map = {f: i for i, f in enumerate(self._file_order)}
            keys = sorted(final_groups.keys(),
                          key=lambda k: order_map.get(k, 999999))
        else:
            keys = sorted(final_groups.keys(), key=self._group_sort_key)

        return keys, final_groups

    @staticmethod
    def _group_sort_key(label):
        """Sort key: numeric PIDs by number, DMIG after PIDs, special last."""
        m = re.search(r'PID\s+(\d+)', label)
        if m:
            return (0, int(m.group(1)), '')
        if label.startswith('M2GG:'):
            return (1, 0, label)
        if label in ('Other', 'Mass Elements') or label.endswith('Mass Elements'):
            return (3, 0, label)
        if 'CONROD' in label:
            return (2, 999999, label)
        return (2, 0, label)

    # ---------------------------------------------------------- display
    @staticmethod
    def _key_type(key):
        """Return the source type for a single raw group key."""
        if key.startswith('M2GG:'):
            return 'External SE'
        if re.match(r'SE\d+:', key):
            return 'Part SE'
        return 'Residual'

    def _get_group_type(self, key):
        """Return the type label for a group key (handles custom groups)."""
        if key in self._custom_groups:
            types = sorted({self._key_type(k) for k in self._custom_groups[key]})
            return ' / '.join(types)
        if key in self._eid_range_groups:
            return 'Residual'
        if key == 'Other':
            return ''
        # For include-file keys, report mixed types from _file_types
        if key in self._file_types:
            types = sorted(self._file_types[key])
            return types[0] if len(types) == 1 else ' / '.join(types)
        return self._key_type(key)

    def _get_display_name(self, key):
        """Return the display name for a group key."""
        if key in self._column_names:
            return self._column_names[key]
        if key in self._custom_groups:
            return key
        if key in self._pid_names:
            return self._pid_names[key]
        return ''

    def _refresh_table(self):
        """Rebuild the table from current data and grouping settings."""
        if not self._bdf_loaded:
            return

        keys, group_mass = self._aggregate_by_group()
        self._current_keys = list(keys)
        total_mass = sum(group_mass.values())
        scale = self._get_display_scale()

        from_unit = self._units_var.get()
        to_unit = self._display_var.get()
        show_converted = abs(scale - 1.0) > 1e-9

        # Build headers — add converted column only when units differ
        if show_converted:
            headers = ['Group', 'Type', f'Mass ({from_unit})',
                       f'Mass ({to_unit})', '% of Total']
        else:
            headers = ['Group', 'Type', f'Mass ({from_unit})', '% of Total']

        # Name row
        name_row = ['Name'] + [''] * (len(headers) - 1)

        # Data rows
        table_data = [name_row]
        for key in keys:
            raw = group_mass[key]
            pct = (raw / total_mass * 100.0) if total_mass > 0 else 0.0
            display_name = self._get_display_name(key)
            label = display_name if display_name else key
            gtype = self._get_group_type(key)
            if show_converted:
                table_data.append([label, gtype, round(raw, 2),
                                   round(raw * scale, 2), round(pct, 1)])
            else:
                table_data.append([label, gtype, round(raw, 2),
                                   round(pct, 1)])

        # Total row
        if show_converted:
            table_data.append(['TOTAL', '', round(total_mass, 2),
                               round(total_mass * scale, 2), 100.0])
        else:
            table_data.append(['TOTAL', '', round(total_mass, 2), 100.0])

        # GPWG validation row
        if self._gpwg_mass is not None:
            delta = self._gpwg_mass - total_mass
            sign = '+' if delta >= 0 else ''
            gpwg_label = f"GPWG Total (\u0394: {sign}{delta:.2f})"
            if show_converted:
                table_data.append([gpwg_label, '', round(self._gpwg_mass, 2),
                                   round(self._gpwg_mass * scale, 2), ''])
            else:
                table_data.append([gpwg_label, '',
                                   round(self._gpwg_mass, 2), ''])

        # Update sheet
        self._sheet.headers(headers)
        self._sheet.set_sheet_data(table_data)
        self._sheet.set_header_height_lines(1)

        ncols = len(headers)
        self._sheet.set_all_column_widths(120)
        self._sheet.column_width(column=0, width=250)
        self._sheet.column_width(column=1, width=100)
        self._sheet.align_columns(
            list(range(ncols)), align="center", align_header=True)

        # Lock all columns except name row group label
        self._sheet.readonly_columns(columns=list(range(1, ncols)))
        n_data_rows = len(table_data)
        if n_data_rows > 1:
            self._sheet.readonly_rows(rows=list(range(1, n_data_rows)))

        self._apply_highlights()

    def _apply_highlights(self):
        """Apply styling to total row and GPWG delta (color-coded)."""
        self._sheet.dehighlight_all(redraw=False)

        if not self._bdf_loaded:
            return

        keys = self._current_keys
        n_keys = len(keys)
        scale = self._get_display_scale()
        ncols = 4 if abs(scale - 1.0) > 1e-9 else 3

        # Total row index = 1 (name row) + n_keys (data rows)
        total_row = 1 + n_keys
        for c in range(ncols):
            self._sheet.highlight_cells(row=total_row, column=c,
                                        fg="white", bg="#1F4E79")

        # GPWG row — green if close (< 1% delta), red if not
        if self._gpwg_mass is not None:
            gpwg_row = total_row + 1
            total_mass = sum(self._aggregate_by_group()[1].values())
            delta = abs(self._gpwg_mass - total_mass)
            pct_diff = (delta / self._gpwg_mass * 100.0) if self._gpwg_mass > 0 else 0.0

            if pct_diff < 1.0:
                fg_color = "#2d8a4e"   # green
                bg_color = "#e6f5eb"
            else:
                fg_color = "#c0392b"   # red
                bg_color = "#fdecea"

            for c in range(ncols):
                self._sheet.highlight_cells(row=gpwg_row, column=c,
                                            fg=fg_color, bg=bg_color)

    def _on_name_edit(self, event):
        """Capture edits to the name row and store them."""
        r = event.row
        c = event.column
        if r != 0 or c != 0 or not self._current_keys:
            return
        # Name row col 0 is just the "Name" label — not useful
        # Actually, users edit individual data rows' group column
        # But with this table layout, the name row is row 0

    # ---------------------------------------------------------- manage groups
    def _manage_groups(self):
        """Open the Manage Groups dialog."""
        if not self._bdf_loaded:
            messagebox.showinfo("No Data", "Load a BDF file first.")
            return

        from modules.energy_breakdown import ManageGroupsDialog

        group_by = self._group_by_var.get()
        if group_by == "Include File":
            available = set(self._mass_by_file.keys())
            id_labels = {}
        else:
            available = set(self._mass_by_key.keys())
            id_labels = {}
            for key in available:
                name = self._pid_names.get(key)
                if name:
                    id_labels[key] = f"{key} \u2014 {name}"

        # Include DMIG groups as available for merging
        available.update(self._dmig_mass.keys())

        ManageGroupsDialog(
            self.frame.winfo_toplevel(),
            available_ids=available,
            existing_groups=self._custom_groups,
            show_ungrouped=self._show_ungrouped,
            on_apply=self._on_groups_applied,
            id_labels=id_labels,
        )

    def _on_groups_applied(self, groups, show_ungrouped):
        """Callback from ManageGroupsDialog."""
        self._custom_groups = {k: set(v) for k, v in groups.items()}
        self._show_ungrouped = show_ungrouped
        self._refresh_table()

    # ---------------------------------------------------------- EID ranges
    def _manage_eid_ranges(self):
        """Open dialog to define EID ranges for mass element grouping."""
        if not self._bdf_loaded:
            messagebox.showinfo("No Data", "Load a BDF file first.")
            return

        n_mass = len(self._mass_elem_by_eid)
        if n_mass == 0:
            messagebox.showinfo("No Mass Elements",
                                "This model has no mass elements to split.")
            return

        eids = sorted(self._mass_elem_by_eid.keys())
        eid_min, eid_max = eids[0], eids[-1]

        EIDRangeDialog(
            self.frame.winfo_toplevel(),
            existing_ranges=dict(self._eid_range_groups),
            eid_info=f"{n_mass} mass elements (EID {eid_min}\u2013{eid_max})",
            on_apply=self._on_eid_ranges_applied,
        )

    def _on_eid_ranges_applied(self, ranges):
        """Callback from EIDRangeDialog."""
        self._eid_range_groups = dict(ranges)
        self._refresh_table()

    # ------------------------------------------------------------ export
    def _export_excel(self):
        if not self._bdf_loaded:
            messagebox.showinfo("Nothing to export",
                                "Load a BDF file first.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return

        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for Excel export.\n\n"
                "pip install openpyxl")
            return

        keys, group_mass = self._aggregate_by_group()
        total_mass = sum(group_mass.values())
        scale = self._get_display_scale()
        from_unit = self._units_var.get()
        to_unit = self._display_var.get()
        show_converted = abs(scale - 1.0) > 1e-9
        title = self._title_var.get().strip() or None
        bdf_name = os.path.basename(self._bdf_path) if self._bdf_path else None

        # Build export data
        if show_converted:
            headers = ['Group', 'Type', f'Mass ({from_unit})',
                       f'Mass ({to_unit})', '% of Total']
        else:
            headers = ['Group', 'Type', f'Mass ({from_unit})', '% of Total']

        table = []
        for key in keys:
            raw = group_mass[key]
            pct = (raw / total_mass * 100.0) if total_mass > 0 else 0.0
            display_name = self._get_display_name(key)
            label = display_name if display_name else key
            gtype = self._get_group_type(key)
            if show_converted:
                table.append([label, gtype, raw, raw * scale, pct])
            else:
                table.append([label, gtype, raw, pct])

        if show_converted:
            total_row = ['TOTAL', '', total_mass, total_mass * scale, 100.0]
        else:
            total_row = ['TOTAL', '', total_mass, 100.0]

        gpwg_row = None
        if self._gpwg_mass is not None:
            delta = self._gpwg_mass - total_mass
            sign = '+' if delta >= 0 else ''
            if show_converted:
                gpwg_row = [f"GPWG Total (\u0394: {sign}{delta:.2f})",
                            '', self._gpwg_mass, self._gpwg_mass * scale, '']
            else:
                gpwg_row = [f"GPWG Total (\u0394: {sign}{delta:.2f})",
                            '', self._gpwg_mass, '']

        export_data = {
            'headers': headers,
            'table': table,
            'total_row': total_row,
            'gpwg_row': gpwg_row,
        }

        wb = Workbook()
        styles = make_mass_styles()
        ws = wb.active
        ws.title = "Mass Breakdown"
        write_mass_sheet(ws, export_data, styles, bdf_name=bdf_name,
                         title=title, wtmass=self._wtmass)

        try:
            wb.save(path)
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))


def main():
    import logging
    logging.getLogger("customtkinter").setLevel(logging.ERROR)
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    root.title("Mass Breakdown")
    root.geometry("900x600")
    mod = MassBreakdownModule(root)
    mod.frame.pack(fill=tk.BOTH, expand=True)
    root.mainloop()


if __name__ == '__main__':
    main()
