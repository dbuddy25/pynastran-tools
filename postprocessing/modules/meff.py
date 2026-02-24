"""Modal effective mass fractions module.

Reads EFMFACS from an OP2 and displays per-mode fractions with
cumulative sums for each direction (Tx-Rz).
"""
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from tksheet import Sheet

import numpy as np
import scipy.sparse

DIRECTIONS = ['Tx', 'Ty', 'Tz', 'Rx', 'Ry', 'Rz']

# Header definitions for tksheet
_SINGLE_HEADERS = ['Mode', 'Freq (Hz)']
for _d in DIRECTIONS:
    _SINGLE_HEADERS.extend([f'{_d} Frac', f'{_d} Sum'])


def _matrix_to_dense(matrix_obj):
    """Convert a pyNastran Matrix object's data to a dense numpy array."""
    data = matrix_obj.data
    if scipy.sparse.issparse(data):
        return data.toarray()
    return np.asarray(data)


# --------------------------------------------------------- Excel helpers

def make_meff_styles():
    """Return a dict of openpyxl style objects for MEFFMASS Excel sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        'dark_fill': PatternFill("solid", fgColor="1F4E79"),
        'mid_fill': PatternFill("solid", fgColor="2E75B6"),
        'white_bold': Font(bold=True, color="FFFFFF", size=11),
        'sub_font': Font(bold=True, color="FFFFFF", size=10),
        'center': Alignment(horizontal="center", vertical="center"),
        'right': Alignment(horizontal="right", vertical="center"),
        'cell_border': Border(bottom=Side(style='thin', color="B4C6E7")),
        'bold_font': Font(bold=True),
        'num2': '0.00',
        'num1': '0.0',
    }


def write_meff_single_sheet(ws, data, styles, op2_name=None, threshold=0.1,
                            title=None):
    """Write a single-file MEFFMASS fraction sheet.

    When title is provided:
      Row 1 = custom title, Row 2 = OP2 filename, Row 3 = direction headers,
      Row 4 = sub-headers, Row 5+ = data.
    When title is None:
      Row 1 = OP2 filename, Row 2 = direction headers, Row 3 = sub-headers,
      Row 4+ = data.
    """
    from openpyxl.utils import get_column_letter

    s = styles
    modes, freqs = data['modes'], data['freqs']
    frac, cumsum = data['frac'], data['cumsum']

    sub = ['Mode', 'Freq (Hz)']
    for _ in DIRECTIONS:
        sub.extend(['Frac', 'Sum'])
    total_cols = len(sub)

    row_offset = 1 if title else 0

    # Row 1: custom title (only when title provided)
    if title:
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=total_cols)
        for ci in range(2, total_cols + 1):
            ws.cell(row=1, column=ci).fill = s['dark_fill']

    # OP2 filename row
    name_row = 1 + row_offset
    name_text = op2_name if op2_name else ""
    cell = ws.cell(row=name_row, column=1, value=name_text)
    cell.font = s['white_bold']
    cell.fill = s['dark_fill']
    cell.alignment = s['center']
    ws.merge_cells(start_row=name_row, start_column=1,
                   end_row=name_row, end_column=total_cols)
    for ci in range(2, total_cols + 1):
        ws.cell(row=name_row, column=ci).fill = s['dark_fill']

    # Direction group headers row
    dir_row = 2 + row_offset
    ws.cell(row=dir_row, column=1, value="").fill = s['dark_fill']
    ws.cell(row=dir_row, column=2, value="").fill = s['dark_fill']
    for idx, d in enumerate(DIRECTIONS):
        c1, c2 = 3 + idx * 2, 4 + idx * 2
        cell = ws.cell(row=dir_row, column=c1, value=d)
        cell.font = s['white_bold']
        cell.fill = s['dark_fill']
        cell.alignment = s['center']
        ws.merge_cells(start_row=dir_row, start_column=c1,
                       end_row=dir_row, end_column=c2)
        ws.cell(row=dir_row, column=c2).fill = s['dark_fill']

    # Sub-headers row
    sub_row = 3 + row_offset
    for ci, h in enumerate(sub, 1):
        cell = ws.cell(row=sub_row, column=ci, value=h)
        cell.font = s['sub_font']
        cell.fill = s['mid_fill']
        cell.alignment = s['center']

    # Data rows
    data_start = 4 + row_offset
    for i in range(len(modes)):
        row = i + data_start
        ws.cell(row=row, column=1, value=int(modes[i])).alignment = s['center']
        c = ws.cell(row=row, column=2, value=float(freqs[i]))
        c.number_format = s['num1']
        c.alignment = s['center']
        for j in range(6):
            fc = ws.cell(row=row, column=3 + j * 2, value=float(frac[i, j]))
            fc.number_format = s['num2']
            fc.alignment = s['center']
            if frac[i, j] >= threshold:
                fc.font = s['bold_font']
            sc = ws.cell(row=row, column=4 + j * 2, value=float(cumsum[i, j]))
            sc.number_format = s['num2']
            sc.alignment = s['center']
        for ci in range(1, total_cols + 1):
            ws.cell(row=row, column=ci).border = s['cell_border']

    # Column widths
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 12
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    freeze_row = data_start
    ws.freeze_panes = f'A{freeze_row}'


# ---------------------------------------------------------------- GUI module

class MeffModule:
    name = "MEFFMASS"

    def __init__(self, parent):
        self.frame = ctk.CTkFrame(parent)
        self.data = None
        self._op2_path = None
        self._threshold_var = tk.StringVar(value='0.1')
        self._title_var = tk.StringVar(value='')
        self._build_ui()

    # ------------------------------------------------------------------ UI
    _GUIDE_TEXT = """\
MEFFMASS Tool — Quick Guide

PURPOSE
Display modal effective mass fractions from a Nastran OP2 file. Shows
per-mode participation in each translational (Tx, Ty, Tz) and rotational
(Rx, Ry, Rz) direction along with cumulative sums.

WORKFLOW
1. Open OP2 — select a Nastran OP2 file containing MEFFMASS data.
   (Requires MEFFMASS(PLOT) = ALL in your case control deck.)
2. Review — the table shows Mode, Frequency, and Frac/Sum columns for
   each direction. Modes with fraction >= threshold are highlighted.
3. Adjust Threshold — change the highlight threshold (default 0.1) to
   focus on significant modes.
4. Export to Excel — save the table as a formatted .xlsx workbook.

THRESHOLD FILTER
Fraction values >= threshold are displayed in bold/blue. Changing the
threshold updates highlights in real time. This helps identify which
modes carry significant mass participation.

TITLE FIELD
Optional title text that appears as a header row in the Excel export.
Leave blank to omit.

EXCEL EXPORT
Produces a styled workbook with:
  - Merged direction group headers (Tx, Ty, ... Rz)
  - Bold formatting for values above threshold
  - Frozen header rows for easy scrolling
  - Color-coded header bands

REQUIREMENTS
  - pyNastran (for OP2 reading)
  - numpy, scipy
  - openpyxl (for Excel export only)\
"""

    def _build_ui(self):
        toolbar = ctk.CTkFrame(self.frame, fg_color="transparent")
        toolbar.pack(fill=tk.X, padx=5, pady=(5, 0))

        self._op2_btn = ctk.CTkButton(toolbar, text="Open OP2\u2026", width=100,
                                      command=self._open_op2)
        self._op2_btn.pack(side=tk.LEFT)

        # Title field
        ctk.CTkLabel(toolbar, text="Title:").pack(side=tk.LEFT, padx=(10, 2))
        ctk.CTkEntry(toolbar, textvariable=self._title_var, width=200).pack(
            side=tk.LEFT, padx=(0, 4))

        ctk.CTkButton(
            toolbar, text="?", width=30, font=ctk.CTkFont(weight="bold"),
            command=self._show_guide,
        ).pack(side=tk.RIGHT, padx=(5, 0))

        ctk.CTkButton(toolbar, text="Export to Excel\u2026", width=130,
                       command=self._export_excel).pack(side=tk.RIGHT)

        # Threshold entry (packed RIGHT so it sits left of Export button)
        ctk.CTkEntry(toolbar, width=50,
                      textvariable=self._threshold_var).pack(
            side=tk.RIGHT, padx=(0, 4))
        ctk.CTkLabel(toolbar, text="Threshold:").pack(
            side=tk.RIGHT, padx=(10, 2))

        self._threshold_var.trace_add('write', self._on_threshold_change)

        # Status label
        self._status_label = ctk.CTkLabel(
            toolbar, text="No OP2 loaded", text_color="gray")
        self._status_label.pack(side=tk.LEFT, padx=(10, 0))

        # Table (tksheet)
        self._sheet = Sheet(
            self.frame,
            headers=list(_SINGLE_HEADERS),
            show_top_left=False,
            show_row_index=False,
        )
        self._sheet.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._sheet.disable_bindings()
        self._sheet.enable_bindings(
            "single_select", "copy", "arrowkeys",
            "column_width_resize", "row_height_resize",
        )
        self._sheet.readonly_columns(
            columns=list(range(len(_SINGLE_HEADERS))))

    def _configure_sheet(self, headers):
        """Reconfigure sheet with new headers and clear data."""
        self._sheet.headers(headers)
        self._sheet.set_sheet_data([])
        self._sheet.readonly_columns(columns=list(range(len(headers))))

    # ---------------------------------------------------------- threshold helpers
    def _get_threshold(self):
        """Parse threshold from entry, default 0.1 on invalid input."""
        try:
            return float(self._threshold_var.get())
        except (ValueError, tk.TclError):
            return 0.1

    def _on_threshold_change(self, *args):
        """Update highlighting when threshold changes."""
        if self.data is not None:
            self._apply_highlights()

    def _apply_highlights(self):
        """Apply threshold-based cell highlighting to the current view."""
        self._sheet.dehighlight_all(redraw=False)
        threshold = self._get_threshold()

        if self.data is not None:
            frac = self.data['frac']
            for i in range(len(self.data['modes'])):
                for j in range(6):
                    col = 2 + j * 2  # Frac columns: indices 2, 4, 6, 8, 10, 12
                    if frac[i, j] >= threshold:
                        self._sheet.highlight_cells(row=i, column=col, fg="blue")

    # ---------------------------------------------------------- Guide
    def _show_guide(self):
        """Open the guide dialog (lazy import to avoid circular dependency)."""
        try:
            from nastran_tools import show_guide
        except ImportError:
            return
        show_guide(self.frame.winfo_toplevel(), "MEFFMASS Guide",
                   self._GUIDE_TEXT)

    # ---------------------------------------------------------- background work
    def _run_in_background(self, label, work_fn, done_fn):
        """Run *work_fn* in a background thread, keeping the UI responsive.

        *label* is shown in the status bar while the work runs.
        *done_fn(result, error)* is called on the main thread when finished.
        The Open OP2 button is disabled during execution.
        """
        self._status_label.configure(text=label, text_color="gray")
        self._op2_btn.configure(state=tk.DISABLED)

        container = {}  # mutable container for thread results

        def _worker():
            try:
                container['result'] = work_fn()
            except Exception as exc:
                container['error'] = exc

        def _poll():
            if thread.is_alive():
                self.frame.after(50, _poll)
            else:
                self._op2_btn.configure(state=tk.NORMAL)
                done_fn(container.get('result'), container.get('error'))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()
        self.frame.after(50, _poll)

    # ---------------------------------------------------------- OP2 loading
    def _open_op2(self):
        """Open a primary OP2 file."""
        path = filedialog.askopenfilename(
            title="Open OP2 File",
            filetypes=[("OP2 files", "*.op2"), ("All files", "*.*")])
        if not path:
            return

        def _work():
            from pyNastran.op2.op2 import OP2
            op2 = OP2(mode='nx')
            op2.read_op2(path)
            return op2

        def _done(op2, error):
            if error is not None:
                messagebox.showerror("Error",
                                     f"Could not read OP2:\n{error}")
                self._status_label.configure(text="Load failed",
                                             text_color="red")
                return

            self._op2_path = path
            self._status_label.configure(
                text=os.path.basename(path),
                text_color=("gray10", "gray90"))
            self.load(op2)

        self._run_in_background("Loading\u2026", _work, _done)

    # -------------------------------------------------------------- load
    def load(self, op2):
        """Populate table from OP2 data."""
        self._configure_sheet(list(_SINGLE_HEADERS))
        self.data = None

        if not op2.eigenvalues:
            messagebox.showwarning(
                "No Eigenvalues",
                "No eigenvalue data found in this OP2.\n\n"
                "Is this a SOL 103 run?")
            return

        eigval_table = next(iter(op2.eigenvalues.values()))
        modes = np.array(eigval_table.mode)
        freqs = np.array(eigval_table.cycles)

        meff_frac = op2.modal_effective_mass_fraction
        if meff_frac is None:
            messagebox.showwarning(
                "No MEFFMASS Data",
                "No MEFFMASS matrices found in this OP2.\n\n"
                "Add to your Nastran case control:\n"
                "  MEFFMASS(PLOT) = ALL")
            return

        raw = _matrix_to_dense(meff_frac)

        frac = raw.T  # (nmodes, 6)
        cumsum = np.cumsum(frac, axis=0)
        nmodes = min(frac.shape[0], len(modes))

        self.data = {
            'modes': modes[:nmodes],
            'freqs': freqs[:nmodes],
            'frac': frac[:nmodes],
            'cumsum': cumsum[:nmodes],
        }

        self._show_single_view()

    # ---------------------------------------------------------- view helpers
    def _show_single_view(self):
        self._configure_sheet(list(_SINGLE_HEADERS))
        if self.data is None:
            return
        modes, freqs = self.data['modes'], self.data['freqs']
        frac, cumsum = self.data['frac'], self.data['cumsum']
        rows = []
        for i in range(len(modes)):
            row = [int(modes[i]), f"{freqs[i]:.1f}"]
            for j in range(6):
                row.extend([f"{frac[i, j]:.2f}", f"{cumsum[i, j]:.2f}"])
            rows.append(row)
        self._sheet.set_sheet_data(rows)
        self._apply_highlights()

    # ------------------------------------------------------------ export
    def _export_excel(self):
        if self.data is None:
            messagebox.showinfo("Nothing to export",
                                "Open an OP2 file first.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return

        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for Excel export.\n\n"
                "pip install openpyxl")
            return

        name_a = os.path.basename(self._op2_path) if self._op2_path else None
        threshold = self._get_threshold()
        title = self._title_var.get().strip() or None

        wb = Workbook()
        styles = make_meff_styles()
        ws = wb.active
        ws.title = "Effective Mass Fractions"
        write_meff_single_sheet(ws, self.data, styles,
                                op2_name=name_a, threshold=threshold,
                                title=title)

        try:
            wb.save(path)
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))


def main():
    import logging
    logging.getLogger("customtkinter").setLevel(logging.ERROR)
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    root.title("MEFFMASS")
    root.geometry("1400x600")
    meff = MeffModule(root)
    meff.frame.pack(fill=tk.BOTH, expand=True)
    root.mainloop()


if __name__ == '__main__':
    main()
