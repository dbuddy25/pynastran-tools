"""FRF Response Limiting: compute a notched drive ASD from FRF data.

Loads a Nastran SOL 111 frequency-response OP2, an input environment ASD,
and a response limit ASD.  Node/direction rows are added to a grid; rows
with X/Y/Z set to 'S' (Show) or 'L' (Limit) are plotted, rows with 'L'
drive the notch calculation.  Results update live as inputs change.
"""

import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from tksheet import Sheet

from .asd_common import (
    RESPONSE_TYPES,
    subcase_options,
    lookup_subcase,
    parse_asd_text,
    parse_asd_file,
    interp_loglog,
    grms_loglog,
)


_DOF_NAMES = ["X", "Y", "Z"]
_DOF_COLS  = [2, 3, 4]   # column indices in the node grid

_NODE_COLORS = (
    "#1f77b4", "#2ca02c", "#9467bd", "#17becf", "#bcbd22",
    "#d62728", "#ff7f0e", "#8c564b", "#e377c2", "#7f7f7f",
)

_THEMES = {
    "dark": {
        "fig_bg": "#2b2b2b", "plot_bg": "#1e1e1e", "grid": "#3a3a3a",
        "text": "#c0c0c0", "spine": "#505050", "legend_bg": "#383838",
    },
    "light": {
        "fig_bg": "#f5f5f5", "plot_bg": "white", "grid": "#cccccc",
        "text": "#222222", "spine": "#888888", "legend_bg": "white",
    },
}

# Cell colours for S / L states
_COL_S  = {"bg": "#1f6aa5", "fg": "white"}   # show only
_COL_L  = {"bg": "#cc6600", "fg": "white"}   # limit (implies show)
_VALID  = {"", "S", "L"}


class ResponseLimitingModule:
    name = "Response Limiting"

    _GUIDE_TEXT = """\
Response Limiting: Quick Guide

PURPOSE
  Compute a notched drive ASD so that the response at selected nodes/DOFs
  stays at or below a specified limit spectrum.

WORKFLOW
  1. Open a SOL 111 FRF OP2 (ACCELERATION(PLOT,PHASE)=ALL required).
  2. Load or paste the Input ASD (freq vs g²/Hz).
  3. Load or paste the Response Limit ASD (freq vs g²/Hz).
     The Response Limit Paste button pre-fills with existing data for editing.
  4. Add nodes via Paste…, Import CSV…, or Add… and set X/Y/Z states.
     Type or use Tab/Enter to cycle states in the grid:
       S = Show:  response curve is plotted
       L = Limit: drives the notch AND is plotted (marked [L])
       (empty)  = not used
  5. Results update live; use Export to save.

NODE GRID
  Add / import format, one per line:  node_id  [label]
    e.g.  1001   or   1001 Tip mass   or   1001, Tip mass
    Nodes arrive with all directions set to L.
    Edit individual X / Y / Z cells in the grid to adjust.

PLOT THEME
  ☾ / ☀ button toggles plot background independent of the system theme.

CURVE PICKER  (Response: single curve view)
  Selects which node/direction to display. [L] marks Limit DOFs.

UNITS
  Set to match OP2 output units (in/s² for slinch/inch models).
  All ASDs are displayed in g²/Hz; GRMS in g.
"""

    def __init__(self, parent):
        self.frame = ctk.CTkFrame(parent)

        # ── State ─────────────────────────────────────────────────────────
        self._op2 = None
        self._op2_path = None
        self._subcase_opts = []
        self._subcase_var = ctk.StringVar(value="(none)")
        self._units_var = ctk.StringVar(value="in/s²")

        self._inputs: list = []        # [{name, freqs_raw, vals_raw, db}]
        self._active_input_idx: int = -1
        self._limit_asd_freqs = None
        self._limit_asd_vals  = None
        self._workmanship_freqs = None
        self._workmanship_vals  = None
        self._workmanship_envelope_var = tk.BooleanVar(value=False)

        self._plot_theme = "light"   # "light" or "dark"

        # Computed results (on FRF frequency grid)
        self._frf_freqs       = None
        self._orig_asd_interp = None
        self._limit_asd_interp = None
        self._notched_asd     = None
        self._response_curves  = {}   # {(nid, dof_idx): (resp_before, resp_after)}
        self._limit_dofs_set   = set()

        self._notch_enabled_var = ctk.BooleanVar(value=False)
        self._notch_db_var      = ctk.StringVar(value="6.0")
        self._scale_var         = ctk.StringVar(value="0.0")

        self._view_var       = ctk.StringVar(value="input")
        self._curve_var      = ctk.StringVar(value="(none)")   # single-curve picker

        self._title_var      = ctk.StringVar(value="")
        self._env_var        = ctk.StringVar(value="")
        self._limit_name_var = ctk.StringVar(value="Response Limit")
        self._yscale_var     = ctk.StringVar(value="Log")
        self._peak_label_var = ctk.StringVar(value="Freq only")

        self._aux_lines        = []
        self._pick_peaks_mode  = False
        self._picked_peaks     = []

        self._drawer_visible = False
        self._debounce_id    = None

        self._build_ui()

    # ── UI construction ────────────────────────────────────────────────────

    def _build_ui(self):
        toolbar = ctk.CTkFrame(self.frame, fg_color="transparent")
        toolbar.pack(fill=tk.X, padx=6, pady=(6, 2))

        # Row 0: OP2 + units + subcase + status + help
        row0 = ctk.CTkFrame(toolbar, fg_color="transparent")
        row0.pack(fill=tk.X, pady=1)

        self._open_btn = ctk.CTkButton(row0, text="Open OP2…", width=110,
                                       command=self._open_op2)
        self._open_btn.pack(side=tk.LEFT, padx=(0, 4))
        ctk.CTkButton(row0, text="Clear", width=60,
                      command=self._clear_op2).pack(side=tk.LEFT, padx=(0, 10))

        ctk.CTkLabel(row0, text="Run:").pack(side=tk.LEFT)
        self._file_label = ctk.CTkLabel(row0, text="(none)", text_color="gray",
                                        width=180, anchor=tk.W)
        self._file_label.pack(side=tk.LEFT, padx=4)

        ctk.CTkLabel(row0, text="Units:").pack(side=tk.LEFT, padx=(8, 2))
        ctk.CTkOptionMenu(row0, variable=self._units_var,
                          values=["in/s²", "m/s²"], width=90,
                          command=lambda _: self._schedule_recompute()).pack(side=tk.LEFT)

        ctk.CTkLabel(row0, text="Subcase:").pack(side=tk.LEFT, padx=(8, 2))
        self._sc_menu = ctk.CTkOptionMenu(row0, variable=self._subcase_var,
                                          values=["(none)"],
                                          command=self._on_subcase_change,
                                          width=160)
        self._sc_menu.pack(side=tk.LEFT)

        self._status_label = ctk.CTkLabel(row0, text="Load an FRF OP2 to begin",
                                          text_color="gray")
        self._status_label.pack(side=tk.LEFT, padx=(12, 0))

        ctk.CTkButton(row0, text="?", width=28,
                      command=self._open_help).pack(side=tk.RIGHT)

        # Row 1: ASD loaders
        row1 = ctk.CTkFrame(toolbar, fg_color="transparent")
        row1.pack(fill=tk.X, pady=1)

        inp = ctk.CTkFrame(row1, fg_color="transparent")
        inp.pack(side=tk.LEFT, padx=(0, 16))
        ctk.CTkLabel(inp, text="Input ASD:", width=72, anchor=tk.W).pack(side=tk.LEFT)
        ctk.CTkButton(inp, text="Load…",  width=60, command=self._load_input_asd).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(inp, text="Paste…", width=60, command=self._paste_input_asd).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(inp, text="Remove", width=68, command=self._remove_active_input).pack(side=tk.LEFT, padx=2)
        self._input_selector = ctk.CTkOptionMenu(
            inp, values=["(none)"], width=200,
            command=self._on_input_selected)
        self._input_selector.pack(side=tk.LEFT, padx=4)

        lim = ctk.CTkFrame(row1, fg_color="transparent")
        lim.pack(side=tk.LEFT)
        ctk.CTkLabel(lim, text="Response Limit:", width=104, anchor=tk.W).pack(side=tk.LEFT)
        ctk.CTkButton(lim, text="Load…",  width=60, command=self._load_limit_asd).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(lim, text="Paste…", width=60, command=self._paste_limit_asd).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(lim, text="Clear",  width=50, command=self._clear_limit_asd).pack(side=tk.LEFT, padx=2)
        self._limit_status = ctk.CTkLabel(lim, text="(none)", text_color="gray",
                                          width=200, anchor=tk.W)
        self._limit_status.pack(side=tk.LEFT, padx=4)

        # Row 1b: Workmanship spec
        row1b = ctk.CTkFrame(toolbar, fg_color="transparent")
        row1b.pack(fill=tk.X, pady=1)
        wm_frame = ctk.CTkFrame(row1b, fg_color="transparent")
        wm_frame.pack(side=tk.LEFT, padx=(0, 16))
        ctk.CTkLabel(wm_frame, text="Workmanship:", width=92, anchor=tk.W).pack(side=tk.LEFT)
        ctk.CTkButton(wm_frame, text="Load…",  width=60, command=self._load_workmanship).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(wm_frame, text="Paste…", width=60, command=self._paste_workmanship).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(wm_frame, text="Clear",  width=50, command=self._clear_workmanship).pack(side=tk.LEFT, padx=2)
        ctk.CTkCheckBox(wm_frame, text="Envelope input",
                        variable=self._workmanship_envelope_var,
                        command=self._on_workmanship_toggle).pack(side=tk.LEFT, padx=(10, 4))
        self._workmanship_status = ctk.CTkLabel(wm_frame, text="(none)", text_color="gray",
                                                width=180, anchor=tk.W)
        self._workmanship_status.pack(side=tk.LEFT, padx=4)

        # Row 2: notch floor
        row2 = ctk.CTkFrame(toolbar, fg_color="transparent")
        row2.pack(fill=tk.X, pady=1)

        ctk.CTkCheckBox(row2, text="Max notch depth (dB):",
                        variable=self._notch_enabled_var,
                        command=self._on_notch_toggle).pack(side=tk.LEFT)
        self._notch_entry = ctk.CTkEntry(row2, textvariable=self._notch_db_var,
                                          width=54, state=tk.DISABLED)
        self._notch_entry.pack(side=tk.LEFT, padx=(4, 20))
        self._notch_db_var.trace_add("write", lambda *_: self._schedule_recompute())

        ctk.CTkLabel(row2, text="Input scale (dB):").pack(side=tk.LEFT)
        ctk.CTkEntry(row2, textvariable=self._scale_var, width=54).pack(side=tk.LEFT, padx=(4, 20))
        self._scale_var.trace_add("write", self._on_scale_change)

        self._export_excel_btn = ctk.CTkButton(row2, text="Export Excel…", width=110,
                                                command=self._export_excel, state=tk.DISABLED)
        self._export_excel_btn.pack(side=tk.LEFT, padx=(0, 4))
        self._export_csv_btn = ctk.CTkButton(row2, text="Export CSV…", width=90,
                                              command=self._export_csv, state=tk.DISABLED)
        self._export_csv_btn.pack(side=tk.LEFT)

        # Row 3: title / environment / limit name
        row3 = ctk.CTkFrame(toolbar, fg_color="transparent")
        row3.pack(fill=tk.X, pady=1)
        ctk.CTkLabel(row3, text="Title:").pack(side=tk.LEFT)
        ctk.CTkEntry(row3, textvariable=self._title_var, width=200,
                     placeholder_text="Plot title").pack(side=tk.LEFT, padx=(4, 12))
        ctk.CTkLabel(row3, text="Env:").pack(side=tk.LEFT)
        ctk.CTkEntry(row3, textvariable=self._env_var, width=180,
                     placeholder_text="Environment / subtitle").pack(side=tk.LEFT, padx=(4, 12))
        ctk.CTkLabel(row3, text="Limit name:").pack(side=tk.LEFT)
        ctk.CTkEntry(row3, textvariable=self._limit_name_var, width=140).pack(side=tk.LEFT, padx=4)
        for v in (self._title_var, self._env_var, self._limit_name_var):
            v.trace_add("write", lambda *_: self._redraw())

        # Row 4: annotations / peak picking / session
        row4 = ctk.CTkFrame(toolbar, fg_color="transparent")
        row4.pack(fill=tk.X, pady=1)
        ctk.CTkLabel(row4, text="Y-axis:").pack(side=tk.LEFT)
        ctk.CTkOptionMenu(row4, variable=self._yscale_var, values=["Log", "Linear"],
                          width=80, command=lambda _: self._redraw()).pack(side=tk.LEFT, padx=(4, 12))
        self._yscale_var.trace_add("write", lambda *_: None)  # handled by command above

        self._pick_btn = ctk.CTkButton(row4, text="Pick Peaks", width=96,
                                       command=self._toggle_pick_mode)
        self._pick_btn.pack(side=tk.LEFT, padx=(0, 4))
        ctk.CTkButton(row4, text="Clear Peaks", width=90,
                      command=self._clear_peaks).pack(side=tk.LEFT, padx=(0, 12))

        ctk.CTkLabel(row4, text="Label:").pack(side=tk.LEFT)
        ctk.CTkOptionMenu(row4, variable=self._peak_label_var,
                          values=["Freq only", "Freq + value"],
                          width=110, command=lambda _: self._redraw()).pack(side=tk.LEFT, padx=(4, 12))

        ctk.CTkButton(row4, text="X/Y Lines…", width=90,
                      command=self._aux_lines_dialog).pack(side=tk.LEFT, padx=(0, 12))
        ctk.CTkButton(row4, text="Save Session…", width=110,
                      command=self._save_session).pack(side=tk.LEFT, padx=(0, 4))
        ctk.CTkButton(row4, text="Open Session…", width=110,
                      command=self._open_session).pack(side=tk.LEFT)

        # ── Body ──────────────────────────────────────────────────────────
        body = ctk.CTkFrame(self.frame, fg_color="transparent")
        body.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)

        # Left panel
        self._left = ctk.CTkScrollableFrame(body, width=320)
        self._left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 6))

        # Node grid section
        ctk.CTkLabel(self._left, text="Nodes  (S = Show, L = Limit)",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W, pady=(4, 2))

        node_btns = ctk.CTkFrame(self._left, fg_color="transparent")
        node_btns.pack(fill=tk.X, pady=(0, 4))
        ctk.CTkButton(node_btns, text="Add…",       width=62,
                      command=self._add_node_dialog).pack(side=tk.LEFT, padx=(0, 2))
        ctk.CTkButton(node_btns, text="Import CSV…",width=90,
                      command=self._import_nodes_csv).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(node_btns, text="Remove",     width=68,
                      command=self._remove_selected_nodes).pack(side=tk.LEFT, padx=2)
        ctk.CTkButton(node_btns, text="Clear All",  width=72,
                      command=self._clear_all_nodes).pack(side=tk.LEFT, padx=2)

        self._node_sheet = Sheet(
            self._left,
            headers=["Node", "Label", "X", "Y", "Z"],
            height=200,
            show_row_index=False,
            theme="dark" if ctk.get_appearance_mode() == "Dark" else "light",
        )
        self._node_sheet.pack(fill=tk.X, pady=(0, 6))
        self._node_sheet.enable_bindings(
            "single_select", "row_select", "edit_cell", "column_width_resize")
        self._node_sheet.column_width(0, 56)
        self._node_sheet.column_width(1, 100)
        self._node_sheet.column_width(2, 40)
        self._node_sheet.column_width(3, 40)
        self._node_sheet.column_width(4, 40)
        self._node_sheet.extra_bindings([("end_edit_cell", self._on_grid_edit)])

        # Separator
        ctk.CTkFrame(self._left, height=1, fg_color="gray40").pack(
            fill=tk.X, pady=(2, 6))

        # View section
        self._view_section = ctk.CTkFrame(self._left, fg_color="transparent")
        self._view_section.pack(fill=tk.X)
        ctk.CTkLabel(self._view_section, text="View",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W, pady=(0, 2))

        for val, label in [
            ("input",        "Input ASDs (orig / notched)"),
            ("response_dof", "Response: single curve"),
            ("response_all", "All responses overlay"),
            ("grms",         "GRMS Summary"),
        ]:
            ctk.CTkRadioButton(self._view_section, text=label, variable=self._view_var,
                               value=val, command=self._redraw).pack(anchor=tk.W, pady=1)

        curve_row = ctk.CTkFrame(self._view_section, fg_color="transparent")
        curve_row.pack(fill=tk.X, pady=(4, 0))
        ctk.CTkLabel(curve_row, text="Curve:").pack(side=tk.LEFT)
        self._curve_menu = ctk.CTkOptionMenu(
            curve_row, variable=self._curve_var,
            values=["(none)"], command=lambda _: self._redraw(), width=150)
        self._curve_menu.pack(side=tk.LEFT, padx=4)

        # Drawer (right side, initially hidden) — must be packed before canvas frame
        self._drawer = ctk.CTkFrame(body, width=290)
        # not packed until toggled

        # Right: canvas + theme button + GRMS overlay
        right = ctk.CTkFrame(body, fg_color="transparent")
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Plot header with theme + drawer + split view toggles
        plot_hdr = ctk.CTkFrame(right, fg_color="transparent")
        plot_hdr.pack(fill=tk.X)
        self._theme_btn = ctk.CTkButton(
            plot_hdr, text="☾ Dark", width=80,
            command=self._toggle_theme)
        self._theme_btn.pack(side=tk.RIGHT, padx=4, pady=2)
        self._drawer_btn = ctk.CTkButton(
            plot_hdr, text="Tables ▶", width=84,
            command=self._toggle_drawer)
        self._drawer_btn.pack(side=tk.RIGHT, padx=(0, 2), pady=2)
        self._split_btn = ctk.CTkButton(
            plot_hdr, text="⊞ Split", width=80,
            command=self._toggle_split)
        self._split_btn.pack(side=tk.RIGHT, padx=(0, 2), pady=2)

        # Pane container (holds 1 or 2 stacked panes)
        self._plots_container = ctk.CTkFrame(right, fg_color="transparent")
        self._plots_container.pack(fill=tk.BOTH, expand=True)

        # Primary pane — shares self._view_var with sidebar radios
        self._panes = [self._make_pane(self._plots_container, view_var=self._view_var)]
        self._panes[0]["frame"].pack(fill=tk.BOTH, expand=True)

        # Drawer content
        self._build_drawer()

        self._draw_idle_plot()

    # ── Pane factory ──────────────────────────────────────────────────────

    _VIEW_KEY_TO_LABEL = {
        "input": "Input ASDs", "response_dof": "Response: single",
        "response_all": "All responses", "grms": "GRMS Summary",
    }
    _VIEW_LABEL_TO_KEY = {v: k for k, v in _VIEW_KEY_TO_LABEL.items()}

    def _make_pane(self, container, view_var=None):
        """Create and return a plot-pane dict. view_var is shared with sidebar radios for pane 0."""
        if view_var is None:
            view_var = ctk.StringVar(value="input")

        frame = ctk.CTkFrame(container, fg_color="transparent")

        # Pane header: view selector (visible in split mode) + Copy Figure
        hdr = ctk.CTkFrame(frame, fg_color="transparent")
        hdr.pack(fill=tk.X)

        # view_label_var holds the display string; synced bidirectionally with view_var (keys)
        view_label_var = ctk.StringVar(
            value=self._VIEW_KEY_TO_LABEL.get(view_var.get(), "Input ASDs"))
        view_menu = ctk.CTkOptionMenu(
            hdr, variable=view_label_var,
            values=list(self._VIEW_KEY_TO_LABEL.values()),
            width=150,
            command=lambda lbl, vv=view_var: vv.set(self._VIEW_LABEL_TO_KEY.get(lbl, "input")))
        view_menu.pack(side=tk.LEFT, padx=4, pady=2)
        view_menu.pack_forget()  # hidden until split mode

        copy_btn = ctk.CTkButton(hdr, text="Copy Figure", width=100,
                                 command=lambda: None)
        copy_btn.pack(side=tk.RIGHT, padx=4, pady=2)

        fig = Figure(figsize=(8, 4))
        ax = fig.add_subplot(111)
        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, frame).update()

        grms_frame = ctk.CTkFrame(frame)
        grms_sheet = Sheet(
            grms_frame,
            headers=["DOF", "Orig Input GRMS (g)", "Notched Input GRMS (g)",
                     "Resp @ Orig Input (GRMS)", "Resp @ Notched Input (GRMS)", "Max Notch (dB)"],
            height=300, show_row_index=False,
        )
        grms_sheet.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        grms_sheet.enable_bindings("column_width_resize")

        pane = {
            "frame": frame, "hdr": hdr,
            "fig": fig, "ax": ax,
            "canvas": canvas, "canvas_widget": canvas_widget,
            "view_var": view_var, "view_label_var": view_label_var, "view_menu": view_menu,
            "copy_btn": copy_btn,
            "grms_frame": grms_frame, "grms_sheet": grms_sheet,
            "last_drawn_curves": [],
        }

        copy_btn.configure(command=lambda p=pane: self._copy_figure(p))
        canvas.mpl_connect("button_press_event",
                           lambda event, p=pane: self._on_canvas_click(event, p))

        # Sync view_label_var when view_var changes (e.g. sidebar radio click)
        def _sync_label(*_, vv=view_var, lv=view_label_var):
            lv.set(self._VIEW_KEY_TO_LABEL.get(vv.get(), lv.get()))
        view_var.trace_add("write", _sync_label)

        # Trigger redraw when view_label_var changes (covers both menu selection and sidebar)
        view_label_var.trace_add("write", lambda *_, p=pane: self._redraw_pane(p))
        return pane

    # ── Drawer ────────────────────────────────────────────────────────────

    def _build_drawer(self):
        ctk.CTkLabel(self._drawer, text="Data Tables",
                     font=ctk.CTkFont(weight="bold")).pack(anchor=tk.W, padx=8, pady=(8, 4))

        ctk.CTkLabel(self._drawer, text="Input ASD",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor=tk.W, padx=8, pady=(4, 2))
        self._input_tbl = Sheet(
            self._drawer,
            headers=["Freq (Hz)", "ASD (g²/Hz)"],
            height=220, show_row_index=False,
            theme="dark" if ctk.get_appearance_mode() == "Dark" else "light",
        )
        self._input_tbl.pack(fill=tk.X, padx=4, pady=(0, 6))
        self._input_tbl.enable_bindings("edit_cell", "column_width_resize", "single_select")
        self._input_tbl.column_width(0, 90)
        self._input_tbl.column_width(1, 100)
        self._input_tbl.extra_bindings([("end_edit_cell", self._on_input_tbl_edit)])

        ctk.CTkFrame(self._drawer, height=1, fg_color="gray40").pack(fill=tk.X, padx=4, pady=4)

        ctk.CTkLabel(self._drawer, text="Response Limit",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor=tk.W, padx=8, pady=(4, 2))
        self._limit_tbl = Sheet(
            self._drawer,
            headers=["Freq (Hz)", "Limit (g²/Hz)"],
            height=220, show_row_index=False,
            theme="dark" if ctk.get_appearance_mode() == "Dark" else "light",
        )
        self._limit_tbl.pack(fill=tk.X, padx=4, pady=(0, 6))
        self._limit_tbl.enable_bindings("edit_cell", "column_width_resize", "single_select")
        self._limit_tbl.column_width(0, 90)
        self._limit_tbl.column_width(1, 100)
        self._limit_tbl.extra_bindings([("end_edit_cell", self._on_limit_tbl_edit)])

        ctk.CTkFrame(self._drawer, height=1, fg_color="gray40").pack(fill=tk.X, padx=4, pady=4)

        ctk.CTkLabel(self._drawer, text="Workmanship",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor=tk.W, padx=8, pady=(4, 2))
        self._workmanship_tbl = Sheet(
            self._drawer,
            headers=["Freq (Hz)", "ASD (g²/Hz)"],
            height=180, show_row_index=False,
            theme="dark" if ctk.get_appearance_mode() == "Dark" else "light",
        )
        self._workmanship_tbl.pack(fill=tk.X, padx=4, pady=(0, 6))
        self._workmanship_tbl.enable_bindings("edit_cell", "column_width_resize", "single_select")
        self._workmanship_tbl.column_width(0, 90)
        self._workmanship_tbl.column_width(1, 100)
        self._workmanship_tbl.extra_bindings([("end_edit_cell", self._on_workmanship_tbl_edit)])

    def _toggle_drawer(self):
        if self._drawer_visible:
            self._drawer.pack_forget()
            self._drawer_btn.configure(text="Tables ▶")
            self._drawer_visible = False
        else:
            self._drawer.pack(side=tk.RIGHT, fill=tk.Y)
            self._drawer_btn.configure(text="Tables ◀")
            self._drawer_visible = True
            self._populate_drawer_tables()

    def _populate_drawer_tables(self):
        inp = self._active_input()
        if inp is not None:
            self._input_tbl.set_sheet_data(
                [[f"{f:.5g}", f"{a:.6g}"]
                 for f, a in zip(inp['freqs_raw'], inp['vals_raw'])])
        else:
            self._input_tbl.set_sheet_data([])
        if self._limit_asd_freqs is not None:
            self._limit_tbl.set_sheet_data(
                [[f"{f:.5g}", f"{a:.6g}"]
                 for f, a in zip(self._limit_asd_freqs, self._limit_asd_vals)])
        else:
            self._limit_tbl.set_sheet_data([])
        if self._workmanship_freqs is not None:
            self._workmanship_tbl.set_sheet_data(
                [[f"{f:.5g}", f"{a:.6g}"]
                 for f, a in zip(self._workmanship_freqs, self._workmanship_vals)])
        else:
            self._workmanship_tbl.set_sheet_data([])

    def _on_input_tbl_edit(self, event):
        inp = self._active_input()
        if inp is None:
            return
        freqs, asds = [], []
        for row in self._input_tbl.get_sheet_data():
            try:
                freqs.append(float(row[0]))
                asds.append(float(row[1]))
            except (ValueError, IndexError, TypeError):
                continue
        if len(freqs) < 2:
            return
        order = np.argsort(freqs)
        inp['freqs_raw'] = np.array(freqs)[order]
        inp['vals_raw']  = np.array(asds)[order]
        inp['name'] = self._asd_status_text("(edited)", inp['freqs_raw'])
        self._refresh_input_selector()
        self._clear_results()
        self._schedule_recompute()

    def _on_limit_tbl_edit(self, event):
        freqs, asds = [], []
        for row in self._limit_tbl.get_sheet_data():
            try:
                freqs.append(float(row[0]))
                asds.append(float(row[1]))
            except (ValueError, IndexError, TypeError):
                continue
        if len(freqs) < 2:
            return
        order = np.argsort(freqs)
        self._limit_asd_freqs = np.array(freqs)[order]
        self._limit_asd_vals  = np.array(asds)[order]
        self._limit_status.configure(
            text=self._asd_status_text("(edited)", self._limit_asd_freqs),
            text_color=("gray10", "gray90"))
        self._clear_results()
        self._schedule_recompute()

    # ── Theme toggle ──────────────────────────────────────────────────────

    def _toggle_theme(self):
        self._plot_theme = "light" if self._plot_theme == "dark" else "dark"
        self._theme_btn.configure(
            text="☾ Dark" if self._plot_theme == "light" else "☀ Light")
        self._redraw()

    def _copy_figure(self, pane=None):
        import io, os, tempfile, subprocess
        if pane is None:
            pane = self._panes[0]
        fig = pane["fig"]
        buf = io.BytesIO()
        try:
            fig.savefig(buf, format='png', dpi=200, bbox_inches='tight',
                        facecolor=fig.get_facecolor())
        except Exception as exc:
            messagebox.showerror("Copy Error", f"Could not render figure:\n{exc}")
            return
        buf.seek(0)
        tmp = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
                f.write(buf.getvalue())
                tmp = f.name
            if os.name == 'nt':
                ps = (
                    'Add-Type -Assembly System.Windows.Forms;'
                    '[Windows.Forms.Clipboard]::SetImage('
                    f'[System.Drawing.Image]::FromFile("{tmp}"))'
                )
                subprocess.run(['powershell', '-Command', ps], check=True)
            elif os.uname().sysname == 'Darwin':
                subprocess.run(
                    ['osascript', '-e',
                     f'set the clipboard to '
                     f'(read (POSIX file "{tmp}") as «class PNGf»)'],
                    check=True)
            else:
                subprocess.run(
                    ['xclip', '-selection', 'clipboard',
                     '-t', 'image/png', '-i', tmp],
                    check=True)
        except Exception as exc:
            messagebox.showerror("Copy Error", str(exc))
        finally:
            if tmp:
                try:
                    os.unlink(tmp)
                except Exception:
                    pass

    # ── Split view toggle ─────────────────────────────────────────────────

    def _toggle_split(self):
        if len(self._panes) == 1:
            # Enter split mode
            pane2 = self._make_pane(self._plots_container)
            self._panes.append(pane2)
            pane2["frame"].pack(fill=tk.BOTH, expand=True)
            # Show view menu in both pane headers
            for pane in self._panes:
                pane["view_menu"].pack(side=tk.LEFT, padx=4, pady=2)
            # Hide sidebar view radios (pane 0's view_menu takes over for pane 0)
            self._view_section.pack_forget()
            self._split_btn.configure(text="◧ Single")
            self._draw_idle_plot()
            self._redraw()
        else:
            # Exit split mode — destroy pane 2
            pane2 = self._panes.pop()
            pane2["frame"].destroy()
            # Hide view menu in pane 0 header, re-show sidebar
            self._panes[0]["view_menu"].pack_forget()
            self._view_section.pack(fill=tk.X)
            self._split_btn.configure(text="⊞ Split")
            self._redraw()

    # ── Help ──────────────────────────────────────────────────────────────

    def _open_help(self):
        win = ctk.CTkToplevel(self.frame.winfo_toplevel())
        win.title("Response Limiting: Guide")
        win.geometry("560x500")
        win.resizable(True, True)
        win.transient(self.frame.winfo_toplevel())
        tb = ctk.CTkTextbox(win, wrap="word")
        tb.pack(fill="both", expand=True, padx=10, pady=(10, 5))
        tb.insert("1.0", self._GUIDE_TEXT)
        tb.configure(state="disabled")
        ctk.CTkButton(win, text="Close", width=80,
                      command=win.destroy).pack(pady=(0, 10))

    # ── Notch floor toggle ────────────────────────────────────────────────

    def _on_notch_toggle(self):
        self._notch_entry.configure(
            state=tk.NORMAL if self._notch_enabled_var.get() else tk.DISABLED)
        self._schedule_recompute()

    # ── Live compute ──────────────────────────────────────────────────────

    def _schedule_recompute(self, delay=400):
        if self._debounce_id is not None:
            self.frame.after_cancel(self._debounce_id)
        self._debounce_id = self.frame.after(delay, self._auto_compute)

    def _auto_compute(self):
        self._debounce_id = None
        if (self._op2 is not None
                and self._active_input() is not None
                and self._limit_asd_freqs is not None
                and self._any_limit_dofs()):
            self._compute_notch(silent=True)

    def _any_limit_dofs(self):
        data = self._node_sheet.get_sheet_data()
        for row in data:
            for col in _DOF_COLS:
                if str(row[col]).upper() == "L":
                    return True
        return False

    # ── Background threading ───────────────────────────────────────────────

    def _run_in_background(self, label, work_fn, done_fn):
        self._status_label.configure(text=label, text_color="gray")
        self._open_btn.configure(state=tk.DISABLED)
        container = {}

        def _worker():
            try:
                container['result'] = work_fn()
            except Exception as exc:
                container['error'] = exc

        def _poll():
            if thread.is_alive():
                self.frame.after(50, _poll)
            else:
                self._open_btn.configure(state=tk.NORMAL)
                done_fn(container.get('result'), container.get('error'))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()
        self.frame.after(50, _poll)

    # ── OP2 loading ────────────────────────────────────────────────────────

    def _open_op2(self):
        path = filedialog.askopenfilename(
            title="Open FRF OP2",
            filetypes=[("OP2 files", "*.op2"), ("All files", "*.*")])
        if not path:
            return

        def _work():
            from pyNastran.op2.op2 import OP2
            op2 = OP2(mode='nx', debug=False)
            op2.read_op2(path)
            return op2

        def _done(op2, error):
            if error is not None:
                messagebox.showerror("Error", f"Could not read OP2:\n{error}")
                self._status_label.configure(text="Load failed", text_color="red")
                return

            cfg = RESPONSE_TYPES['Acceleration']
            frf_dict = getattr(op2, cfg['frf_attr'], None) or {}
            if not frf_dict:
                messagebox.showwarning(
                    "No FRF Data",
                    "This OP2 has no frequency-response acceleration data.\n\n"
                    "The deck must include:\n  ACCELERATION(PLOT,PHASE) = ALL")
                self._file_label.configure(text="(no FRF data)", text_color="orange")
                return

            self._op2 = op2
            self._op2_path = path

            sc_pairs = subcase_options(frf_dict)
            self._subcase_opts = sc_pairs
            labels = [lbl for _, lbl in sc_pairs]
            self._sc_menu.configure(values=labels)
            self._subcase_var.set(labels[0] if labels else "(none)")

            stem = os.path.splitext(os.path.basename(path))[0]
            self._file_label.configure(text=stem, text_color=("gray10", "gray90"))
            self._status_label.configure(
                text=f"{os.path.basename(path)}: {len(sc_pairs)} subcase(s)",
                text_color=("gray10", "gray90"))
            self._clear_results()
            self._schedule_recompute()

        self._run_in_background("Loading OP2…", _work, _done)

    def _clear_op2(self):
        self._op2 = None
        self._op2_path = None
        self._subcase_opts = []
        self._subcase_var.set("(none)")
        self._sc_menu.configure(values=["(none)"])
        self._file_label.configure(text="(none)", text_color="gray")
        self._status_label.configure(text="Load an FRF OP2 to begin", text_color="gray")
        self._clear_results()
        self._draw_idle_plot()

    def _on_subcase_change(self, _=None):
        self._clear_results()
        self._schedule_recompute()

    def _get_subcase_int(self):
        label = self._subcase_var.get()
        for sc_id, lbl in self._subcase_opts:
            if lbl == label:
                return sc_id
        return None

    # ── ASD loaders ───────────────────────────────────────────────────────

    @staticmethod
    def _asd_status_text(label, freqs):
        return f"{label} ({len(freqs)} pts, {freqs[0]:.1f}–{freqs[-1]:.1f} Hz)"

    def _load_input_asd(self):
        path = filedialog.askopenfilename(
            title="Load Input ASD",
            filetypes=[("Text/CSV", "*.txt *.csv"), ("All files", "*.*")])
        if not path:
            return
        try:
            freqs, asds, file_name = parse_asd_file(path)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Load Error", str(exc))
            return
        name = file_name or self._asd_status_text(os.path.basename(path), freqs)
        self._inputs.append({"name": name, "freqs_raw": freqs, "vals_raw": asds, "db": "0.0"})
        self._active_input_idx = len(self._inputs) - 1
        self._refresh_input_selector()
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _paste_input_asd(self):
        text = self._paste_dialog(
            "Paste Input ASD", "Paste 2-column data (freq  g²/Hz), one row per line:")
        if text is None:
            return
        freqs, asds, text_name = parse_asd_text(text)
        if freqs is None:
            messagebox.showerror("Parse Error", "Need at least 2 valid frequency rows.")
            return
        if text_name:
            name = text_name
        else:
            n = sum(1 for e in self._inputs if e['name'].startswith("(pasted)"))
            name = self._asd_status_text(f"(pasted {n + 1})" if n else "(pasted)", freqs)
        self._inputs.append({"name": name, "freqs_raw": freqs, "vals_raw": asds, "db": "0.0"})
        self._active_input_idx = len(self._inputs) - 1
        self._refresh_input_selector()
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _remove_active_input(self):
        if not self._inputs:
            return
        if 0 <= self._active_input_idx < len(self._inputs):
            self._inputs.pop(self._active_input_idx)
        self._active_input_idx = min(self._active_input_idx, len(self._inputs) - 1)
        self._refresh_input_selector()
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    # ── Input helpers ─────────────────────────────────────────────────────

    def _active_input(self):
        if 0 <= self._active_input_idx < len(self._inputs):
            return self._inputs[self._active_input_idx]
        return None

    def _refresh_input_selector(self):
        if self._inputs:
            names = [e['name'] for e in self._inputs]
            self._input_selector.configure(values=names)
            idx = max(0, min(self._active_input_idx, len(self._inputs) - 1))
            self._active_input_idx = idx
            self._input_selector.set(names[idx])
            # Load this input's saved dB into scale entry
            self._scale_var.set(self._inputs[idx].get('db', '0.0'))
        else:
            self._active_input_idx = -1
            self._input_selector.configure(values=["(none)"])
            self._input_selector.set("(none)")
            self._scale_var.set("0.0")

    def _on_input_selected(self, name):
        for i, inp in enumerate(self._inputs):
            if inp['name'] == name:
                self._active_input_idx = i
                self._scale_var.set(inp.get('db', '0.0'))
                break
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _on_scale_change(self, *_):
        inp = self._active_input()
        if inp is not None:
            inp['db'] = self._scale_var.get()
        self._schedule_recompute()

    def _effective_input(self, freqs, scale):
        """Return interpolated + scaled effective input; applies workmanship envelope when toggled."""
        inp = self._active_input()
        base = interp_loglog(inp['freqs_raw'], inp['vals_raw'], freqs) * scale
        if self._workmanship_envelope_var.get() and self._workmanship_vals is not None:
            wm = interp_loglog(self._workmanship_freqs, self._workmanship_vals, freqs)
            return np.maximum(base, wm)
        return base

    # ── Workmanship spec ──────────────────────────────────────────────────

    def _load_workmanship(self):
        path = filedialog.askopenfilename(
            title="Load Workmanship Spec",
            filetypes=[("Text/CSV", "*.txt *.csv"), ("All files", "*.*")])
        if not path:
            return
        try:
            freqs, asds, _ = parse_asd_file(path)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Load Error", str(exc))
            return
        self._workmanship_freqs, self._workmanship_vals = freqs, asds
        self._workmanship_status.configure(
            text=self._asd_status_text(os.path.basename(path), freqs),
            text_color=("gray10", "gray90"))
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _paste_workmanship(self):
        initial = ""
        if self._workmanship_freqs is not None:
            initial = "\n".join(f"{f:.5g}  {a:.6g}"
                                for f, a in zip(self._workmanship_freqs, self._workmanship_vals))
        text = self._paste_dialog(
            "Paste / Edit Workmanship Spec",
            "2-column data (freq  g²/Hz), one row per line:",
            initial=initial)
        if text is None:
            return
        freqs, asds, _ = parse_asd_text(text)
        if freqs is None:
            messagebox.showerror("Parse Error", "Need at least 2 valid frequency rows.")
            return
        self._workmanship_freqs, self._workmanship_vals = freqs, asds
        self._workmanship_status.configure(
            text=self._asd_status_text("(pasted)", freqs),
            text_color=("gray10", "gray90"))
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _clear_workmanship(self):
        self._workmanship_freqs = self._workmanship_vals = None
        self._workmanship_status.configure(text="(none)", text_color="gray")
        if self._drawer_visible:
            self._workmanship_tbl.set_sheet_data([])
        self._clear_results()
        self._schedule_recompute()

    def _on_workmanship_toggle(self):
        self._clear_results()
        self._schedule_recompute()

    def _on_workmanship_tbl_edit(self, event):
        freqs, asds = [], []
        for row in self._workmanship_tbl.get_sheet_data():
            try:
                freqs.append(float(row[0]))
                asds.append(float(row[1]))
            except (ValueError, IndexError, TypeError):
                continue
        if len(freqs) < 2:
            return
        order = np.argsort(freqs)
        self._workmanship_freqs = np.array(freqs)[order]
        self._workmanship_vals  = np.array(asds)[order]
        self._workmanship_status.configure(
            text=self._asd_status_text("(edited)", self._workmanship_freqs),
            text_color=("gray10", "gray90"))
        self._clear_results()
        self._schedule_recompute()

    def _load_limit_asd(self):
        path = filedialog.askopenfilename(
            title="Load Response Limit ASD",
            filetypes=[("Text/CSV", "*.txt *.csv"), ("All files", "*.*")])
        if not path:
            return
        try:
            freqs, asds, _ = parse_asd_file(path)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Load Error", str(exc))
            return
        self._limit_asd_freqs, self._limit_asd_vals = freqs, asds
        self._limit_status.configure(
            text=self._asd_status_text(os.path.basename(path), freqs),
            text_color=("gray10", "gray90"))
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _paste_limit_asd(self):
        # Pre-populate with current limit data so the user can edit in place
        initial = ""
        if self._limit_asd_freqs is not None:
            lines = [f"{f:.5g}  {a:.6g}"
                     for f, a in zip(self._limit_asd_freqs, self._limit_asd_vals)]
            initial = "\n".join(lines)
        text = self._paste_dialog(
            "Paste / Edit Response Limit ASD",
            "2-column data (freq  g²/Hz), one row per line:",
            initial=initial)
        if text is None:
            return
        freqs, asds, _ = parse_asd_text(text)
        if freqs is None:
            messagebox.showerror("Parse Error", "Need at least 2 valid frequency rows.")
            return
        self._limit_asd_freqs, self._limit_asd_vals = freqs, asds
        self._limit_status.configure(
            text=self._asd_status_text("(pasted)", freqs),
            text_color=("gray10", "gray90"))
        if self._drawer_visible:
            self._populate_drawer_tables()
        self._clear_results()
        self._schedule_recompute()

    def _clear_limit_asd(self):
        self._limit_asd_freqs = self._limit_asd_vals = None
        self._limit_status.configure(text="(none)", text_color="gray")
        self._clear_results()

    # ── Node grid management ──────────────────────────────────────────────

    def _add_node_to_grid(self, nid, label="", x="", y="", z=""):
        """Insert a row in the node sheet. x/y/z should be '', 'S', or 'L'."""
        nrows = len(self._node_sheet.get_sheet_data())
        self._node_sheet.insert_row(row=[str(nid), label,
                                        x.upper(), y.upper(), z.upper()])
        for ci, val in zip(_DOF_COLS, [x, y, z]):
            self._apply_cell_color(nrows, ci, val.upper())

    def _apply_cell_color(self, row, col, val):
        if val == "S":
            self._node_sheet.highlight_cells(
                row=row, column=col,
                bg=_COL_S["bg"], fg=_COL_S["fg"])
        elif val == "L":
            self._node_sheet.highlight_cells(
                row=row, column=col,
                bg=_COL_L["bg"], fg=_COL_L["fg"])
        else:
            self._node_sheet.dehighlight_cells(row=row, column=col)

    def _on_grid_edit(self, event):
        r, c = event.row, event.column
        if c not in _DOF_COLS:
            # Label or Node ID changed; just schedule recompute
            self._schedule_recompute()
            return
        raw = str(self._node_sheet.get_cell_data(r, c)).strip().upper()
        if raw not in _VALID:
            raw = ""
        self._node_sheet.set_cell_data(r, c, raw)
        self._apply_cell_color(r, c, raw)
        self._schedule_recompute()

    def _add_node_dialog(self):
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("Add Nodes")
        dlg.geometry("380x310")
        dlg.transient(self.frame.winfo_toplevel())
        dlg.grab_set()

        ctk.CTkLabel(
            dlg,
            text="Enter one node per line.  Optional label:\n"
                 "  1001        1001 Tip mass        1001, Tip mass",
            justify=tk.LEFT, anchor=tk.W,
        ).pack(padx=12, pady=(12, 4), fill=tk.X)

        tb = ctk.CTkTextbox(dlg, wrap="none")
        tb.pack(fill=tk.BOTH, expand=True, padx=12)

        def _ok():
            self._parse_and_add_nodes(tb.get("1.0", "end"))
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=12, pady=8)
        ctk.CTkButton(btn_row, text="Add", command=_ok).pack(side=tk.LEFT)
        ctk.CTkButton(btn_row, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=5)
        dlg.bind("<Return>", lambda _: _ok())

    def _parse_and_add_nodes(self, text):
        existing = set()
        for r in self._node_sheet.get_sheet_data():
            try:
                existing.add(int(str(r[0]).strip()))
            except (ValueError, TypeError):
                pass
        added = 0
        for raw in text.splitlines():
            line = raw.strip()
            if not line or line.startswith('#') or line.startswith('$'):
                continue
            parts = [p.strip() for p in line.split(',', 1)] if ',' in line \
                else line.split(None, 1)
            try:
                nid = int(parts[0])
            except (ValueError, IndexError):
                continue
            if nid in existing:
                continue
            existing.add(nid)
            label = parts[1].strip() if len(parts) > 1 and parts[1].strip() else ""
            self._add_node_to_grid(nid, label, x="L", y="L", z="L")
            added += 1
        if added:
            self._status_label.configure(text=f"Added {added} node(s).",
                                         text_color=("gray10", "gray90"))
            self._schedule_recompute()

    def _import_nodes_csv(self):
        path = filedialog.askopenfilename(
            title="Import Nodes",
            filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv"),
                       ("All files", "*.*")])
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.csv':
                text = self._read_csv_as_text(path)
            else:
                text = self._read_text_node_file(path)
        except Exception as exc:
            messagebox.showerror("Import Error", str(exc))
            return
        self._parse_and_add_nodes(text)

    @staticmethod
    def _read_csv_as_text(path):
        import csv as _csv
        with open(path, newline='', encoding='utf-8-sig') as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = _csv.Sniffer().sniff(sample)
                has_header = _csv.Sniffer().has_header(sample)
            except _csv.Error:
                dialect = _csv.excel
                has_header = False
            reader = _csv.reader(f, dialect)
            if has_header:
                next(reader, None)
            lines = []
            for row in reader:
                if not row:
                    continue
                try:
                    gid = int(str(row[0]).strip())
                except (ValueError, IndexError):
                    continue
                label = row[1].strip() if len(row) > 1 and str(row[1]).strip() else ""
                lines.append(f"{gid},{label}" if label else str(gid))
        return "\n".join(lines)

    @staticmethod
    def _read_text_node_file(path):
        lines = []
        with open(path, encoding='utf-8') as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith('#') or line.startswith('$'):
                    continue
                parts = [p.strip() for p in line.split(',', 1)] if ',' in line \
                    else line.split(None, 1)
                try:
                    gid = int(parts[0])
                except (ValueError, IndexError):
                    continue
                label = parts[1].strip() if len(parts) > 1 and parts[1].strip() else ""
                lines.append(f"{gid},{label}" if label else str(gid))
        return "\n".join(lines)

    def _remove_selected_nodes(self):
        selected = self._node_sheet.get_selected_rows()
        if not selected:
            return
        for r in sorted(selected, reverse=True):
            self._node_sheet.delete_rows([r])
        self._schedule_recompute()

    def _clear_all_nodes(self):
        n = len(self._node_sheet.get_sheet_data())
        if n:
            self._node_sheet.delete_rows(list(range(n)))
        self._clear_results()

    def _get_grid_dofs(self):
        """Return (plot_dofs, limit_dofs, color_map, label_map) from current grid."""
        data     = self._node_sheet.get_sheet_data()
        plot_dofs  = []   # [(nid, dof_idx), ...]
        limit_dofs = []
        color_map  = {}   # (nid, dof_idx) → hex color
        label_map  = {}   # (nid, dof_idx) → display label
        color_idx = 0
        for row in data:
            try:
                nid = int(str(row[0]).strip())
            except (ValueError, TypeError):
                continue
            user_label = str(row[1]).strip()
            for ci, dof_idx in zip(_DOF_COLS, range(3)):
                val = str(row[ci]).strip().upper()
                if val in ("S", "L"):
                    key = (nid, dof_idx)
                    plot_dofs.append(key)
                    color_map[key] = _NODE_COLORS[color_idx % len(_NODE_COLORS)]
                    color_idx += 1
                    base = user_label if user_label else str(nid)
                    label_map[key] = f"{base} {_DOF_NAMES[dof_idx]}"
                if val == "L":
                    limit_dofs.append((nid, dof_idx))
        return plot_dofs, limit_dofs, color_map, label_map

    # ── Paste dialog ──────────────────────────────────────────────────────

    def _paste_dialog(self, title, prompt, initial=""):
        result = {}
        win = ctk.CTkToplevel(self.frame.winfo_toplevel())
        win.title(title)
        win.geometry("460x340")
        win.resizable(True, True)
        win.transient(self.frame.winfo_toplevel())
        win.grab_set()

        ctk.CTkLabel(win, text=prompt, anchor=tk.W).pack(fill=tk.X, padx=10, pady=(10, 4))
        tb = ctk.CTkTextbox(win, wrap="none")
        tb.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        if initial:
            tb.insert("1.0", initial)

        btn_row = ctk.CTkFrame(win, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=10, pady=(4, 10))

        def _ok():
            result['text'] = tb.get("1.0", tk.END)
            win.destroy()

        ctk.CTkButton(btn_row, text="OK", width=80, command=_ok).pack(side=tk.LEFT, padx=4)
        ctk.CTkButton(btn_row, text="Cancel", width=80,
                      command=win.destroy).pack(side=tk.LEFT, padx=4)
        win.wait_window()
        return result.get('text')

    # ── Core computation ──────────────────────────────────────────────────

    def _compute_notch(self, silent=False):
        if self._op2 is None:
            if not silent:
                messagebox.showwarning("No OP2", "Load an FRF OP2 first.")
            return
        if self._active_input() is None:
            if not silent:
                messagebox.showwarning("No Input ASD", "Load an Input ASD first.")
            return
        if self._limit_asd_freqs is None:
            if not silent:
                messagebox.showwarning("No Limit ASD", "Load a Response Limit ASD first.")
            return

        plot_dofs, limit_dofs, color_map, label_map = self._get_grid_dofs()
        if not limit_dofs:
            if not silent:
                messagebox.showwarning("No Limit DOFs",
                                       "Set at least one X/Y/Z cell to L in the node grid.")
            return

        sc = self._get_subcase_int()
        if sc is None:
            return

        cfg = RESPONSE_TYPES['Acceleration']
        unit_factor = cfg['unit_factors'].get(self._units_var.get(), 386.089)
        id_attr = cfg['id_attr']

        frf_dict = getattr(self._op2, cfg['frf_attr'], None) or {}
        frf_tbl  = lookup_subcase(frf_dict, sc)
        if frf_tbl is None:
            return

        freqs = frf_tbl._times
        arr   = getattr(frf_tbl, id_attr)
        entity_ids = arr[:, 0] if id_attr == 'node_gridtype' else arr

        try:
            scale_db = float(self._scale_var.get())
        except ValueError:
            scale_db = 0.0
        scale = 10.0 ** (scale_db / 10.0)

        orig_interp  = self._effective_input(freqs, scale)
        limit_interp = interp_loglog(self._limit_asd_freqs, self._limit_asd_vals, freqs)

        min_allowed = np.full(len(freqs), np.inf)
        missing     = []
        limit_dofs_set = set()

        for nid, dof_idx in limit_dofs:
            hits = np.where(entity_ids == nid)[0]
            if not len(hits):
                missing.append(f"{nid} {_DOF_NAMES[dof_idx]}")
                continue
            H_g = frf_tbl.data[:, hits[0], dof_idx] / unit_factor
            H2  = H_g.real**2 + H_g.imag**2
            with np.errstate(divide='ignore', invalid='ignore'):
                min_allowed = np.minimum(
                    min_allowed, np.where(H2 > 0, limit_interp / H2, np.inf))
            limit_dofs_set.add((nid, dof_idx))

        notched = np.minimum(orig_interp, min_allowed)
        notched = np.where(np.isinf(notched), orig_interp, notched)

        if self._notch_enabled_var.get():
            try:
                db = float(self._notch_db_var.get())
            except ValueError:
                db = 6.0
            notched = np.maximum(notched, orig_interp * 10**(-db / 10.0))

        response_curves = {}
        for nid, dof_idx in plot_dofs:
            hits = np.where(entity_ids == nid)[0]
            if not len(hits):
                if f"{nid} {_DOF_NAMES[dof_idx]}" not in missing:
                    missing.append(f"{nid} {_DOF_NAMES[dof_idx]}")
                continue
            H_g = frf_tbl.data[:, hits[0], dof_idx] / unit_factor
            H2  = H_g.real**2 + H_g.imag**2
            response_curves[(nid, dof_idx)] = (H2 * orig_interp, H2 * notched)

        self._frf_freqs        = freqs
        self._orig_asd_interp  = orig_interp
        self._limit_asd_interp = limit_interp
        self._notched_asd      = notched
        self._response_curves  = response_curves
        self._limit_dofs_set   = limit_dofs_set
        self._color_map        = color_map
        self._label_map        = label_map

        # Update curve picker
        labels = []
        for key in response_curves:
            tag = " [L]" if key in limit_dofs_set else ""
            labels.append(label_map.get(key, f"Node {key[0]} {_DOF_NAMES[key[1]]}") + tag)
        self._curve_menu.configure(values=labels if labels else ["(none)"])
        if labels and (self._curve_var.get() == "(none)" or
                       self._curve_var.get() not in labels):
            self._curve_var.set(labels[0])

        self._export_excel_btn.configure(state=tk.NORMAL)
        self._export_csv_btn.configure(state=tk.NORMAL)

        parts = [f"Notched: {len(response_curves)} curve(s)."]
        if missing:
            parts.append(f"Not found: {', '.join(missing[:4])}")
        self._status_label.configure(text="  ".join(parts),
                                     text_color=("gray10", "gray90"))
        self._redraw()

    def _clear_results(self):
        self._frf_freqs        = None
        self._orig_asd_interp  = None
        self._limit_asd_interp = None
        self._notched_asd      = None
        self._response_curves  = {}
        self._limit_dofs_set   = set()
        self._color_map        = {}
        self._label_map        = {}
        self._picked_peaks.clear()
        for pane in self._panes:
            pane["last_drawn_curves"] = []
        self._export_excel_btn.configure(state=tk.DISABLED)
        self._export_csv_btn.configure(state=tk.DISABLED)
        self._curve_menu.configure(values=["(none)"])
        self._curve_var.set("(none)")

    # ── Plot helpers ──────────────────────────────────────────────────────

    def _get_theme(self):
        return _THEMES[self._plot_theme]

    def _format_plot(self, pane, title="", ylabel="ASD (g²/Hz)"):
        ax = pane["ax"]
        th = self._get_theme()
        pane["fig"].set_facecolor(th["fig_bg"])
        ax.set_facecolor(th["plot_bg"])
        ax.tick_params(colors=th["text"])
        ax.xaxis.label.set_color(th["text"])
        ax.yaxis.label.set_color(th["text"])
        ax.title.set_color(th["text"])
        for spine in ax.spines.values():
            spine.set_edgecolor(th["spine"])
        ax.set_xscale("log")
        ax.set_yscale(self._yscale_var.get().lower())
        ax.grid(True, which="both", color=th["grid"], linewidth=0.5)
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel(ylabel)
        user_title = self._title_var.get().strip()
        env = self._env_var.get().strip()
        display_title = user_title or title
        if display_title:
            ax.set_title(display_title, color=th["text"], fontsize=12, weight="bold",
                         pad=20 if env else 6)
        if env:
            ax.text(0.5, 1.01, env, transform=ax.transAxes, ha="center", va="bottom",
                    color=th["text"], fontsize=10, style="italic", alpha=0.75)

    def _legend(self, ax):
        th = self._get_theme()
        leg = ax.legend(facecolor=th["legend_bg"],
                        edgecolor=th["spine"],
                        labelcolor=th["text"],
                        fontsize=8)
        return leg

    def _draw_idle_plot(self):
        th = self._get_theme()
        for pane in self._panes:
            pane["grms_frame"].pack_forget()
            pane["canvas_widget"].pack(fill=tk.BOTH, expand=True)
            pane["fig"].clf()
            ax = pane["fig"].add_subplot(111)
            pane["ax"] = ax
            self._format_plot(pane, title="Response Limiting")
            ax.text(0.5, 0.5,
                    "1. Open FRF OP2\n2. Load Input ASD + Response Limit\n"
                    "3. Add nodes, set X/Y/Z to L\n"
                    "Results update automatically",
                    ha='center', va='center', transform=ax.transAxes,
                    color=th["text"], fontsize=11)
            pane["canvas"].draw_idle()

    # ── View dispatch ──────────────────────────────────────────────────────

    def _redraw(self):
        for pane in self._panes:
            self._redraw_pane(pane)

    def _redraw_pane(self, pane):
        if self._notched_asd is None:
            return
        view = pane["view_var"].get()
        if view == "grms":
            pane["canvas_widget"].pack_forget()
            pane["grms_frame"].pack(fill=tk.BOTH, expand=True)
            self._draw_grms_view(pane)
        else:
            pane["grms_frame"].pack_forget()
            pane["canvas_widget"].pack(fill=tk.BOTH, expand=True)
            pane["fig"].clf()
            pane["ax"] = pane["fig"].add_subplot(111)
            if view == "input":
                self._draw_input_view(pane)
            elif view == "response_dof":
                self._draw_response_dof_view(pane)
            elif view == "response_all":
                self._draw_response_all_view(pane)
            pane["canvas"].draw_idle()

    def _draw_input_view(self, pane):
        ax = pane["ax"]
        freqs = self._frf_freqs
        try:
            scale_db = float(self._scale_var.get())
        except ValueError:
            scale_db = 0.0
        scale = 10.0 ** (scale_db / 10.0)
        lbl_scale = f" ({scale_db:+.3g} dB)" if scale_db != 0.0 else ""
        pane["last_drawn_curves"] = []
        inp = self._active_input()
        if inp is not None:
            vals = inp['vals_raw'] * scale
            lbl = f"Original Input{lbl_scale}"
            ax.plot(inp['freqs_raw'], vals, color="#1f77b4", label=lbl)
            pane["last_drawn_curves"].append({"freqs": inp['freqs_raw'], "data": vals, "label": lbl})
        if self._workmanship_freqs is not None:
            ax.plot(self._workmanship_freqs, self._workmanship_vals,
                    color="#7f7f7f", linestyle="--", linewidth=1.2, label="Workmanship")
            pane["last_drawn_curves"].append({"freqs": self._workmanship_freqs,
                                              "data": self._workmanship_vals, "label": "Workmanship"})
        if (self._workmanship_envelope_var.get() and self._workmanship_freqs is not None
                and inp is not None):
            eff_lbl = "Effective Input (enveloped)"
            ax.plot(freqs, self._orig_asd_interp, color="#1f77b4", linewidth=2.5,
                    linestyle="-", alpha=0.85, label=eff_lbl)
            pane["last_drawn_curves"].append({"freqs": freqs, "data": self._orig_asd_interp,
                                              "label": eff_lbl})
        ax.plot(freqs, self._notched_asd, color="#d62728", label="Notched Input")
        pane["last_drawn_curves"].append({"freqs": freqs, "data": self._notched_asd,
                                          "label": "Notched Input"})
        og = np.sqrt(max(0.0, grms_loglog(freqs, self._orig_asd_interp)))
        ng = np.sqrt(max(0.0, grms_loglog(freqs, self._notched_asd)))
        self._draw_aux_lines(ax)
        self._draw_picked_peaks(ax)
        self._legend(ax).set_title(
            f"Orig:    {og:.3g} g GRMS\nNotched: {ng:.3g} g GRMS")
        self._format_plot(pane, title="Input ASD: Original vs Notched")
        ax.set_xlim(freqs[0], freqs[-1])

    def _draw_response_dof_view(self, pane):
        ax = pane["ax"]
        freqs = self._frf_freqs
        sel = self._curve_var.get()
        limit_name = self._limit_name_var.get().strip() or "Response Limit"

        selected_key = None
        for key in self._response_curves:
            tag = " [L]" if key in self._limit_dofs_set else ""
            lbl = self._label_map.get(key, f"Node {key[0]} {_DOF_NAMES[key[1]]}") + tag
            if lbl == sel:
                selected_key = key
                break
        if selected_key is None and self._response_curves:
            selected_key = next(iter(self._response_curves))

        pane["last_drawn_curves"] = []
        rb_grms = ra_grms = 0.0
        if selected_key is not None:
            rb, ra = self._response_curves[selected_key]
            col = self._color_map.get(selected_key, "#1f77b4")
            lbl_b = "Before notch"
            lbl_a = "After notch"
            ax.plot(freqs, rb, color=col, linestyle="--", label=lbl_b)
            ax.plot(freqs, ra, color=col, label=lbl_a)
            pane["last_drawn_curves"].append({"freqs": freqs, "data": rb, "label": lbl_b})
            pane["last_drawn_curves"].append({"freqs": freqs, "data": ra, "label": lbl_a})
            rb_grms = np.sqrt(max(0.0, grms_loglog(freqs, rb)))
            ra_grms = np.sqrt(max(0.0, grms_loglog(freqs, ra)))

        ax.plot(freqs, self._limit_asd_interp,
                color="#ff7f0e", linestyle=":", linewidth=2, label=limit_name)
        pane["last_drawn_curves"].append({"freqs": freqs, "data": self._limit_asd_interp,
                                          "label": limit_name})
        self._draw_aux_lines(ax)
        self._draw_picked_peaks(ax)
        leg = self._legend(ax)
        leg.set_title(f"Before: {rb_grms:.3g} g\nAfter:  {ra_grms:.3g} g")

        if selected_key:
            lbl = self._label_map.get(selected_key,
                                       f"Node {selected_key[0]} {_DOF_NAMES[selected_key[1]]}")
        else:
            lbl = "(none)"
        self._format_plot(pane, title=f"Response: {lbl}")
        ax.set_xlim(freqs[0], freqs[-1])

    def _draw_response_all_view(self, pane):
        ax = pane["ax"]
        freqs = self._frf_freqs
        limit_name = self._limit_name_var.get().strip() or "Response Limit"
        pane["last_drawn_curves"] = []
        for key, (_, ra) in self._response_curves.items():
            col = self._color_map.get(key, "#aaaaaa")
            tag = " [L]" if key in self._limit_dofs_set else ""
            lbl = self._label_map.get(key, f"Node {key[0]} {_DOF_NAMES[key[1]]}") + tag
            ax.plot(freqs, ra, color=col, label=lbl)
            pane["last_drawn_curves"].append({"freqs": freqs, "data": ra, "label": lbl})
        ax.plot(freqs, self._limit_asd_interp,
                color="#ff7f0e", linestyle=":", linewidth=2, label=limit_name)
        pane["last_drawn_curves"].append({"freqs": freqs, "data": self._limit_asd_interp,
                                          "label": limit_name})
        self._draw_aux_lines(ax)
        self._draw_picked_peaks(ax)
        self._legend(ax)
        self._format_plot(pane, title=f"All Responses (notched) vs {limit_name}  ([L] = limit DOF)")
        ax.set_xlim(freqs[0], freqs[-1])

    def _draw_grms_view(self, pane):
        freqs = self._frf_freqs
        og = np.sqrt(max(0.0, grms_loglog(freqs, self._orig_asd_interp)))
        ng = np.sqrt(max(0.0, grms_loglog(freqs, self._notched_asd)))
        with np.errstate(divide='ignore', invalid='ignore'):
            ratio = np.where(self._notched_asd > 0,
                             self._orig_asd_interp / self._notched_asd, 1.0)
        max_db = 10.0 * np.log10(max(float(ratio.max()), 1.0))
        rows = []
        for key, (rb, ra) in self._response_curves.items():
            tag = " [L]" if key in self._limit_dofs_set else ""
            lbl = self._label_map.get(key, f"Node {key[0]} {_DOF_NAMES[key[1]]}") + tag
            rows.append([
                lbl,
                f"{og:.4g}", f"{ng:.4g}",
                f"{np.sqrt(max(0.0, grms_loglog(freqs, rb))):.4g}",
                f"{np.sqrt(max(0.0, grms_loglog(freqs, ra))):.4g}",
                f"{max_db:.2f}",
            ])
        pane["grms_sheet"].set_sheet_data(rows)

    # ── Aux reference lines ───────────────────────────────────────────────

    def _draw_aux_lines(self, ax):
        for ln in self._aux_lines:
            if not ln.get('visible', True):
                continue
            try:
                val = float(ln['value'])
            except (KeyError, ValueError):
                continue
            kw = dict(color=ln.get('color', '#888888'),
                      linestyle=ln.get('linestyle', '--'),
                      linewidth=float(ln.get('linewidth', 1.0)),
                      alpha=0.8)
            if ln.get('axis', 'x') == 'x':
                ax.axvline(val, **kw)
            else:
                ax.axhline(val, **kw)
            lbl = ln.get('label', '').strip()
            if lbl:
                if ln.get('axis', 'x') == 'x':
                    ax.text(val, ax.get_ylim()[1], f" {lbl}", color=ln.get('color', '#888888'),
                            fontsize=7, va='top', ha='left', clip_on=True)
                else:
                    ax.text(ax.get_xlim()[0], val, f" {lbl}", color=ln.get('color', '#888888'),
                            fontsize=7, va='bottom', ha='left', clip_on=True)

    def _aux_lines_dialog(self):
        import copy
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("X/Y Reference Lines")
        dlg.geometry("660x380")
        dlg.transient(self.frame.winfo_toplevel())
        dlg.grab_set()

        working = copy.deepcopy(self._aux_lines)
        row_widgets = []

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill=tk.BOTH, expand=True, padx=8, pady=(8, 0))

        hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        hdr.pack(fill=tk.X, pady=(0, 2))
        for txt, w in [("Vis", 28), ("Axis", 50), ("Value", 80), ("Label", 170),
                       ("Color", 80), ("Style", 80), ("LW", 50), ("", 28)]:
            ctk.CTkLabel(hdr, text=txt, width=w, anchor=tk.W,
                         font=ctk.CTkFont(weight="bold")).pack(side=tk.LEFT, padx=2)

        def _make_row(ln):
            from tkinter import colorchooser as cc
            rf = ctk.CTkFrame(scroll, fg_color="transparent")
            rf.pack(fill=tk.X, pady=1)

            vis_var  = tk.BooleanVar(value=ln.get('visible', True))
            axis_var = ctk.StringVar(value=ln.get('axis', 'x').upper())
            val_var  = ctk.StringVar(value=str(ln.get('value', 0.0)))
            lbl_var  = ctk.StringVar(value=ln.get('label', ''))
            col_var  = ctk.StringVar(value=ln.get('color', '#888888'))
            ls_var   = ctk.StringVar(value=ln.get('linestyle', '--'))
            lw_var   = ctk.StringVar(value=str(ln.get('linewidth', 1.0)))

            ctk.CTkCheckBox(rf, text="", variable=vis_var, width=28).pack(side=tk.LEFT, padx=2)
            ctk.CTkOptionMenu(rf, variable=axis_var, values=["X", "Y"],
                              width=50).pack(side=tk.LEFT, padx=2)
            ctk.CTkEntry(rf, textvariable=val_var, width=80).pack(side=tk.LEFT, padx=2)
            ctk.CTkEntry(rf, textvariable=lbl_var, width=170).pack(side=tk.LEFT, padx=2)

            swatch = ctk.CTkButton(rf, text="", width=30, height=22,
                                   fg_color=ln.get('color', '#888888'),
                                   hover_color=ln.get('color', '#888888'),
                                   command=lambda: None)

            def _pick(cv=col_var, sw=swatch):
                result = cc.askcolor(color=cv.get(), title="Pick Color", parent=dlg)
                if result and result[1]:
                    cv.set(result[1])
                    sw.configure(fg_color=result[1], hover_color=result[1])

            swatch.configure(command=_pick)
            swatch.pack(side=tk.LEFT, padx=2)

            ctk.CTkOptionMenu(rf, variable=ls_var,
                              values=["-", "--", ":", "-."],
                              width=80).pack(side=tk.LEFT, padx=2)
            ctk.CTkEntry(rf, textvariable=lw_var, width=50).pack(side=tk.LEFT, padx=2)

            def _del(r=rf):
                r.destroy()
                for item in row_widgets:
                    if item['frame'] is r:
                        row_widgets.remove(item)
                        break

            ctk.CTkButton(rf, text="x", width=28,
                          fg_color="transparent", hover_color=("gray75", "gray30"),
                          text_color=("gray40", "gray60"),
                          command=_del).pack(side=tk.LEFT, padx=2)

            row_widgets.append({
                'frame': rf, 'vis': vis_var, 'axis': axis_var,
                'value': val_var, 'label': lbl_var, 'color': col_var,
                'linestyle': ls_var, 'linewidth': lw_var,
            })

        for ln in working:
            _make_row(ln)

        def _add():
            _make_row({"axis": "x", "value": 100.0, "label": "",
                       "color": "#888888", "linestyle": "--",
                       "linewidth": 1.0, "visible": True})

        add_row = ctk.CTkFrame(dlg, fg_color="transparent")
        add_row.pack(fill=tk.X, padx=8, pady=(4, 0))
        ctk.CTkButton(add_row, text="+ Add Line", width=100, command=_add).pack(side=tk.LEFT)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill=tk.X, padx=8, pady=(4, 10))

        def _ok():
            new_lines = []
            for w in row_widgets:
                try:
                    val = float(w['value'].get())
                except ValueError:
                    continue
                try:
                    lw = float(w['linewidth'].get())
                except ValueError:
                    lw = 1.0
                new_lines.append({
                    "axis": w['axis'].get().lower(),
                    "value": val,
                    "label": w['label'].get(),
                    "color": w['color'].get() or "#888888",
                    "linestyle": w['linestyle'].get(),
                    "linewidth": lw,
                    "visible": w['vis'].get(),
                })
            self._aux_lines = new_lines
            self._redraw()
            dlg.destroy()

        ctk.CTkButton(btn_row, text="OK", width=80, command=_ok).pack(side=tk.LEFT)
        ctk.CTkButton(btn_row, text="Cancel", width=80,
                      command=dlg.destroy).pack(side=tk.LEFT, padx=5)

    # ── Peak picking ──────────────────────────────────────────────────────

    def _draw_picked_peaks(self, ax):
        mode = self._peak_label_var.get()
        th = self._get_theme()
        for pk in self._picked_peaks:
            f, v = pk['freq'], pk['value']
            if f <= 0 or v <= 0:
                continue
            ax.axvline(f, color='gray', linestyle=':', linewidth=0.8, alpha=0.6)
            if mode == "Freq only":
                txt = f"{f:.1f} Hz"
            else:
                txt = f"{f:.1f} Hz\n{v:.3g}"
            ax.annotate(txt, xy=(f, v), xytext=(4, 4), textcoords='offset points',
                        fontsize=7, color=th["text"], clip_on=True)

    def _toggle_pick_mode(self):
        self._pick_peaks_mode = not self._pick_peaks_mode
        cursor = "cross" if self._pick_peaks_mode else ""
        if self._pick_peaks_mode:
            self._pick_btn.configure(fg_color=("#1f6aa5", "#1f6aa5"), text="Pick Peaks (ON)")
        else:
            self._pick_btn.configure(fg_color=("gray75", "gray25"), text="Pick Peaks")
        for pane in self._panes:
            pane["canvas_widget"].configure(cursor=cursor)

    def _clear_peaks(self):
        self._picked_peaks.clear()
        self._redraw()

    def _on_canvas_click(self, event, pane):
        if not self._pick_peaks_mode:
            return
        if event.inaxes is None or event.xdata is None or event.ydata is None:
            return
        if event.xdata <= 0:
            return

        from scipy.signal import find_peaks as _find_peaks

        click_lx = np.log10(max(event.xdata, 1e-12))

        best = None
        for curve in pane["last_drawn_curves"]:
            freqs = curve["freqs"]
            data = curve["data"]
            if len(data) < 3:
                continue
            peak_idx, _ = _find_peaks(data)
            for i in peak_idx:
                f, v = float(freqs[i]), float(data[i])
                if f <= 0 or v <= 0:
                    continue
                d = abs(np.log10(f) - click_lx)
                if best is None or d < best[0]:
                    best = (d, f, v, curve["label"])

        if best is None or best[0] > 0.3:
            return
        self._picked_peaks.append({
            "freq": best[1], "value": best[2], "label": best[3],
        })
        self._redraw()

    # ── Session save / load ───────────────────────────────────────────────

    def _suggested_export_name(self):
        stem = (self._title_var.get().strip()
                or self._env_var.get().strip()
                or "Response_Limiting")
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in stem)
        return f"{safe.strip()}.xlsx"

    def _save_session(self):
        import json, datetime
        if self._op2_path is None and not self._inputs:
            messagebox.showwarning("Nothing to save",
                                   "Load an OP2 or ASD data before saving a session.")
            return
        stem = (self._title_var.get().strip()
                or self._env_var.get().strip()
                or "rl_session")
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in stem)
        path = filedialog.asksaveasfilename(
            title="Save Session",
            defaultextension=".rllimit.json",
            filetypes=[("Response Limiting session", "*.rllimit.json"),
                       ("All files", "*.*")],
            initialfile=f"{safe.strip()}.rllimit.json",
        )
        if not path:
            return

        # Serialize node grid
        nodes_data = []
        for row in self._node_sheet.get_sheet_data():
            try:
                nid = int(str(row[0]).strip())
            except (ValueError, TypeError):
                continue
            nodes_data.append({
                "id": nid,
                "label": str(row[1]).strip() if len(row) > 1 else "",
                "x": str(row[2]).strip().upper() if len(row) > 2 else "",
                "y": str(row[3]).strip().upper() if len(row) > 3 else "",
                "z": str(row[4]).strip().upper() if len(row) > 4 else "",
            })

        data = {
            "version": 2,
            "tool": "Response Limiting",
            "saved_at": datetime.datetime.now().isoformat(timespec='seconds'),
            "op2_path": self._op2_path,
            "units": self._units_var.get(),
            "subcase": self._subcase_var.get(),
            "inputs": [
                {"name": e["name"], "freqs": e["freqs_raw"].tolist(),
                 "vals": e["vals_raw"].tolist(), "db": e.get("db", "0.0")}
                for e in self._inputs
            ],
            "active_input_idx": self._active_input_idx,
            "limit_asd": ({"freqs": self._limit_asd_freqs.tolist(),
                           "vals":  self._limit_asd_vals.tolist()}
                          if self._limit_asd_freqs is not None else None),
            "workmanship": ({"freqs": self._workmanship_freqs.tolist(),
                             "vals":  self._workmanship_vals.tolist()}
                            if self._workmanship_freqs is not None else None),
            "workmanship_envelope": self._workmanship_envelope_var.get(),
            "nodes": nodes_data,
            "notch_enabled": self._notch_enabled_var.get(),
            "notch_db": self._notch_db_var.get(),
            "scale_db": self._scale_var.get(),
            "aux_lines": list(self._aux_lines),
            "view": {
                "view": self._view_var.get(),
                "yscale": self._yscale_var.get(),
                "title": self._title_var.get(),
                "env": self._env_var.get(),
                "limit_name": self._limit_name_var.get(),
                "peak_label": self._peak_label_var.get(),
            },
            "pane_views": [p["view_var"].get() for p in self._panes],
            "split_active": len(self._panes) > 1,
            "picked_peaks": [
                {"freq": pk["freq"], "value": pk["value"], "label": pk.get("label", "")}
                for pk in self._picked_peaks
            ],
        }
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        except Exception as exc:
            messagebox.showerror("Save Error", str(exc))
            return
        self._show_export_done_dialog(path)

    def _open_session(self):
        import json
        path = filedialog.askopenfilename(
            title="Open Session",
            filetypes=[("Response Limiting session", "*.rllimit.json"),
                       ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
        except Exception as exc:
            messagebox.showerror("Load Error", f"Could not read session:\n{exc}")
            return
        if data.get('tool') != "Response Limiting" or data.get('version') not in (1, 2):
            messagebox.showerror("Unsupported File",
                                 "This file does not appear to be a Response Limiting session.")
            return

        view = data.get('view', {})
        self._view_var.set(view.get('view', 'input'))
        self._yscale_var.set(view.get('yscale', 'Log'))
        self._title_var.set(view.get('title', ''))
        self._env_var.set(view.get('env', ''))
        self._limit_name_var.set(view.get('limit_name', 'Response Limit'))
        self._peak_label_var.set(view.get('peak_label', 'Freq only'))

        self._aux_lines = data.get('aux_lines', [])
        self._picked_peaks = [
            {"freq": pk["freq"], "value": pk["value"], "label": pk.get("label", "")}
            for pk in data.get('picked_peaks', [])
        ]

        self._units_var.set(data.get('units', 'in/s²'))
        self._notch_enabled_var.set(data.get('notch_enabled', False))
        self._notch_db_var.set(data.get('notch_db', '6.0'))
        self._notch_entry.configure(
            state=tk.NORMAL if self._notch_enabled_var.get() else tk.DISABLED)
        self._scale_var.set(data.get('scale_db', '0.0'))

        # Load inputs — support v1 (single input_asd) and v2 (inputs list)
        self._inputs = []
        if data.get('version') == 1:
            inp = data.get('input_asd')
            if inp:
                freqs = np.array(inp['freqs'])
                vals  = np.array(inp['vals'])
                self._inputs = [{"name": self._asd_status_text("(session)", freqs),
                                  "freqs_raw": freqs, "vals_raw": vals,
                                  "db": data.get('scale_db', '0.0')}]
                self._active_input_idx = 0
        else:
            for e in data.get('inputs', []):
                freqs = np.array(e['freqs'])
                vals  = np.array(e['vals'])
                self._inputs.append({"name": e.get('name', self._asd_status_text("(session)", freqs)),
                                     "freqs_raw": freqs, "vals_raw": vals,
                                     "db": e.get('db', '0.0')})
            self._active_input_idx = data.get('active_input_idx', 0 if self._inputs else -1)
        self._refresh_input_selector()

        wm = data.get('workmanship')
        if wm:
            self._workmanship_freqs = np.array(wm['freqs'])
            self._workmanship_vals  = np.array(wm['vals'])
            self._workmanship_status.configure(
                text=self._asd_status_text("(session)", self._workmanship_freqs),
                text_color=("gray10", "gray90"))
        self._workmanship_envelope_var.set(data.get('workmanship_envelope', False))

        lim = data.get('limit_asd')
        if lim:
            self._limit_asd_freqs = np.array(lim['freqs'])
            self._limit_asd_vals  = np.array(lim['vals'])
            self._limit_status.configure(
                text=self._asd_status_text("(session)", self._limit_asd_freqs),
                text_color=("gray10", "gray90"))

        # Restore node grid
        n = len(self._node_sheet.get_sheet_data())
        if n:
            self._node_sheet.delete_rows(list(range(n)))
        for nd in data.get('nodes', []):
            self._add_node_to_grid(nd['id'], nd.get('label', ''),
                                   nd.get('x', ''), nd.get('y', ''), nd.get('z', ''))

        if self._drawer_visible:
            self._populate_drawer_tables()

        # Restore split view and per-pane view modes
        pane_views = data.get('pane_views', [self._view_var.get()])
        split_active = data.get('split_active', False)
        if split_active and len(self._panes) == 1:
            self._toggle_split()
        elif not split_active and len(self._panes) > 1:
            self._toggle_split()
        for i, pane in enumerate(self._panes):
            if i < len(pane_views):
                pane["view_var"].set(pane_views[i])

        op2_path = data.get('op2_path')
        if op2_path and os.path.isfile(op2_path):
            self._status_label.configure(text=f"Reloading OP2 from session…", text_color="gray")

            def _work():
                from pyNastran.op2.op2 import OP2
                op2 = OP2(mode='nx', debug=False)
                op2.read_op2(op2_path)
                return op2

            _saved_sc = data.get('subcase', '')

            def _done(op2, error):
                if error is not None:
                    self._status_label.configure(
                        text=f"OP2 reload failed: {error}", text_color="orange")
                    return
                cfg = RESPONSE_TYPES['Acceleration']
                frf_dict = getattr(op2, cfg['frf_attr'], None) or {}
                if not frf_dict:
                    self._status_label.configure(text="No FRF data in OP2", text_color="orange")
                    return
                self._op2 = op2
                self._op2_path = op2_path
                sc_pairs = subcase_options(frf_dict)
                self._subcase_opts = sc_pairs
                labels = [lbl for _, lbl in sc_pairs]
                self._sc_menu.configure(values=labels)
                sc_to_set = _saved_sc if _saved_sc in labels else (labels[0] if labels else "(none)")
                self._subcase_var.set(sc_to_set)
                stem = os.path.splitext(os.path.basename(op2_path))[0]
                self._file_label.configure(text=stem, text_color=("gray10", "gray90"))
                self._schedule_recompute()

            self._run_in_background("Reloading OP2…", _work, _done)
        else:
            if op2_path:
                self._status_label.configure(
                    text=f"Session loaded (OP2 not found: {os.path.basename(op2_path)})",
                    text_color="orange")
            self._schedule_recompute()

    # ── Export done dialog ────────────────────────────────────────────────

    def _show_export_done_dialog(self, path):
        import subprocess
        dlg = ctk.CTkToplevel(self.frame.winfo_toplevel())
        dlg.title("Export complete")
        dlg.geometry("340x130")
        dlg.resizable(False, False)
        dlg.transient(self.frame.winfo_toplevel())

        ctk.CTkLabel(dlg, text=f"Saved: {os.path.basename(path)}",
                     anchor=tk.W).pack(padx=14, pady=(14, 8), fill=tk.X)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(padx=14, pady=(0, 14), fill=tk.X)

        def _open_file():
            try:
                if os.name == 'nt':
                    os.startfile(path)
                elif os.uname().sysname == 'Darwin':
                    subprocess.run(["open", path], check=False)
                else:
                    subprocess.run(["xdg-open", path], check=False)
            except Exception:
                pass
            dlg.destroy()

        def _open_folder():
            folder = os.path.dirname(path)
            try:
                if os.name == 'nt':
                    os.startfile(folder)
                elif os.uname().sysname == 'Darwin':
                    subprocess.run(["open", folder], check=False)
                else:
                    subprocess.run(["xdg-open", folder], check=False)
            except Exception:
                pass
            dlg.destroy()

        ctk.CTkButton(btn_row, text="Open File", command=_open_file,
                      width=90).pack(side=tk.LEFT, padx=(0, 6))
        ctk.CTkButton(btn_row, text="Open Folder", command=_open_folder,
                      width=100).pack(side=tk.LEFT, padx=(0, 6))
        ctk.CTkButton(btn_row, text="Close", command=dlg.destroy,
                      width=70).pack(side=tk.LEFT)

    # ── Export ────────────────────────────────────────────────────────────

    def _export_excel(self):
        if self._notched_asd is None:
            messagebox.showwarning("No Data", "Compute notch first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save Excel", defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
            initialfile=self._suggested_export_name())
        if not path:
            return
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Missing Library", "openpyxl is required for Excel export.")
            return

        freqs = self._frf_freqs
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Notched ASD"
        ws.append(["Freq (Hz)", "Original ASD (g²/Hz)", "Notched ASD (g²/Hz)"])
        for f, o, n in zip(freqs, self._orig_asd_interp, self._notched_asd):
            ws.append([float(f), float(o), float(n)])

        ws2 = wb.create_sheet("Limit ASD")
        ws2.append(["Freq (Hz)", "Limit ASD (g²/Hz)"])
        for f, lim in zip(freqs, self._limit_asd_interp):
            ws2.append([float(f), float(lim)])

        for key, (rb, ra) in self._response_curves.items():
            lbl = self._label_map.get(key, f"N{key[0]}_{_DOF_NAMES[key[1]]}")
            sname = lbl[:31]
            ws_d = wb.create_sheet(sname)
            ws_d.append(["Freq (Hz)", "Resp Before (g²/Hz)", "Resp After (g²/Hz)"])
            for f, r_b, r_a in zip(freqs, rb, ra):
                ws_d.append([float(f), float(r_b), float(r_a)])

        ws_g = wb.create_sheet("GRMS Summary")
        ws_g.append(["DOF", "Orig Input GRMS (g)", "Notched Input GRMS (g)",
                      "Resp Before (g)", "Resp After (g)", "Max Notch (dB)"])
        og = np.sqrt(max(0.0, grms_loglog(freqs, self._orig_asd_interp)))
        ng = np.sqrt(max(0.0, grms_loglog(freqs, self._notched_asd)))
        with np.errstate(divide='ignore', invalid='ignore'):
            ratio = np.where(self._notched_asd > 0,
                             self._orig_asd_interp / self._notched_asd, 1.0)
        max_db = 10.0 * np.log10(max(float(ratio.max()), 1.0))
        for key, (rb, ra) in self._response_curves.items():
            tag = " [L]" if key in self._limit_dofs_set else ""
            lbl = self._label_map.get(key, f"Node {key[0]} {_DOF_NAMES[key[1]]}") + tag
            ws_g.append([lbl, round(og, 6), round(ng, 6),
                          round(np.sqrt(max(0.0, grms_loglog(freqs, rb))), 6),
                          round(np.sqrt(max(0.0, grms_loglog(freqs, ra))), 6),
                          round(max_db, 4)])

        wb.save(path)
        self._show_export_done_dialog(path)

    def _export_csv(self):
        if self._notched_asd is None:
            messagebox.showwarning("No Data", "Compute notch first.")
            return
        stem = "".join(c if c.isalnum() or c in " _-" else "_"
                       for c in (self._title_var.get().strip()
                                 or self._env_var.get().strip()
                                 or "Response_Limiting"))
        path = filedialog.asksaveasfilename(
            title="Save CSV", defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
            initialfile=f"{stem.strip()}.csv")
        if not path:
            return
        import csv as _csv
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = _csv.writer(f)
            writer.writerow(["Freq (Hz)", "Notched ASD (g^2/Hz)"])
            for freq, asd in zip(self._frf_freqs, self._notched_asd):
                writer.writerow([float(freq), float(asd)])
        self._show_export_done_dialog(path)
